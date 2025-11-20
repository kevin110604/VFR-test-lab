from flask import Flask, request, render_template, session, redirect, url_for, jsonify, flash, send_from_directory, Response, stream_with_context, abort, template_rendered, send_file
from config import SECRET_KEY, local_main, UPLOAD_FOLDER, TEST_GROUPS, local_complete, SO_GIO_TEST, TEAMS_WEBHOOK_URL_TRF, TEAMS_WEBHOOK_URL_RATE, TEMPLATE_MAP
from excel_utils import get_item_code, get_col_idx, copy_row_with_style, write_tfr_to_excel, append_row_to_trf, export_expired_samples_to_excel

from image_utils import allowed_file, get_img_urls
from auth import login, get_user_type
from test_logic import load_group_notes, get_group_test_status, is_drop_test, is_impact_test, is_rotational_test,  TEST_GROUP_TITLES, TEST_TYPE_VI, DROP_ZONES, DROP_LABELS, GT68_FACE_LABELS, GT68_FACE_ZONES
from test_logic import IMPACT_ZONES, IMPACT_LABELS, ROT_LABELS, ROT_ZONES, RH_IMPACT_ZONES, RH_VIB_ZONES, RH_SECOND_IMPACT_ZONES, RH_STEP12_ZONES, TWO_C_NP_STEP5_ZONES, update_group_note_file, get_group_note_value, F2057_TEST_TITLES
from notify_utils import send_teams_message
from counter_utils import update_counter, check_and_reset_counter, log_report_complete
from docx_utils import approve_request_fill_docx_pdf, fill_cover_from_excel_generic, try_convert_to_pdf
from file_utils import (
    safe_write_json, safe_read_json, safe_save_excel, safe_load_excel,
    safe_write_text, safe_read_text, safe_append_backup_json   # <‚Äî th√™m h√†m n√†y
)
import re, os, pytz, json, openpyxl, random, subprocess, regex, traceback, calendar, time, tempfile, uuid, secrets, copy, glob, zipfile, io
from contextlib import contextmanager
from datetime import datetime, timedelta
from waitress import serve
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from collections import defaultdict, OrderedDict
from threading import Lock
from contextlib import contextmanager
from vfr3 import vfr3_bp
from werkzeug.utils import secure_filename
from qr_print import qr_bp
from testlab_dashboard import dashboard_bp
from flask_session import Session
from collections import OrderedDict
from markupsafe import Markup
from dateutil.relativedelta import relativedelta
from datetime import datetime, date
from uuid import uuid4
from io import BytesIO

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.register_blueprint(vfr3_bp)
app.register_blueprint(qr_bp)
app.register_blueprint(dashboard_bp, url_prefix="/testlab")

# ==== Server-side session config ====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config.update(
    SESSION_TYPE='filesystem',  # c√≥ th·ªÉ ƒë·ªïi sang 'redis' / 'sqlalchemy' n·∫øu mu·ªën
    SESSION_FILE_DIR=os.path.join(BASE_DIR, '.flask_session'),
    SESSION_PERMANENT=False,  # cookie phi√™n (ƒë√≥ng tr√¨nh duy·ªát l√† h·∫øt)
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    SESSION_COOKIE_NAME='vfr_session',
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',  # ƒë·ªïi 'Strict' n·∫øu kiosk
    # SESSION_COOKIE_SECURE=True,   # b·∫≠t n·∫øu ch·∫°y HTTPS
)
Session(app)
# Nh·ªØng test d√πng giao di·ªán Hot & Cold
HOTCOLD_LIKE = set(["hot_cold", "standing_water", "stain", "corrosion"])
INDOOR_GROUPS = {"indoor_chuyen", "indoor_thuong", "indoor_stone", "indoor_metal","outdoor_finishing"}
REPORT_NO_LOCK = Lock()
BLANK_TOKENS = {"", "-", "‚Äî"}
BOXES_FILE = os.path.join(os.path.dirname(__file__), "boxes.json")

_LAST_TEST_LRU_LIMIT = 50  # l∆∞u t·ªëi ƒëa 50 report g·∫ßn nh·∫•t

def read_boxes():
    data = safe_read_json(BOXES_FILE)
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and isinstance(data.get("boxes"), list):
        return data["boxes"]
    return ["B1-S1","B1-S2","B2-S1","B2-S2","B3-S1"]

def write_boxes(boxes):
    safe_write_json(BOXES_FILE, {"boxes": boxes})

def _set_limited_mapping(session_key: str, subkey: str, value, limit: int = _LAST_TEST_LRU_LIMIT):
    """
    L∆∞u value v√†o session[session_key][subkey] v·ªõi gi·ªõi h·∫°n s·ªë ph·∫ßn t·ª≠.
    D√πng OrderedDict nh∆∞ LRU: n·∫øu qu√° 'limit' th√¨ pop ph·∫ßn t·ª≠ c≈© nh·∫•t.
    """
    mapping = session.get(session_key)
    if not isinstance(mapping, dict):
        mapping = {}

    od = OrderedDict(mapping)
    if subkey in od:
        od.move_to_end(subkey)
    od[subkey] = value

    while len(od) > limit:
        od.popitem(last=False)

    session[session_key] = dict(od)

def set_last_test(report: str, group_code: str, display_title: str = None):
    """
    Ghi ƒë·ªìng th·ªùi 'last_test_code' (group) v√† 'last_test_type' (display title).
    N·∫øu kh√¥ng truy·ªÅn display_title s·∫Ω d√πng t√™n nh√≥m (VI) qua get_group_title.
    """
    if display_title is None:
        display_title = get_group_title(group_code) or group_code.upper()
    _set_limited_mapping("last_test_code", str(report), group_code)
    _set_limited_mapping("last_test_type", str(report), display_title)

def set_last_test_type(report: str, display_title: str):
    """Ch·ªâ c·∫≠p nh·∫≠t 'last_test_type' (kh√¥ng ƒë·ª•ng t·ªõi last_test_code)."""
    _set_limited_mapping("last_test_type", str(report), display_title)

def get_last_test_code(report: str, default=None):
    return (session.get("last_test_code") or {}).get(str(report), default)

def get_last_test_type(report: str, default=None):
    return (session.get("last_test_type") or {}).get(str(report), default)

def _is_blank_cell(v):
    if v is None:
        return True
    if isinstance(v, str):
        s = (v.replace("\u00A0","").replace("\u200B","")
               .replace("\r","").replace("\n","").replace("\t","").strip())
        return s in BLANK_TOKENS or s == ""
    return False

def row_is_filled_for_report(excel_path, report_no):
    """True n·∫øu d√≤ng c√≥ B == report_no ƒê√É c√≥ d·ªØ li·ªáu ·ªü b·∫•t k·ª≥ c·ªôt C..X; False n·∫øu v·∫´n tr·ªëng."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    target_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value  # c·ªôt B
        if (str(v).strip() if v is not None else "") == str(report_no).strip():
            target_row = r
            break
    if target_row is None:
        wb.close()
        # Kh√¥ng th·∫•y m√£ trong c·ªôt B (kh√°c thi·∫øt k·∫ø) -> coi nh∆∞ ƒë√£ d√πng ƒë·ªÉ tr√°nh ghi b·∫≠y
        return True
    for c in range(3, 25):  # C..X
        if not _is_blank_cell(ws.cell(row=target_row, column=c).value):
            wb.close()
            return True   # ƒê√É c√≥ d·ªØ li·ªáu
    wb.close()
    return False          # C..X ƒë·ªÅu tr·ªëng => CH∆ØA d√πng

def format_excel_date_short(dt):
    """Convert Python datetime/date -> format 'd-mmm' (e.g., 7-Aug) cho Excel."""
    if isinstance(dt, str):
        # Th·ª≠ parse v·ªÅ date
        try:
            dt = datetime.strptime(dt, "%Y-%m-%d")
        except:
            try:
                dt = datetime.strptime(dt, "%d/%m/%Y")
            except:
                try:
                    dt = datetime.strptime(dt, "%m/%d/%Y")
                except:
                    return dt  # Tr·∫£ nguy√™n n·∫øu kh√¥ng parse ƒë∆∞·ª£c
    # Tr·∫£ v·ªÅ d·∫°ng 'd-mmm'
    return f"{dt.day}-{calendar.month_abbr[dt.month]}"

def try_parse_excel_date(dt):
    """Parse dt v·ªÅ ki·ªÉu datetime/date n·∫øu c√≥ th·ªÉ, tr·∫£ v·ªÅ None n·∫øu kh√¥ng h·ª£p l·ªá."""
    if isinstance(dt, datetime):
        return dt
    if isinstance(dt, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(dt, fmt)
            except Exception:
                continue
    return None

def calculate_default_etd(request_date: str, test_group: str, *, all_reqs=None) -> str:
    if not request_date:
        return ""

    # Chu·∫©n ho√° group
    g = _group_of(test_group)
    if g in ("CONSTRUCTION", "TRANSIT"):
        base = 2   # 3 ng√†y t√≠nh c·∫£ ng√†y request => +2
    elif g in ("FINISHING", "MATERIAL"):
        base = 4   # 5 ng√†y t√≠nh c·∫£ ng√†y request => +4
    else:
        base = 2

    # --- Chu·∫©n b·ªã d·ªØ li·ªáu Pending + Archive ---
    # Pending: ch·ªâ l·∫•y Submitted
    pending_submitted = []
    if isinstance(all_reqs, list):
        for r in all_reqs:
            try:
                if (r.get("status") or "").strip() == "Submitted":
                    pending_submitted.append(r)
            except Exception:
                continue

    # Archive: d√πng tr·ª±c ti·∫øp test_group n·∫øu c√≥, fallback map t·ª´ report_no
    try:
        archive_list = safe_read_json(ARCHIVE_LOG) or []
    except Exception:
        archive_list = []

    rep2grp = _build_reportno_to_group_map()

    archive_mapped = []
    for a in archive_list:
        try:
            req_date = (a.get("request_date") or "").strip()
            trq     = (a.get("trq_id") or "").strip()
            rep_no  = (a.get("report_no") or "").strip()
            grp0    = (a.get("test_group") or "").strip()  # ∆∞u ti√™n group ƒë√£ l∆∞u
            grp     = grp0 if grp0 else rep2grp.get(rep_no, "")
            if not req_date or not grp:
                continue
            archive_mapped.append({
                "request_date": req_date,
                "test_group": grp,       # ƒë·ªÉ _group_of d√πng
                "trq_id": trq or rep_no, # fallback report_no n·∫øu thi·∫øu trq_id
                "status": "Approved",
            })
        except Exception:
            continue

    # --- ƒê·∫øm TRQ duy nh·∫•t cho (request_date, group) ---
    target_date = (request_date or "").strip()
    target_grp  = g
    uniq_trq = set()

    # Pending Submitted
    for r in pending_submitted:
        try:
            r_date = (r.get("request_date") or "").strip()
            r_grp  = _group_of(r.get("test_group") or r.get("type_of_test"))
            if r_date == target_date and r_grp == target_grp:
                tid = (r.get("trq_id") or "").strip()
                if not tid:
                    # ƒë·ªÉ kh√¥ng crash, coi 1 d√≤ng l√† 1 "TRQ"
                    tid = f"__row_{id(r)}"
                uniq_trq.add(tid)
        except Exception:
            continue

    # Archive Approved (ƒë√£ map group & trq)
    for a in archive_mapped:
        try:
            if a["request_date"] == target_date and _group_of(a["test_group"]) == target_grp:
                tid = (a.get("trq_id") or "").strip()
                if not tid:
                    tid = a.get("report_no", "")
                if tid:
                    uniq_trq.add(tid)
        except Exception:
            continue

    cnt = len(uniq_trq)  # s·ªë TRQ duy nh·∫•t ƒë√£ c√≥ TR∆Ø·ªöC request m·ªõi

    # --- Extra theo ng∆∞·ª°ng t·ª´ng nh√≥m ---
    extra = 0
    if g in ("CONSTRUCTION", "TRANSIT"):
        # ƒëang l√† # (cnt + 1); n·∫øu ƒë√£ c√≥ ‚â•5 th√¨ request m·ªõi r∆°i v√†o #6..#10
        if cnt >= 10:      # ƒëang l√† #11..#15
            extra = 2
        elif cnt >= 5:     # ƒëang l√† #6..#10
            extra = 1
    elif g in ("FINISHING", "MATERIAL"):
        if cnt >= 30:      # ƒëang l√† #31..#45
            extra = 4
        elif cnt >= 15:    # ƒëang l√† #16..#30
            extra = 2

    try:
        d0 = datetime.strptime(request_date, "%Y-%m-%d").date()
    except Exception:
        try:
            d0 = datetime.strptime(request_date, "%d/%m/%Y").date()
        except Exception:
            return ""

    etd = d0 + timedelta(days=base + extra)
    return etd.strftime("%Y-%m-%d")

# ---- c√°c h√†m helper kh√¥ng ƒë·ªïi (gi·ªØ nguy√™n) ----
def get_group_title(group):
    for g_id, g_name in TEST_GROUPS:
        if g_id == group:
            return g_name
    return None

def generate_unique_trq_id(existing_ids):
    yy = str(datetime.now().year)[-2:]  # 2 s·ªë cu·ªëi c·ªßa nƒÉm hi·ªán t·∫°i
    while True:
        num = random.randint(10000, 99999)
        new_id = f"TL-{yy}{num}"
        if new_id not in existing_ids:
            return new_id

ARCHIVE_LOG = "tfr_archive.json"
TFR_LOG_FILE = "tfr_requests.json"

@contextmanager
def report_lock():
    lock_dir = tempfile.gettempdir()              # Windows: C:\Users\<user>\AppData\Local\Temp
    lock_path = os.path.join(lock_dir, "tfr_report.lock")
    # Optional: timeout ƒë·ªÉ kh√¥ng ch·ªù v√¥ h·∫°n
    timeout_s = 60
    t0 = time.time()
    fd = None
    while True:
        try:
            # t·∫°o m·ªõi, n·∫øu ƒë√£ c√≥ -> FileExistsError
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            # ghi ch√∫t info ƒë·ªÉ debug stale lock
            os.write(fd, f"pid={os.getpid()} run={uuid.uuid4()}".encode("utf-8"))
            break
        except FileExistsError:
            # lock l√¢u qu√° coi nh∆∞ stale => c·ªë g·∫Øng xo√°
            if time.time() - t0 > timeout_s:
                _try_unlink_with_retry(lock_path)
            else:
                time.sleep(0.05 + random.random() * 0.15)
    try:
        yield
    finally:
        try:
            if fd is not None:
                os.close(fd)
        except Exception:
            pass
        _try_unlink_with_retry(lock_path)

def _try_unlink_with_retry(path, retries=8, delay=0.08):
    # Windows c√≥ th·ªÉ v∆∞·ªõng PermissionError do AV; retry ng·∫Øn s·∫Ω qua ƒë∆∞·ª£c
    for i in range(retries):
        try:
            os.unlink(path)
            return True
        except FileNotFoundError:
            return True
        except PermissionError:
            time.sleep(delay * (1.5 ** i))  # backoff
        except Exception:
            time.sleep(delay)
    # fallback: ƒë·ªïi t√™n ƒë·ªÉ kh√¥ng c·∫£n tr·ªü l·∫ßn sau
    try:
        os.rename(path, path + ".stale")
    except Exception:
        pass
    return False

def bump_report_no(s):
    m = re.search(r'(\d+)$', str(s))
    if not m:
        return f"{s}-1"
    start, end = m.span(1)
    n = int(m.group(1)) + 1
    width = end - start
    return f"{s[:start]}{str(n).zfill(width)}"

def report_no_exists(report_no, tfr_requests):
    """
    ƒê√É D√ôNG khi:
    - D√≤ng B==report_no trong Excel c√≥ d·ªØ li·ªáu C..X (kh√¥ng c√≤n tr·ªëng), HO·∫∂C
    - File ƒë·∫ßu ra cho m√£ ƒë√≥ ƒë√£ t·ªìn t·∫°i (pdf/docx), HO·∫∂C
    - M√£ n√†y ƒë√£ n·∫±m trong archive/log (ƒë√£ approve).
    """
    # 1) Excel: d√≤ng ƒë√£ c√≥ d·ªØ li·ªáu?
    try:
        if row_is_filled_for_report(local_main, report_no):
            return True
    except Exception:
        pass

    # 2) Tr√πng file ƒë√£ sinh?
    output_folder = os.path.join('static', 'TFR')
    if os.path.exists(os.path.join(output_folder, f"{report_no}.pdf")):
        return True
    if os.path.exists(os.path.join(output_folder, f"{report_no}.docx")):
        return True

    # 3) Tr√πng trong log pending ƒëang d√πng?
    for r in tfr_requests:
        if str(r.get("report_no") or "").strip() == str(report_no):
            return True

    # 4) Tr√πng trong archive (ƒë√£ approve)?
    try:
        archive = safe_read_json(ARCHIVE_LOG)
        for r in archive:
            if str(r.get("report_no") or "").strip() == str(report_no):
                return True
    except Exception:
        pass

    return False

def allocate_unique_report_no(make_report_func, req, tfr_requests, max_try=2):
    """
    C·∫•p v√† c·ªë ƒë·ªãnh report_no ƒë√∫ng logic:
    - N·∫øu req ƒë√£ c√≥ report_no: ki·ªÉm tra d√≤ng B==report_no c√≤n tr·ªëng (C..X). N·∫øu ƒë√£ c√≥ d·ªØ li·ªáu -> b√°o l·ªói.
    - N·∫øu ch∆∞a c√≥: ƒë·ªÉ make_report_func ch·ªçn D√íNG TR·ªêNG (C..X tr·ªëng) v√† tr·∫£ v·ªÅ report_no t∆∞∆°ng ·ª©ng.
    - C√≥ retry nh·∫π (bump s·ªë) n·∫øu hi h·ªØu b·ªã chi·∫øm ch·ªó gi·ªØa ch·ª´ng.
    """
    tries = 0

    # Case A: preset report_no trong req
    fixed_req = dict(req)
    preset = str(fixed_req.get("report_no", "")).strip()
    if preset:
        if row_is_filled_for_report(local_main, preset):
            raise RuntimeError(f"M√£ report {preset} ƒë√£ c√≥ d·ªØ li·ªáu, kh√¥ng th·ªÉ ghi ƒë√®.")
        pdf_path, report_no = make_report_func(fixed_req)  # docx_utils ∆∞u ti√™n s·ªë ƒë√£ set
        return pdf_path, report_no

    # Case B: ch∆∞a c√≥ -> ƒë·ªÉ make_report_func ch·ªçn d√≤ng C..X tr·ªëng
    while True:
        pdf_path, report_no = make_report_func(req)

        # X√°c nh·∫≠n l·∫°i: d√≤ng v·∫´n c√≤n tr·ªëng?
        if not row_is_filled_for_report(local_main, report_no):
            return pdf_path, report_no

        # Hi h·ªØu: ai ƒë√≥ v·ª´a ƒëi·ªÅn v√†o d√≤ng n√†y gi·ªØa ch·ª´ng -> th·ª≠ l·∫°i (bump s·ªë + regenerate)
        tries += 1
        if tries >= max_try:
            raise RuntimeError("Kh√¥ng t√¨m ƒë∆∞·ª£c d√≤ng tr·ªëng ƒë·ªÉ c·∫•p m√£ report.")

        # Xo√° file v·ª´a sinh (ƒëi nh·∫ßm d√≤ng)
        try:
            outdir = os.path.join('static', 'TFR')
            for ext in ('.pdf', '.docx'):
                fp = os.path.join(outdir, f"{report_no}{ext}")
                if os.path.exists(fp):
                    os.remove(fp)
        except Exception:
            pass

        # Bump s·ªë v√† t√°i t·∫°o v·ªõi s·ªë c·ªë ƒë·ªãnh
        tries += 1
        if tries >= max_try:
            raise RuntimeError("Kh√¥ng c·∫•p ƒë∆∞·ª£c report_no duy nh·∫•t sau nhi·ªÅu l·∫ßn th·ª≠")

        bumped = bump_report_no(report_no)
        fixed_req = dict(req)
        fixed_req["report_no"] = bumped
        pdf_path, report_no = make_report_func(fixed_req)

# ---- ARCHIVE REQUEST LOG ----
def archive_request(short_data):
    now = datetime.now()
    archive = safe_read_json(ARCHIVE_LOG)
    def get_dt(d):
        if "-" in d: return datetime.strptime(d, "%Y-%m-%d")
        else: return datetime.strptime(d, "%d/%m/%Y")
    archive = [r for r in archive if (now - get_dt(r["request_date"])).days < 14]
    archive.append(short_data)
    safe_write_json(ARCHIVE_LOG, archive)
    safe_append_backup_json(ARCHIVE_LOG, short_data) 

# --- ADD NEW: cleanup archive file (>14 ng√†y) ---
def cleanup_archive_json(days=14):
    """
    X√≥a c√°c b·∫£n ghi archive qu√° 'days' ng√†y (x√≥a th·∫≠t trong JSON).
    ∆Øu ti√™n ARCHIVE_LOG / TFR_ARCHIVE_FILE n·∫øu c√≥; n·∫øu kh√¥ng suy ra t·ª´ TFR_LOG_FILE.
    """
    try:
        archive_path = globals().get("ARCHIVE_LOG") or globals().get("TFR_ARCHIVE_FILE")
        if not archive_path:
            base, ext = os.path.splitext(TFR_LOG_FILE)
            archive_path = f"{base}_archive.json"

        data = safe_read_json(archive_path)
        if not isinstance(data, list) or not data:
            return
        def _parse_date(s):
            if not s:
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(str(s), fmt).date()
                except Exception:
                    pass
            return None

        today = datetime.now(pytz.timezone("Asia/Ho_Chi_Minh")).date()
        kept = []
        for r in data:
            d = None
            if isinstance(r, dict):
                d = _parse_date(r.get("approved_date")) or _parse_date(r.get("etd")) or _parse_date(r.get("request_date"))
            if not d or (today - d).days <= days:
                kept.append(r)
        if len(kept) != len(data):
            safe_write_json(archive_path, kept)
    except Exception as _e:
        print("cleanup_archive_json error:", _e)

# ---- HOME PAGE ----
@app.route("/", methods=["GET", "POST"])
def home():
    message = None

    # Handle item# search and login/set_staff_id POST actions
    if request.method == "POST":
        item_search = request.form.get("item_search", "").strip()
        if item_search:
            return redirect(url_for("home", item_search=item_search))
        if request.form.get("action") == "login":
            password_input = request.form.get("password", "")
            if login(password_input):
                # auth.login ƒë√£ set session['auth_ok'] v√† session['user_type']
                session['role'] = get_user_type()  # 'stl' / 'wtl' / 'vfr3'
                return redirect(url_for("home"))
            else:
                message = "Incorrect password. Please try again."
        elif request.form.get("action") == "set_staff_id":
            staff_id = request.form.get("staff_id", "").strip()
            # Regex: s·ªë - h·ªç t√™n (h·ªó tr·ª£ Unicode ti·∫øng Vi·ªát)
            pattern = r'^\d+\s*-\s*[\p{L}]+(?:\s+[\p{L}]+){1,}$'
            if not staff_id:
                message = "Please enter your ID!"
            elif not regex.match(pattern, staff_id):
                message = ("Please enter the correct format: Staff ID - Full name "
                          "(e.g., 19797 - Nguyen Van A)")
            else:
                session["staff_id"] = staff_id
                return redirect(url_for("home"))


    # ==== Load Excel data ====
    full_report_list = []
    type_of_set = set()
    all_statuses = ['LATE', 'MUST', 'DUE', 'ACTIVE', 'COMPLETE', 'DONE']
    raw_status_set = set()
    try:
        wb = load_workbook(local_main)
        ws = wb.active
        def clean_col(s):
            s = str(s).lower().strip()
            s = re.sub(r'[^a-z0-9#]+', '', s)
            return s
        headers = {}
        for col in range(2, ws.max_column + 1):
            name = ws.cell(row=1, column=col).value
            if name:
                clean = clean_col(name)
                headers[clean] = col
        report_col    = headers.get("report#")
        item_col      = headers.get("item#")
        status_col    = headers.get("status")
        test_date_col = headers.get("logindate")
        type_of_col   = headers.get("typeof")
        etd_col       = headers.get("etd")
        if None in (report_col, item_col, status_col, test_date_col):
            message = f"Missing columns in Excel file! Found: {headers}"
        else:
            for row in range(2, ws.max_row + 1):
                status_raw = ws.cell(row=row, column=status_col).value
                status = str(status_raw).strip().upper() if status_raw else ""
                report = ws.cell(row=row, column=report_col).value
                item = ws.cell(row=row, column=item_col).value
                etd = ws.cell(row=row, column=etd_col).value if etd_col else ""
                type_of = ws.cell(row=row, column=type_of_col).value if type_of_col else ""
                log_date = ws.cell(row=row, column=test_date_col).value
                log_date_str = str(log_date).strip() if log_date else ""
                if log_date_str: log_date_str = log_date_str.split()[0]
                r_dict = {
                    "report": str(report).strip() if report else "",
                    "item": str(item).strip() if item else "",
                    "status": status,
                    "type_of": str(type_of).strip() if type_of else "",
                    "log_date": log_date_str,
                    "etd": etd if etd is not None else ""
                }
                full_report_list.append(r_dict)
                if r_dict["type_of"]:
                    type_of_set.add(r_dict["type_of"])
                if status:
                    raw_status_set.add(status)
        type_of_set = sorted(type_of_set)
    except Exception as e:
        message = f"Error reading list: {e}"

    status_set = []
    all_statuses = ['LATE', 'MUST', 'DUE', 'ACTIVE', 'COMPLETE', 'DONE']
    for s in all_statuses:
        if s in raw_status_set: status_set.append(s)
    for s in sorted(raw_status_set):
        if s not in status_set: status_set.append(s)

    # --- LOGIC L·ªåC STATUS ---
    selected_status = request.args.getlist("status")
    filter_from_user = "status" in request.args

    if not filter_from_user:
        # M·ªõi v√†o trang, m·∫∑c ƒë·ªãnh l·ªçc theo LATE, MUST, DUE
        selected_status = ["LATE", "MUST", "DUE"]
    else:
        # N·∫øu form l·ªçc ƒë∆∞·ª£c g·ª≠i (d√π b·∫•m All hay ch·ªçn t·ª´ng status)
        # N·∫øu kh√¥ng ch·ªçn g√¨ ho·∫∑c ch·ªâ tick All ‚Üí ALL (kh√¥ng filter theo status)
        if not selected_status or selected_status == [""]:
            selected_status = []
        elif "" in selected_status:
            # N·∫øu c√≥ tick c·∫£ All + c√°c status kh√°c, v·∫´n xem nh∆∞ ALL
            selected_status = []

    selected_type = request.args.get("type_of", "")
    item_search = request.args.get("item_search", "").strip()

    report_list = full_report_list
    if item_search:
        # Khi t√¨m item th√¨ lu√¥n t√¨m tr√™n to√†n b·ªô danh s√°ch, kh√¥ng l·ªçc theo tr·∫°ng th√°i!
        report_list = [r for r in full_report_list if item_search.lower() in (r["item"] or "").lower()]
        if selected_type:
            report_list = [r for r in report_list if r["type_of"] == selected_type]
        def safe_report_key(r):
            try:
                return int(r["report"])
            except:
                return r["report"]
        report_list = sorted(report_list, key=safe_report_key)
    else:
        if selected_type:
            report_list = [r for r in report_list if r["type_of"] == selected_type]
        if selected_status:
            report_list = [r for r in report_list if r["status"] in selected_status]

    try:
        page = int(request.args.get("page", "1"))
    except:
        page = 1
    try:
        page_size = int(request.args.get("page_size", "10"))
    except:
        page_size = 10
    if page_size not in [10, 15, 20]: page_size = 10
    total_reports = len(report_list)
    total_pages = max((total_reports + page_size - 1) // page_size, 1)
    if page < 1: page = 1
    if page > total_pages: page = total_pages
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    report_list_page = report_list[start_idx:end_idx]

    type_shortname = {
        "CONSTRUCTION": "CON",
        "FINISHING": "FIN",
        "MATERIAL": "MAT",
        "PACKING": "PAC",
        "GENERAL": "GEN",
    }
    summary_by_type = []
    for t in type_of_set:
        late   = sum(1 for r in full_report_list if r['type_of'] == t and r['status'] == "LATE")
        due    = sum(1 for r in full_report_list if r['type_of'] == t and r['status'] == "DUE")
        must   = sum(1 for r in full_report_list if r['type_of'] == t and r['status'] == "MUST")
        active = sum(1 for r in full_report_list if r['type_of'] == t and r['status'] == "ACTIVE")
        total  = late + due + must + active

        t_up = (t or "").upper()
        if t_up.startswith("OUTSOURCE"):
            # L·∫•y 3 k√Ω t·ª± cu·ªëi sau d·∫•u '-' n·∫øu c√≥; n·∫øu kh√¥ng c√≥ '-' th√¨ l·∫•y 3 k√Ω t·ª± cu·ªëi c·ªßa to√†n chu·ªói
            parts = t_up.split("-")
            short = (parts[-1][-3:] if len(parts) >= 2 else t_up[-3:])
        else:
            short = type_shortname.get(t, t_up[:3])

        summary_by_type.append({
            "type_of": t,
            "short": short,
            "late": late,
            "due": due,
            "must": must,
            "active": active,
            "total": total,
        })

    counter = {"office": 0, "ot": 0}
    path = "counter_stats.json"
    if os.path.exists(path):
        counter = safe_read_json(path)

    return render_template(
        "home.html",
        message=message,
        type_of_set=type_of_set,
        selected_type=selected_type,
        status_set=status_set,
        selected_status=selected_status,
        session=session,
        report_list=report_list,
        summary_by_type=summary_by_type,
        counter=counter,
        page=page,
        total_pages=total_pages,
        page_size=page_size,
        total_reports=total_reports,
        request=request,
        darkmode=session.get('darkmode', '0'),
        slang=session.get('lang', 'vi'),
    )

def get_category_component_position(finishing_type, material_type):
    # material_type: ch·ªâ nh·∫≠n WOOD ho·∫∑c METAL (n√™n x·ª≠ l√Ω hoa th∆∞·ªùng h√≥a)
    if not finishing_type or not material_type:
        return ""
    finishing_type = finishing_type.strip().upper()
    material_type = material_type.strip().upper()
    if finishing_type == "QA TEST":
        if material_type == "WOOD":
            return "COLOR PANEL"
        elif material_type == "METAL":
            return "METAL"
    elif finishing_type == "LINE TEST":
        if material_type == "WOOD":
            return "LINE TEST_COLOR"
        elif material_type == "METAL":
            return "LINE TEST_METAL"
    return ""

def _load_pending_locked():
    with PENDING_LOCK:
        return safe_read_json(TFR_LOG_FILE)

def _write_pending_locked(data):
    with PENDING_LOCK:
        safe_write_json(TFR_LOG_FILE, data)

# ==== Helpers nh√≥m test ====
def _group_of(test_group: str) -> str:
    """
    Chu·∫©n ho√° nh√≥m test ƒë·ªÉ t√≠nh ETD: CONSTRUCTION / TRANSIT / FINISHING / MATERIAL
    """
    g = (test_group or "").strip().upper()
    if "CONSTRUCTION" in g: return "CONSTRUCTION"
    if "TRANSIT" in g:      return "TRANSIT"
    if "FINISHING" in g:    return "FINISHING"
    if "MATERIAL" in g:     return "MATERIAL"
    return g or "OTHER"

def compute_request_date_now(cutoff_hour: int = 15) -> str:
    """
    Quy t·∫Øc request_date:
    - Tr∆∞·ªõc 15:00  -> h√¥m nay
    - T·ª´ 15:00 tr·ªü ƒëi -> ng√†y mai
    """
    now = datetime.now()
    today = now.date()
    if now.hour >= cutoff_hour:
        return (today + timedelta(days=1)).strftime("%Y-%m-%d")
    return today.strftime("%Y-%m-%d")

def _count_by_date_and_group(all_reqs, req_date: str, group_name: str) -> int:
    """
    ƒê·∫øm s·ªë request theo (request_date, group) tr√™n 1 danh s√°ch 'all_reqs'.
    ƒê·∫øm THEO REQUEST (m·ªói record = 1), v√† ch·ªâ t√≠nh c√°c record c√≥ status != Declined
    (v·ªõi pending), nh·∫±m ph·ª•c v·ª• t√≠nh ETD.
    """
    gn = _group_of(group_name)
    dd = (req_date or "").strip()
    c = 0
    for r in (all_reqs or []):
        try:
            r_date = (r.get("request_date") or "").strip()
            r_group = _group_of(r.get("test_group") or r.get("type_of_test"))
            if r_date == dd and r_group == gn:
                st = (r.get("status") or "").strip()
                # Pending: ch·ªâ t√≠nh Submitted; n·∫øu record ƒë·∫øn t·ª´ archive c√≥ th·ªÉ kh√¥ng c√≥ status -> v·∫´n t√≠nh
                if st and st != "Submitted" and st != "Approved":
                    # lo·∫°i Declined v√† c√°c tr·∫°ng th√°i pending kh√°c
                    continue
                c += 1
        except Exception:
            continue
    return c


def _build_reportno_to_group_map():
    """
    D√≤ Excel 'local_main' ƒë·ªÉ map report_no -> group chu·∫©n ho√° (CONSTRUCTION/TRANSIT/FINISHING/MATERIAL/OTHER)
    D·ª±a v√†o c·ªôt 'type of' m√† approve_all_one() ƒë√£ ƒëi·ªÅn.
    """
    try:
        wb = safe_load_excel(local_main)
        ws = wb.active
        col_report = get_col_idx(ws, "report#")
        col_typeof = get_col_idx(ws, "type of")
        if not col_report or not col_typeof:
            return {}

        mapping = {}
        for row in range(2, ws.max_row + 1):
            rep = ws.cell(row=row, column=col_report).value
            tp  = ws.cell(row=row, column=col_typeof).value
            rep_s = ("" if rep is None else str(rep)).strip()
            tp_s  = ("" if tp  is None else str(tp )).strip()
            if not rep_s:
                continue
            # Excel l∆∞u "type of" KH√îNG c√≥ " TEST", n√™n th√™m " TEST" ƒë·ªÉ _group_of() hi·ªÉu,
            # ho·∫∑c b·∫°n c√≥ th·ªÉ map th·∫≥ng n·∫øu mu·ªën.
            grp = _group_of(tp_s + " TEST") if tp_s else "OTHER"
            mapping[rep_s] = grp
        return mapping
    except Exception:
        return {}


def calculate_default_etd(request_date: str, test_group: str, *, all_reqs=None) -> str:
    """
    ETD m·∫∑c ƒë·ªãnh, t√≠nh t·ª´ request_date (t√≠nh C·∫¢ ng√†y request).
    T·∫¢I TRONG NG√ÄY = Pending (ch·ªâ Submitted) + Approved (Archive), lo·∫°i b·ªè Declined.

    - CONSTRUCTION / TRANSIT: 3 ng√†y  -> base +2 ng√†y
      * t·∫£i (trong c√πng request_date), ƒë·∫øm THEO REQUEST:
        - ƒë√£ c√≥ ‚â•5  req  (ƒëang l√† req #6..#10)  -> +1 ng√†y
        - ƒë√£ c√≥ ‚â•10 req (ƒëang l√† req #11..#15) -> +2 ng√†y

    - FINISHING / MATERIAL : 5 ng√†y  -> base +4 ng√†y
      * t·∫£i:
        - ƒë√£ c√≥ ‚â•15 req (ƒëang l√† #16..#30) -> +2 ng√†y
        - ƒë√£ c√≥ ‚â•30 req (ƒëang l√† #31..#45) -> +4 ng√†y
    """
    if not request_date:
        return ""

    g = _group_of(test_group)
    if g in ("CONSTRUCTION", "TRANSIT"):
        base = 2   # 3 ng√†y t√≠nh c·∫£ ng√†y request => +2
    elif g in ("FINISHING", "MATERIAL"):
        base = 4   # 5 ng√†y t√≠nh c·∫£ ng√†y request => +4
    else:
        base = 2

    # GH√âP Pending (ch·ªâ Submitted) + Archive ƒë·ªÉ ƒë·∫øm t·∫£i theo ng√†y/type
    try:
        archive_list = safe_read_json(ARCHIVE_LOG) or []
    except Exception:
        archive_list = []

    # Archive kh√¥ng c√≥ test_group, n√™n join qua Excel ƒë·ªÉ g√°n group
    rep2grp = _build_reportno_to_group_map()
    archive_mapped = []
    for a in archive_list:
        try:
            # gi·ªØ c√°c kh√≥a c·∫ßn cho _count_by_date_and_group
            req_date = (a.get("request_date") or "").strip()
            rep_no   = (a.get("report_no") or "").strip()
            grp      = rep2grp.get(rep_no, "")
            if not req_date or not grp:
                continue
            archive_mapped.append({
                "request_date": req_date,
                "test_group": grp,       # ƒë·ªÉ _group_of hi·ªÉu
                "status": "Approved",    # ƒë√°nh d·∫•u ƒë·ªÉ l·ªçc h·ª£p l·ªá
            })
        except Exception:
            continue

    combined = []
    # Pending (all_reqs) ‚Äì ch·ªâ mu·ªën Submitted
    if isinstance(all_reqs, list):
        combined += [r for r in all_reqs if (r.get("status") or "").strip() == "Submitted"]
    # Approved (archive ƒë√£ map)
    combined += archive_mapped

    cnt = _count_by_date_and_group(combined, request_date, g)

    extra = 0
    if g in ("CONSTRUCTION", "TRANSIT"):
        if cnt >= 10:      # ƒëang l√† #11..#15
            extra = 2
        elif cnt >= 5:     # ƒëang l√† #6..#10
            extra = 1
    elif g in ("FINISHING", "MATERIAL"):
        if cnt >= 30:      # ƒëang l√† #31..#45
            extra = 4
        elif cnt >= 15:    # ƒëang l√† #16..#30
            extra = 2

    d0 = datetime.strptime(request_date, "%Y-%m-%d").date()
    etd = d0 + timedelta(days=base + extra)
    return etd.strftime("%Y-%m-%d")

TFR_INIT_DIR = os.path.join('static', 'TFR_INIT')
os.makedirs(TFR_INIT_DIR, exist_ok=True)

def _save_initial_img(file_storage, trq_id):
    """L∆∞u ·∫£nh initial theo TRQ-ID, tr·∫£ v·ªÅ ƒë∆∞·ªùng d·∫´n t∆∞∆°ng ƒë·ªëi d∆∞·ªõi /static (v√≠ d·ª•: 'TFR_INIT/TRQ123_20250101_120102.jpg')."""
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return None
    fname = secure_filename(file_storage.filename)
    ext = (fname.rsplit('.', 1)[-1] if '.' in fname else 'jpg').lower()
    if not allowed_file(fname):
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"{trq_id}_{stamp}.{ext}"
    abs_path = os.path.join(TFR_INIT_DIR, out_name)
    file_storage.save(abs_path)
    return f"TFR_INIT/{out_name}"

@app.route("/tfr_request_form", methods=["GET", "POST"])
def tfr_request_form():
    tfr_requests = _load_pending_locked()
    error = ""
    form_data = {}
    missing_fields = []

    # L·∫•y tham s·ªë t·ª´ URL ho·∫∑c t·ª´ POST
    trq_id = request.args.get("trq_id") or request.form.get("trq_id")
    edit_idx = request.args.get("edit_idx") or request.form.get("edit_idx")
    editing = False

    # N·∫øu c√≥ trq_id + edit_idx -> ƒëang ·ªü ch·∫ø ƒë·ªô EDIT: n·∫°p s·∫µn d·ªØ li·ªáu v√†o form_data
    if trq_id:
        try:
            # L·∫•y t·∫•t c·∫£ v·ªã tr√≠ c√≥ c√πng TRQ-ID trong file g·ªëc
            matches = [i for i, req in enumerate(tfr_requests) if (req.get("trq_id") or "").strip() == str(trq_id).strip()]

            if matches:
                # N·∫øu c√≥ nhi·ªÅu b·∫£n ghi c√πng TRQ (tr∆∞·ªùng h·ª£p admin gi·ªØ TRQ khi duplicate)
                # v√† edit_idx POST l√™n l√† ordinal trong 'matches' th√¨ d√πng, ng∆∞·ª£c l·∫°i l·∫•y ph·∫ßn t·ª≠ ƒë·∫ßu ti√™n.
                sel = 0
                if edit_idx is not None:
                    try:
                        _ordinal = int(edit_idx)
                        if 0 <= _ordinal < len(matches):
                            sel = _ordinal
                    except Exception:
                        pass

                abs_idx = matches[sel]
                form_data = tfr_requests[abs_idx].copy()
                editing = True
                # ƒê·∫£m b·∫£o hidden edit_idx l√† "ch·ªâ s·ªë tuy·ªát ƒë·ªëi" ƒë·ªÉ c√°c l·∫ßn submit sau kh√¥ng l·ªách
                form_data["edit_idx"] = str(abs_idx)
            else:
                # Kh√¥ng t√¨m th·∫•y theo TRQ-ID -> coi nh∆∞ t·∫°o m·ªõi
                editing = False
        except Exception:
            editing = False

    if request.method == "POST":
        form = request.form

        required_fields = [
            "requestor", "employee_id", "department", "request_date",
            "sample_description", "test_status", "quantity", "sample_return"
        ]
        for field in required_fields:
            if not form.get(field) and not form.get(f"{field}_na"):
                missing_fields.append(field)

        test_group = form.get("test_group", "")
        if not test_group:
            missing_fields.append("test_group")
            error = "Ph·∫£i ch·ªçn lo·∫°i test!"

        furniture_testing = form.get("furniture_testing", "")
        if not furniture_testing:
            missing_fields.append("furniture_testing")
            error = "Ph·∫£i ch·ªçn Indoor ho·∫∑c Outdoor!"

        finishing_type = form.get("finishing_type", "")
        material_type = form.get("material_type", "")

        # form_data ƒë·ªÉ render l·∫°i khi l·ªói
        form_data = form.to_dict(flat=True)
        form_data["test_group"] = test_group
        form_data["furniture_testing"] = furniture_testing
        form_data["trq_id"] = form.get("trq_id", trq_id)
        form_data["employee_id"] = form.get("employee_id", "").strip()
        form_data["sample_return"] = form.get("sample_return", "")
        form_data["remark"] = form.get("remark", "").strip()
        form_data["finishing_type"] = finishing_type
        form_data["material_type"] = material_type

        # gi·ªØ l·∫°i edit_idx qua POST n·∫øu c√≥
        if edit_idx is not None:
            form_data["edit_idx"] = edit_idx

        def na_or_value(key):
            na_key = key + "_na"
            if form.get(na_key):
                return "N/A"
            return form.get(key, "").strip()

        if test_group == "FINISHING TEST" and not finishing_type:
            missing_fields.append("finishing_type")
            error = "Ph·∫£i ch·ªçn QA TEST ho·∫∑c LINE TEST!"

        # --- Rule ri√™ng: N·∫øu Department = VFR5 th√¨ Subcon b·∫Øt bu·ªôc v√† kh√¥ng ƒë∆∞·ª£c N/A ---
        department = form.get("department", "").strip()
        subcon_val = form.get("subcon", "").strip()
        subcon_na  = form.get("subcon_na")

        if department.upper() == "VFR5":
            if not subcon_val or subcon_na:
                missing_fields.append("subcon")
                error = "If Department is VFR5, you need to fill Subcon."

        # N·∫øu c√≥ thi·∫øu, tr·∫£ v·ªÅ form k√®m l·ªói
        if missing_fields:
            if not error:
                error = "Vui l√≤ng ƒëi·ªÅn ƒë·ªß c√°c tr∆∞·ªùng b·∫Øt bu·ªôc (*)"
            return render_template(
                "tfr_request_form.html",
                darkmode=session.get("darkmode", "0"),
                lang=session.get("lang", "vi"),
                error=error,
                form_data=form_data,
                missing_fields=missing_fields,
                editing=bool(edit_idx is not None),
                trq_id=trq_id,
                edit_idx=edit_idx
            )

        # ---- Build new_request t·ª´ form ----
        # L·∫•y request_date: n·∫øu user ƒë·ªÉ tr·ªëng -> d√πng rule 15:00 (prefill v·∫´n cho s·ª≠a)
        request_date_input = (form.get("request_date") or "").strip()
        if not request_date_input:
            request_date_input = compute_request_date_now()

        item_code = na_or_value("item_code")
        supplier = na_or_value("supplier")
        subcon = na_or_value("subcon")

        test_status = form.get("test_status")
        if test_status == "nth":
            nth = form.get("test_status_nth", "").strip()
            test_status = nth + "th" if nth.isdigit() else "nth"

        remark = form.get("remark", "").strip()
        if test_group == "FINISHING TEST" and finishing_type:
            remark = f"{remark} ({finishing_type})" if remark else finishing_type

        new_request = {
            "trq_id": form.get("trq_id", trq_id),
            "requestor": form.get("requestor"),
            "employee_id": form.get("employee_id", ""),
            "department": form.get("department"),
            "request_date": request_date_input,  # <-- cho s·ª≠a, nh∆∞ng n·∫øu tr·ªëng ƒë√£ auto set ·ªü tr√™n
            "sample_description": na_or_value("sample_description"),
            "item_code": item_code,
            "supplier": supplier,
            "subcon": subcon,
            "test_status": test_status,
            "furniture_testing": furniture_testing,
            "quantity": form.get("quantity"),
            "sample_return": form.get("sample_return", ""),
            "remark": remark,
            "test_group": test_group,
            "finishing_type": finishing_type,
            "material_type": material_type,
            "status": "Submitted",
            "decline_reason": "",
            "report_no": ""
        }

        # ‚úÖ T·ª± t√≠nh ETD theo rule + t·∫£i, d·ª±a tr√™n danh s√°ch hi·ªán c√≥ (ƒë·ªÉ ƒë·∫øm theo request_date & group)
        new_request["etd"] = calculate_default_etd(
            new_request.get("request_date", ""),
            new_request.get("test_group", ""),
            all_reqs=tfr_requests   # <‚Äî th√™m d√≤ng n√†y
        )

        # N·∫øu l√† EDIT: gi·ªØ l·∫°i c√°c tr∆∞·ªùng h·ªá th·ªëng c≈© (PDF/DOCX/report_no/etd/status/decline_reason)
        if editing or (trq_id and edit_idx is not None):
            try:
                _edit_idx_int = int(edit_idx)
                matches = [i for i, req in enumerate(tfr_requests) if req.get("trq_id") == trq_id]
                if len(matches) > _edit_idx_int:
                    old = tfr_requests[matches[_edit_idx_int]]
                    for k in ("status", "report_no", "pdf_path", "docx_path", "etd", "decline_reason"):
                        if k in old and old.get(k) not in (None, ""):
                            new_request[k] = old.get(k)
            except Exception:
                pass

        # ---- Ghi ƒë√® item c≈© ho·∫∑c append m·ªõi ----
        if trq_id and edit_idx is not None:
            try:
                _abs = int(edit_idx)
                if 0 <= _abs < len(tfr_requests) and tfr_requests[_abs].get("trq_id") == trq_id:
                    tfr_requests[_abs] = new_request
                else:
                    # Fallback theo ordinal trong nh√≥m c√πng trq_id
                    matches = [i for i, req in enumerate(tfr_requests) if req.get("trq_id") == trq_id]
                    if len(matches) > _abs:
                        tfr_requests[matches[_abs]] = new_request
                    else:
                        tfr_requests.append(new_request)
            except Exception:
                tfr_requests.append(new_request)
        else:
            # T·∫°o m·ªõi: n·∫øu ch∆∞a c√≥ TRQ-ID (v√≠ d·ª• truy c·∫≠p tr·ª±c ti·∫øp POST), th√¨ sinh m·ªõi
            if not new_request.get("trq_id"):
                existing_ids = {r.get("trq_id") for r in tfr_requests if r.get("trq_id")}
                new_request["trq_id"] = generate_unique_trq_id(existing_ids)
            tfr_requests.append(new_request)

        # ·∫¢NH BAN ƒê·∫¶U (INITIAL PRODUCT IMAGE)
        init_files = request.files.getlist("initial_img")  # input name="initial_img" + multiple
        delete_flag = (form.get("delete_initial_img") == "1")

        # L·∫•y ·∫£nh c≈© n·∫øu ƒëang edit (ƒë·ªÉ gi·ªØ nguy√™n khi kh√¥ng upload m·ªõi)
        old_initial_img = None
        old_initial_images = []
        if editing:
            old_list = safe_read_json(TFR_LOG_FILE) or []
            try:
                idx_keep = int(form.get("edit_idx", "-1"))
                if 0 <= idx_keep < len(old_list):
                    old_initial_img = old_list[idx_keep].get("initial_img")
                    old_initial_images = old_list[idx_keep].get("initial_images") or []
                    # n·∫øu b·∫£n c≈© ch·ªâ c√≥ initial_img (chu·ªói), convert th√†nh list cho ƒë·ªìng b·ªô
                    if (not old_initial_images) and isinstance(old_initial_img, str) and old_initial_img:
                        old_initial_images = [old_initial_img]
            except Exception:
                pass

        new_initial_images = []

        if delete_flag:
            # Ng∆∞·ªùi d√πng y√™u c·∫ßu x√≥a ·∫£nh initial khi edit
            new_request["initial_img"] = None
            new_request["initial_images"] = []
        else:
            if init_files:
                # C√≥ upload m·ªõi: l∆∞u t·ªëi ƒëa 2 ·∫£nh h·ª£p l·ªá
                for f in init_files[:2]:
                    if not f or not f.filename:
                        continue
                    saved = _save_initial_img(f, new_request["trq_id"])  # tr·∫£ v·ªÅ "TFR_INIT/xxx.ext"
                    if saved:
                        new_initial_images.append(saved)

                if new_initial_images:
                    new_request["initial_images"] = new_initial_images
                    new_request["initial_img"] = new_initial_images[0]  # gi·ªØ key c≈© cho UI c≈©
                else:
                    # Kh√¥ng c√≥ file h·ª£p l·ªá -> n·∫øu ƒëang edit th√¨ gi·ªØ ·∫£nh c≈©, ng∆∞·ª£c l·∫°i None
                    if editing and old_initial_images:
                        new_request["initial_images"] = old_initial_images
                        new_request["initial_img"] = old_initial_images[0]
                    else:
                        new_request["initial_images"] = []
                        new_request["initial_img"] = None
            else:
                # Kh√¥ng upload m·ªõi -> n·∫øu edit th√¨ gi·ªØ ·∫£nh c≈©, n·∫øu t·∫°o m·ªõi th√¨ None
                if editing and old_initial_images:
                    new_request["initial_images"] = old_initial_images
                    new_request["initial_img"] = old_initial_images[0]
                else:
                    new_request["initial_images"] = []
                    new_request["initial_img"] = None

        # Ghi log nh∆∞ c≈©
        safe_write_json(TFR_LOG_FILE, tfr_requests)
        safe_append_backup_json(TFR_LOG_FILE, new_request)

        message = (
            f"üìù [TRF] C√≥ y√™u c·∫ßu Test Request m·ªõi!\n"
            f"- Ng∆∞·ªùi g·ª≠i: {new_request.get('requestor')}\n"
            f"- B·ªô ph·∫≠n: {new_request.get('department')}\n"
            f"- Ng√†y g·ª≠i: {new_request.get('request_date')}\n"
            f"- Nh√≥m test: {new_request.get('test_group')}\n"
            f"- S·ªë l∆∞·ª£ng: {new_request.get('quantity')}\n"
            f"- M√£ TRQ-ID: {new_request.get('trq_id')}"
        )
        send_teams_message(TEAMS_WEBHOOK_URL_TRF, message)

        return redirect(url_for('tfr_request_status'))

    # ===== GET l·∫ßn ƒë·∫ßu (kh√¥ng EDIT) -> auto fill employee_id, requestor t·ª´ session
    if not editing:
        staff_id_full = session.get("staff_id", "").strip()
        if staff_id_full and "-" in staff_id_full:
            emp_id, name = staff_id_full.split("-", 1)
            emp_id = emp_id.strip()
            name = name.strip()
        else:
            emp_id = staff_id_full
            name = ""
        form_data.setdefault("employee_id", emp_id)
        form_data.setdefault("requestor", name)

    # T·∫°o TRQ-ID m·ªõi n·∫øu ch∆∞a c√≥
    if not form_data.get("trq_id"):
        form_data["trq_id"] = generate_unique_trq_id({r.get("trq_id") for r in tfr_requests if "trq_id" in r})

    # Prefill request_date theo rule 15:00 (nh∆∞ng user v·∫´n c√≥ th·ªÉ s·ª≠a ·ªü form)
    if not form_data.get("request_date"):
        form_data["request_date"] = compute_request_date_now()

    return render_template(
        "tfr_request_form.html",
        darkmode=session.get("darkmode", "0"),
        lang=session.get("lang", "vi"),
        error=error,
        form_data=form_data,
        missing_fields=missing_fields,
        editing=editing,
        trq_id=trq_id,
        edit_idx=edit_idx
    )

PENDING_LOCK = Lock()
CANCEL_FLAGS = {} 

def _read_pending():
    return safe_read_json(TFR_LOG_FILE)

def _write_pending(new_list):
    safe_write_json(TFR_LOG_FILE, new_list)

def _merge_update_etd(updates):
    """
    C·∫≠p nh·∫≠t ETD an to√†n theo d·ªØ li·ªáu m·ªõi nh·∫•t trong file:
    - N·∫øu update c√≥ c·∫£ idx & trq_id: ∆∞u ti√™n kh·ªõp trq_id, r·ªìi m·ªõi r∆°i v·ªÅ idx.
    - N·∫øu ch·ªâ c√≥ trq_id: d√πng trq_id.
    - N·∫øu ch·ªâ c√≥ idx: d√πng idx, nh∆∞ng v·∫´n check bounds.
    """
    with PENDING_LOCK:
        cur = _read_pending()
        # T·∫°o map {trq_id: index} tr√™n d·ªØ li·ªáu M·ªöI NH·∫§T
        id_to_idx = {}
        for i, r in enumerate(cur):
            tid = (r.get("trq_id") or "").strip()
            if tid:
                id_to_idx[tid] = i

        changed = False
        for u in updates:
            tid = (u.get("trq_id") or "").strip()
            etd = (u.get("etd") or "").strip()
            idx = u.get("idx")

            # ∆Øu ti√™n d√πng trq_id
            if tid and tid in id_to_idx:
                cur[id_to_idx[tid]]["etd"] = etd
                changed = True
            # Fallback d√πng idx n·∫øu h·ª£p l·ªá
            elif isinstance(idx, int) and 0 <= idx < len(cur):
                cur[idx]["etd"] = etd
                changed = True

        if changed:
            _write_pending(cur)
        return cur  # tr·∫£ v·ªÅ snapshot m·ªõi nh·∫•t sau khi ƒë√£ merge ETD

def _remove_approved_from_file(approved_trq_ids):
    """
    X√≥a c√°c request ƒë√£ Approved RA KH·ªéI FILE theo trq_id (merge an to√†n):
    - Lu√¥n ƒë·ªçc file m·ªõi nh·∫•t
    - L·ªçc b·ªè c√°c ph·∫ßn t·ª≠ c√≥ trq_id thu·ªôc t·∫≠p approved_trq_ids
    - Kh√¥ng ƒë·ª•ng ch·∫°m c√°c request m·ªõi ph√°t sinh
    """
    if not approved_trq_ids:
        return

    with PENDING_LOCK:
        cur = _read_pending()
        keep = []
        approved_set = {tid.strip() for tid in approved_trq_ids if tid}
        for r in cur:
            tid = (r.get("trq_id") or "").strip()
            if tid and tid in approved_set:
                continue  # b·ªè c√°c request v·ª´a approve
            keep.append(r)
        _write_pending(keep)

def make_id_index_map(pending_list):
    """
    (gi·ªØ n·∫øu b·∫°n ƒëang g·ªçi n∆°i kh√°c) ‚Äì map {trq_id: last_index}
    """
    mapping = {}
    if not isinstance(pending_list, list):
        return mapping
    for i, row in enumerate(pending_list):
        try:
            tid = (row.get("trq_id") or "").strip()
        except Exception:
            tid = ""
        if tid:
            mapping[tid] = i
    return mapping

# --- H√ÄM DUY·ªÜT 1 REQUEST (gi·ªØ nguy√™n n·∫øu app b·∫°n ƒëang x√†i) ---
def approve_all_one(req):
    """
    Approve 1 request:
      - c·∫•p report_no + t·∫°o DOCX/PDF
      - c·∫≠p nh·∫≠t Excel + TRF.xlsx
      - ƒë·∫©y v√†o archive
      - tr·∫£ v·ªÅ req ƒë√£ c·∫≠p nh·∫≠t (status/report_no/pdf_path/docx_path)
    """
    # üîí Kho√° file-level cho c·∫£ phi√™n c·∫•p s·ªë -> ghi Excel -> archive
    with report_lock():                              # ‚Üê TH√äM D√íNG N√ÄY
        # (gi·ªØ thread-lock nh∆∞ c≈© ƒë·ªÉ an to√†n trong c√πng process)
        with REPORT_NO_LOCK:
            current_list = safe_read_json(TFR_LOG_FILE)
            pdf_path, report_no = allocate_unique_report_no(
                approve_request_fill_docx_pdf, req, current_list
            )

        req["status"] = "Approved"
        req["decline_reason"] = ""
        req["report_no"] = report_no

        output_folder = os.path.join('static', 'TFR')
        output_docx = os.path.join(output_folder, f"{report_no}.docx")
        output_pdf  = os.path.join(output_folder, f"{report_no}.pdf")

        try:
            if not os.path.exists(output_pdf):
                try_convert_to_pdf(output_docx, output_pdf)
        except Exception as _pdf_e:
            print("PDF convert failed, fallback to DOCX:", _pdf_e)

        if os.path.exists(output_pdf):
            req['pdf_path'] = f"TFR/{report_no}.pdf"
            req['docx_path'] = None
        else:
            req['pdf_path'] = None
            req['docx_path'] = f"TFR/{report_no}.docx"

    # Ghi Excel & TRF.xlsx & archive (gi·ªØ nguy√™n nh∆∞ b·∫°n ƒëang c√≥)
        try:
            write_tfr_to_excel(local_main, report_no, req)
            wb = load_workbook(local_main)
            ws = wb.active
            report_col = get_col_idx(ws, "report#")
            row_idx = None
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=report_col).value
                if v and str(v).strip() == str(report_no):
                    row_idx = row
                    break
            if row_idx:
                def set_val(col_name, value, is_date_col=False):
                    col_idx = get_col_idx(ws, col_name)
                    if col_idx:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if is_date_col:
                            dt_val = try_parse_excel_date(value)
                            if dt_val:
                                cell.value = dt_val
                                cell.number_format = 'dd-mmm'   # <- ƒë·ªïi d-mmm -> dd-mmm
                            else:
                                cell.value = value
                        else:
                            cell.value = value.upper() if isinstance(value, str) else value

                def clean_type_of(val):
                    return val[:-5].strip() if val and isinstance(val, str) and val.upper().endswith(" TEST") else val

                set_val("item#", req.get("item_code", ""))
                set_val("type of", clean_type_of(req.get("test_group", "")))
                set_val("item name/ description", req.get("sample_description", ""))
                set_val("furniture testing", req.get("furniture_testing", ""))
                set_val("submiter in", req.get("requestor", ""))
                set_val("submited", req.get("department", ""))
                set_val("qa comment", req.get("remark", ""))

                etd_val = req.get("etd", "")
                set_val("etd", etd_val, is_date_col=True)  # <-- b·ªè format_excel_date_short


                vn_tz = pytz.timezone("Asia/Ho_Chi_Minh")
                req_login = (req.get("request_date") or "").strip()
                val_for_excel = req_login if req_login else datetime.now(vn_tz).strftime("%Y-%m-%d")
                set_val("log in date", val_for_excel, is_date_col=True)

                finishing_type = req.get("finishing_type", "")
                material_type  = req.get("material_type", "")
                cat_comp_pos   = get_category_component_position(finishing_type, material_type)
                set_val("category / component name / position", cat_comp_pos)
                wb.save(local_main)
        except Exception as e:
            print("Ghi v√†o Excel b·ªã l·ªói:", e)

        try:
            append_row_to_trf(report_no, local_main, "TRF.xlsx", trq_id=req.get("trq_id", ""))
        except Exception as e:
            print("Append TRF l·ªói:", e)

        try:
            vn_tz = pytz.timezone("Asia/Ho_Chi_Minh")
            short_data = {
                "trq_id": req.get("trq_id", ""),
                "report_no": req.get("report_no", ""),
                "requestor": req.get("requestor", ""),
                "department": req.get("department", ""),
                "request_date": req.get("request_date", ""),
                "item_code": req.get("item_code", ""),
                "status": req.get("status", ""),
                "pdf_path": req.get("pdf_path"),
                "docx_path": req.get("docx_path"),
                "employee_id": req.get("employee_id", ""),
                "approved_date": datetime.now(vn_tz).strftime("%Y-%m-%d"),
                "test_group": req.get("test_group", ""),
            }
            archive_request(short_data)
        except Exception as e:
            print("Archive l·ªói:", e)

    return req

# ================== ROUTE: APPROVE ALL (STREAM) ‚Äî ƒê√É S·ª¨A ==================
@app.post("/approve_all_stream")
def approve_all_stream():
    """
    S·ª≠a ch√≠nh:
      1) C·∫≠p nh·∫≠t ETD theo file M·ªöI NH·∫§T (merge) => kh√¥ng ƒë√® m·∫•t request m·ªõi.
      2) Sau M·ªñI request ƒë∆∞·ª£c approve, x√≥a request ƒë√≥ kh·ªèi file b·∫±ng ph√©p "l·ªçc theo trq_id"
         tr√™n d·ªØ li·ªáu M·ªöI NH·∫§T => kh√¥ng bao gi·ªù overwrite c√°c request m·ªõi v·ª´a ƒë∆∞·ª£c g·ª≠i.
      3) Kh√¥ng c√≤n final write "ghi ƒë√® c·∫£ file" theo snapshot c≈© n·ªØa.
    """
    def gen():
        run_id = str(uuid4())
        CANCEL_FLAGS[run_id] = False

        # Nh·∫≠n input
        try:
            data = request.get_json(silent=True) or {}
            updates = data.get("updates", []) or []
        except Exception as e:
            yield json.dumps({"type": "error", "message": f"Parse JSON: {e}"}) + "\n"
            CANCEL_FLAGS.pop(run_id, None)
            return

        # (1) Merge c·∫≠p nh·∫≠t ETD v√†o file hi·ªán t·∫°i (an to√†n)
        try:
            pending_after_etd = _merge_update_etd(updates)
        except Exception as e:
            yield json.dumps({"type": "error", "message": f"Bulk ETD update: {e}"}) + "\n"
            pending_after_etd = _read_pending()

        # (2) L·∫≠p danh s√°ch c·∫ßn duy·ªát (submitted + c√≥ ETD)
        id_to_idx = make_id_index_map(pending_after_etd)
        todo = []
        for u in updates:
            idx = u.get("idx")
            tid = (u.get("trq_id") or "").strip()

            # ∆Øu ti√™n idx n·∫øu c√≤n h·ª£p l·ªá v√† kh·ªõp trq_id (n·∫øu c√≥)
            picked = None
            if isinstance(idx, int) and 0 <= idx < len(pending_after_etd):
                item = pending_after_etd[idx]
                if item and item.get("status") == "Submitted" and (item.get("etd") or "").strip():
                    if not tid or tid == (item.get("trq_id") or "").strip():
                        picked = (idx, (item.get("trq_id") or "").strip(), item)

            # Fallback theo trq_id
            if not picked and tid and tid in id_to_idx:
                j = id_to_idx[tid]
                item = pending_after_etd[j]
                if item and item.get("status") == "Submitted" and (item.get("etd") or "").strip():
                    picked = (j, tid, item)

            if picked:
                todo.append(picked)
        
        def _parse_dt(s: str):
            s = (s or "").strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(s, fmt)
                except Exception:
                    pass
            return datetime.max

        def _norm_type(rec: dict):
            t = (rec.get("type_of_test") or rec.get("test_group") or "")
            return t.replace(" TEST", "").strip().lower()

        # todo l√† list (idx, trq_id, item)
        todo.sort(key=lambda x: (
            _parse_dt(x[2].get("request_date")),
            _norm_type(x[2]),
            (x[2].get("trq_id") or "")
        ))

        yield json.dumps({"type": "start", "total": len(todo), "run_id": run_id}) + "\n"

        # (3) Duy·ªát t·ª´ng request + m·ªói l·∫ßn xong th√¨ g·ª° kh·ªèi file b·∫±ng merge-remove
        done = 0
        approved_tids = []

        for _, tid, item in todo:
            try:
                approved = approve_all_one(dict(item))  # d√πng b·∫£n copy ƒë·ªÉ tr√°nh side-effect
                report_no = (approved or {}).get("report_no") or item.get("report_no")

                # Ghi nh·∫≠n ti·∫øn ƒë·ªô
                done += 1
                approved_tids.append(tid)
                yield json.dumps({
                    "type": "progress",
                    "done": done,
                    "total": len(todo),
                    "trq_id": tid,
                    "report_no": report_no
                }) + "\n"

                # X√≥a request ƒë√£ approve ra kh·ªèi file (MERGE theo tr·∫°ng th√°i file m·ªõi nh·∫•t)
                _remove_approved_from_file([tid])

                # Ng∆∞·ªùi d√πng b·∫•m Cancel -> d·ª´ng sau khi xong request hi·ªán t·∫°i
                if CANCEL_FLAGS.get(run_id):
                    yield json.dumps({"type": "cancelled", "done": done, "total": len(todo)}) + "\n"
                    CANCEL_FLAGS.pop(run_id, None)
                    return

            except Exception as e:
                yield json.dumps({"type": "error", "message": str(e), "trq_id": tid}) + "\n"

        # (4) K·∫øt th√∫c b√¨nh th∆∞·ªùng
        yield json.dumps({"type": "done", "done": done, "total": len(todo)}) + "\n"
        CANCEL_FLAGS.pop(run_id, None)

    return Response(stream_with_context(gen()), mimetype="application/json")


# (tu·ª≥ ch·ªçn) Route cancel gi·ªØ nguy√™n
@app.post("/approve_all_cancel")
def approve_all_cancel():
    data = request.get_json(silent=True) or {}
    run_id = data.get("run_id")
    if not run_id:
        return jsonify(success=False, message="Thi·∫øu run_id"), 400
    if run_id not in CANCEL_FLAGS:
        return jsonify(success=False, message="Run ID kh√¥ng t·ªìn t·∫°i ho·∫∑c ƒë√£ k·∫øt th√∫c"), 404
    CANCEL_FLAGS[run_id] = True
    return jsonify(success=True)

@app.route("/tfr_request_status", methods=["GET", "POST"])
def tfr_request_status():
    try:
        import pandas as pd
    except Exception:
        pd = None

    # ===== Helpers nh·ªè trong route =====
    def _parse_date(s):
        if not s:
            return datetime.max
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        return datetime.max

    def _norm_type(r):
        t = (r.get("type_of_test") or r.get("test_group") or "").strip()
        return t.replace(" TEST", "").strip().lower()

    def _tie_break(r):
        tid = (r.get("trq_id") or "").strip()
        m = re.search(r"(\d+)$", tid)
        return int(m.group(1)) if m else tid

    def _redirect_back():
        back = request.form.get("return_url")
        if back:
            return redirect(back)
        return redirect(url_for('tfr_request_status'))

    # ===== Load & quy·ªÅn =====
    tfr_requests = safe_read_json(TFR_LOG_FILE) or []
    is_admin = session.get("user_type") in ("stl", "superadmin")

    # ===== L·∫•y Staff ID & t√°ch (ID - T√™n) =====
    viewer_staff_id = (session.get("staff_id") or request.args.get("staff_id") or "").strip()
    if viewer_staff_id and "-" in viewer_staff_id:
        _emp_id, _name = viewer_staff_id.split("-", 1)
        viewer_emp_id = _emp_id.strip()
        viewer_name   = _name.strip()
    else:
        viewer_emp_id = ""
        viewer_name   = viewer_staff_id.strip()

    def _eq(a, b):
        return (str(a or "").strip().lower() == str(b or "").strip().lower())

    # ===== L·ªçc hi·ªÉn th·ªã: user th∆∞·ªùng ch·ªâ th·∫•y request c·ªßa m√¨nh (T√™n HO·∫∂C Employee ID) =====
    if not is_admin and (viewer_name or viewer_emp_id):
        tfr_requests = [
            r for r in tfr_requests
            if _eq(r.get("requestor"), viewer_name) or _eq(r.get("employee_id"), viewer_emp_id)
        ]

    # ===== POST actions =====
    if request.method == "POST":
        action = request.form.get("action")
        current = safe_read_json(TFR_LOG_FILE) or []   # snapshot m·ªõi nh·∫•t

        # ---------- APPROVE ALL ----------
        if is_admin and action == "approve_all":
            approved_count = 0
            new_pending = []
            for req in current:
                if (req.get("status") == "Submitted") and (req.get("etd") or "").strip():
                    try:
                        # ‚úÖ log_in_date = request_date
                        req["log_in_date"] = req.get("request_date")
                        approve_all_one(req)
                        approved_count += 1
                        continue
                    except Exception as e:
                        print("Approve one (approve_all) error:", e)
                new_pending.append(req)

            safe_write_json(TFR_LOG_FILE, new_pending)
            flash(f"ƒê√£ duy·ªát {approved_count} request!")
            return _redirect_back()

        # ---------- APPROVE SINGLE ----------
        elif is_admin and action == "approve":
            trq_id = request.form.get("trq_id")
            edit_idx = int(request.form.get("edit_idx", 0)) if "edit_idx" in request.form else None
            matches = [i for i, req in enumerate(current) if req.get("trq_id") == trq_id]
            idx = matches[edit_idx] if edit_idx is not None and edit_idx < len(matches) else (matches[0] if matches else None)
            if idx is not None:
                req = current[idx]
                etd = (request.form.get("etd", "") or "").strip()
                if not etd:
                    flash("B·∫°n c·∫ßn ƒëi·ªÅn Estimated Completion Date (ETD) tr∆∞·ªõc khi approve!")
                    return _redirect_back()

                req["etd"] = etd
                req["estimated_completion_date"] = etd
                # ‚úÖ log_in_date = request_date
                req["log_in_date"] = req.get("request_date")

                try:
                    approve_all_one(req)
                    del current[idx]
                    safe_write_json(TFR_LOG_FILE, current)
                except Exception as e:
                    print("Approve one (single) error:", e)
                    flash("C√≥ l·ªói khi approve, vui l√≤ng th·ª≠ l·∫°i.")
            return _redirect_back()

        # ---------- DECLINE ----------
        elif is_admin and action == "decline":
            trq_id = request.form.get("trq_id")
            reason = (request.form.get("decline_reason", "") or "").strip()
            matches = [i for i, req in enumerate(current) if req.get("trq_id") == trq_id]
            idx = matches[0] if matches else None
            if idx is not None:
                current[idx]["status"] = "Declined"
                current[idx]["decline_reason"] = reason
            safe_write_json(TFR_LOG_FILE, current)
            return _redirect_back()

        # ---------- DUPLICATE ----------
        elif action == "duplicate":
            trq_id  = request.form.get("trq_id")
            edit_idx = int(request.form.get("edit_idx", 0)) if "edit_idx" in request.form else None

            matches = [i for i, req in enumerate(current) if str(req.get("trq_id")) == str(trq_id)]
            idx = matches[edit_idx] if (edit_idx is not None and 0 <= edit_idx < len(matches)) else (matches[0] if matches else None)

            if idx is not None:
                old_req = current[idx]
                new_req = old_req.copy()

                # reset fields cho b·∫£n dup
                new_req["report_no"] = ""
                new_req["status"] = "Submitted"
                new_req["pdf_path"] = ""
                new_req["decline_reason"] = ""

                if is_admin:
                    # Admin: gi·ªØ nguy√™n TRQ-ID (h√†nh vi c≈©)
                    # -> v·∫´n ch√®n ngay sau b·∫£n g·ªëc ƒë·ªÉ ti·ªán edit
                    insert_pos = idx + 1
                    current.insert(insert_pos, new_req)
                    safe_write_json(TFR_LOG_FILE, current)
                    # Admin v·∫´n quay v·ªÅ trang danh s√°ch nh∆∞ c≈©
                    return _redirect_back()
                else:
                    # X√°c th·ª±c ch·ªß s·ªü h·ªØu theo T√äN ho·∫∑c EMPLOYEE ID
                    viewer_staff_id_post = (session.get("staff_id") or request.form.get("staff_id") or request.args.get("staff_id") or "").strip()
                    if viewer_staff_id_post and "-" in viewer_staff_id_post:
                        _emp2, _name2 = viewer_staff_id_post.split("-", 1)
                        owner_emp_id = _emp2.strip()
                        owner_name   = _name2.strip()
                    else:
                        owner_emp_id = ""
                        owner_name   = viewer_staff_id_post.strip()

                    is_owner = _eq(old_req.get("requestor"), owner_name) or _eq(old_req.get("employee_id"), owner_emp_id)
                    if not is_owner:
                        return _redirect_back()

                    # Ng∆∞·ªùi th∆∞·ªùng: t·∫°o TRQ m·ªõi + request_date & ETD m·ªõi (lu√¥n t√≠nh ETD)
                    existing_ids = [str(r.get("trq_id")) for r in current if r.get("trq_id")]
                    new_req["trq_id"] = generate_unique_trq_id(existing_ids)
                    new_req["request_date"] = compute_request_date_now()
                    new_req["etd"] = calculate_default_etd(
                        new_req["request_date"],
                        new_req.get("test_group", ""),
                        all_reqs=current
                    )
                    new_req["estimated_completion_date"] = new_req["etd"]

                    # Ch√®n ngay sau b·∫£n g·ªëc
                    insert_pos = idx + 1
                    current.insert(insert_pos, new_req)
                    safe_write_json(TFR_LOG_FILE, current)

                    # üîÅ NEW: Sau khi Dup th√†nh c√¥ng, chuy·ªÉn th·∫≥ng t·ªõi form edit c·ªßa b·∫£n m·ªõi
                    return redirect(url_for(
                        'tfr_request_form',
                        trq_id=new_req["trq_id"],
                        edit_idx=insert_pos
                    ))

            return _redirect_back()

        # ---------- DELETE ----------
        elif action == "delete":
            trq_id = request.form.get("trq_id")
            edit_idx = request.form.get("edit_idx")
            if edit_idx is not None:
                try:
                    edit_idx = int(edit_idx)
                    deleted_req = current.pop(edit_idx)
                    send_teams_message(
                        TEAMS_WEBHOOK_URL_TRF,
                        f"üóëÔ∏è [TRF] ƒê√£ c√≥ y√™u c·∫ßu b·ªã x√≥a!\n- TRQ-ID: {deleted_req.get('trq_id')}\n- Ng∆∞·ªùi thao t√°c: {session.get('staff_id', 'Kh√¥ng r√µ')}"
                    )
                except Exception as e:
                    print("X√≥a b·ªã l·ªói:", e)
            else:
                for i, req in enumerate(current):
                    if req.get("trq_id") == trq_id:
                        deleted_req = current.pop(i)
                        send_teams_message(
                            TEAMS_WEBHOOK_URL_TRF,
                            f"üóëÔ∏è [TRF] ƒê√£ c√≥ y√™u c·∫ßu b·ªã x√≥a!\n- TRQ-ID: {deleted_req.get('trq_id')}\n- Ng∆∞·ªùi thao t√°c: {session.get('staff_id', 'Kh√¥ng r√µ')}"
                        )
                        break
            safe_write_json(TFR_LOG_FILE, current)
            return _redirect_back()

    # ===== GET view (KH√îNG reload l·∫°i full list; d√πng danh s√°ch ƒë√£ l·ªçc) =====
    sort_mode = request.args.get("sort", "date")

    pairs_declined  = [(i, r) for i, r in enumerate(tfr_requests) if (r.get("status") or "").strip() == "Declined"]
    pairs_submitted = [(i, r) for i, r in enumerate(tfr_requests) if (r.get("status") or "").strip() == "Submitted"]

    if sort_mode == "type":
        key_fn = lambda it: (_norm_type(it[1]), _parse_date(it[1].get("request_date")), _tie_break(it[1]))
        pairs_declined.sort(key=key_fn)
        pairs_submitted.sort(key=key_fn)
        ordered_pairs = pairs_declined + pairs_submitted
    else:
        ordered_pairs = pairs_declined + pairs_submitted  # gi·ªØ th·ª© t·ª± JSON

    real_indices  = [i for i, _ in ordered_pairs]
    show_requests = [r.copy() for _, r in ordered_pairs]  # copy ƒë·ªÉ g√°n _rank

    # ===== T√≠nh th·ª© t·ª± trong ng√†y theo nh√≥m (ƒë·ªÉ t√¥ m√†u) =====
    if is_admin:
        # 1) Seed b·ªô ƒë·∫øm t·ª´ archive theo (date, group) -> count TRQ DUY NH·∫§T
        try:
            archive_all = safe_read_json(ARCHIVE_LOG) or []
        except Exception:
            archive_all = []

        rep2grp = _build_reportno_to_group_map()

        # 1) Seed theo TRQ duy nh·∫•t ƒë√£ c√≥ trong archive (Approved) theo (ng√†y, nh√≥m)
        base_seen = {}   # (date, group) -> set(TRQ)
        for a in archive_all:
            try:
                d0  = (a.get("request_date") or "").strip()
                rep = (a.get("report_no") or "").strip()
                g0  = rep2grp.get(rep, "")
                tid = (a.get("trq_id") or "").strip() or rep  # fallback report_no
                if not (d0 and g0 and tid):
                    continue
                key = (d0, _group_of(g0))
                s = base_seen.get(key)
                if s is None:
                    s = set()
                    base_seen[key] = s
                s.add(tid)
            except Exception:
                continue

        base_count = {k: len(v) for k, v in base_seen.items()}  # (date, group) -> s·ªë TRQ duy nh·∫•t ƒë√£ c√≥

        # 2) Duy·ªát c√°c request ƒëang hi·ªÉn th·ªã v√† ƒë√°nh s·ªë ti·∫øp THEO TRQ DUY NH·∫§T
        running_seen = {}   # (date, group) -> set(TRQ) ƒë√£ g·∫∑p trong batch hi·ªán t·∫°i
        running_rank = {}   # (date, group) -> {trq: rank}
        running_count = {}  # (date, group) -> next rank (kh·ªüi t·ª´ base_count)

        for r in show_requests:
            d  = (r.get("request_date") or "").strip()
            g  = _group_of(r.get("test_group") or r.get("type_of_test"))
            st = (r.get("status") or "").strip()
            key = (d, g)

            # chu·∫©n b·ªã c·∫•u tr√∫c
            if key not in running_seen:
                running_seen[key] = set()
                running_rank[key] = {}
                running_count[key] = base_count.get(key, 0)

            # m·∫∑c ƒë·ªãnh
            r["_rank_color"] = None
            r["_group_norm"] = g

            if st != "Submitted":
                # ch·ªâ t√¥ m√†u cho Submitted theo y√™u c·∫ßu
                continue

            trq = (r.get("trq_id") or "").strip()
            if not trq:
                # tr√°nh v·ª°: n·∫øu thi·∫øu TRQ th√¨ coi m·ªói d√≤ng 1 "TRQ"
                trq = f"__row_{id(r)}"

            if trq in running_rank[key]:
                # d√≤ng th·ª© 2, th·ª© 3... c·ªßa c√πng TRQ -> d√πng l·∫°i c√πng rank
                r["_rank_color"] = running_rank[key][trq]
            else:
                # l·∫ßn ƒë·∫ßu g·∫∑p TRQ n√†y trong batch
                # N·∫øu TRQ ƒë√£ n·∫±m trong seed (Approved c√πng ng√†y, nh√≥m), rank l·ªãch s·ª≠ c·ªßa n√≥ <= base_count
                if trq in base_seen.get(key, set()):
                    # set rank = base_count hi·ªán t·∫°i (ƒë·ªß ƒë·ªÉ x√°c ƒë·ªãnh qua/v∆∞·ª£t m·ªëc 5 hay ch∆∞a)
                    rank = base_count.get(key, 0)
                    running_rank[key][trq] = rank
                    running_seen[key].add(trq)
                    r["_rank_color"] = rank
                else:
                    # TRQ m·ªõi trong ng√†y+nh√≥m -> +1 theo TRQ (kh√¥ng theo s·ªë d√≤ng)
                    running_count[key] += 1
                    rank = running_count[key]
                    running_rank[key][trq] = rank
                    running_seen[key].add(trq)
                    r["_rank_color"] = rank

    # ====== NEW: Banner c·∫£nh b√°o t·ª´ DS: bi·∫øn trong config l√† `local_main` ======
    def _find_local_main_path():
        candidates = []
        try:
            if local_main:
                candidates.append(local_main)
        except Exception:
            pass

        # Fallback t√™n m·∫∑c ƒë·ªãnh
        candidates += [
            "local_main.xlsx",
            "local_main.xlsm",
            os.path.join("data", "local_main.xlsx"),
            os.path.join("data", "local_main.xlsm"),
        ]
        for p in candidates:
            try:
                if p and os.path.exists(p):
                    return p
            except Exception:
                continue
        return None  # kh√¥ng t√¨m th·∫•y

    REPORT_RE = re.compile(r"^\d{2}-\d{1,5}$")  # YY-serial (serial 1‚Äì5 ch·ªØ s·ªë)

    def _row_is_empty_except(df_row, ignore_cols=("QR Code", "Report #")):
        """R·ªóng khi t·∫•t c·∫£ c·ªôt (tr·ª´ ignore) ƒë·ªÅu tr·ªëng/NaN/chu·ªói tr·∫Øng."""
        for col in df_row.index:
            if col in ignore_cols:
                continue
            v = df_row[col]
            if pd.isna(v):
                continue
            if isinstance(v, str):
                if v.strip() == "":
                    continue
                return False
            # s·ªë/ki·ªÉu kh√°c -> coi l√† c√≥ th√¥ng tin
            return False
        return True

    def _extract_report_code(s):
        """Chu·∫©n h√≥a chu·ªói th√†nh d·∫°ng YY-serial n·∫øu c√≥ th·ªÉ."""
        if s is None:
            return None
        s = str(s).strip()
        if not s:
            return None
        if REPORT_RE.match(s):
            return s
        # normalize: 2025-123 -> 25-123 ; 25/0123 ; 25_12345 ...
        m = re.match(r"^(?:20)?(\d{2})[-_/ ]?(\d{1,5})$", s)
        if m:
            return f"{m.group(1)}-{int(m.group(2))}"
        return None

    def _pick_latest_from_codes(codes):
        """∆Øu ti√™n nƒÉm hi·ªán t·∫°i; c√πng nƒÉm ch·ªçn serial l·ªõn nh·∫•t; n·∫øu kh√¥ng c√≥, ch·ªçn (yy, serial) l·ªõn nh·∫•t."""
        if not codes:
            return None
        parsed = []
        for c in codes:
            try:
                yy, ser = c.split("-")
                parsed.append((int(yy), int(ser), c))
            except Exception:
                continue
        if not parsed:
            return None
        cur_yy = int(datetime.now().strftime("%y"))
        pool = [x for x in parsed if x[0] == cur_yy] or parsed
        pool.sort(key=lambda x: (x[0], x[1]), reverse=True)
        return pool[0][2]

    def _next_code_from_codes(codes):
        """Sinh m√£ k·∫ø ti·∫øp trong nƒÉm hi·ªán t·∫°i, serial = max+1 (kh√¥ng padding)."""
        cur_yy = int(datetime.now().strftime("%y"))
        max_ser = 0
        for c in codes or []:
            try:
                yy, ser = c.split("-")
                if int(yy) == cur_yy:
                    max_ser = max(max_ser, int(ser))
            except Exception:
                continue
        return f"{cur_yy}-{max_ser + 1 if max_ser >= 0 else 1}"

    def _build_ds_report_info():
        """ƒê·ªçc `local_main`, x√°c ƒë·ªãnh d√≤ng cu·ªëi c√≥ th√¥ng tin (d√≤ng sau r·ªóng), l·∫•y m√£ report v√† tr·∫£ chu·ªói c·∫£nh b√°o ti·∫øng Anh."""
        if pd is None:
            return None  # kh√¥ng c√≥ pandas -> b·ªè qua banner
        path = _find_local_main_path()
        if not path:
            return None
        try:
            df = pd.read_excel(path)
        except Exception:
            try:
                df = pd.read_excel(path, engine="openpyxl")
            except Exception:
                return None

        if df is None or df.empty:
            return None

        # C√°c t√™n c·ªôt kh·∫£ dƒ© cho Report #
        report_cols = ["Report #", "Report", "Report No", "Report no", "report_no", "report", "report_id", "Report_ID"]
        col_found = None
        for c in df.columns:
            if c in report_cols:
                col_found = c
                break

        # Thu th·∫≠p to√†n b·ªô code h·ª£p l·ªá ƒë·ªÉ d·ª± ph√≤ng
        all_codes = []
        if col_found:
            for v in df[col_found].tolist():
                c = _extract_report_code(v)
                if c:
                    all_codes.append(c)

        # T√¨m d√≤ng cu·ªëi c√≥ th√¥ng tin: row i not-empty & row i+1 empty
        last_idx = None
        n = len(df)
        for i in range(0, n):
            this_empty = _row_is_empty_except(df.iloc[i])
            next_empty = True
            if i + 1 < n:
                next_empty = _row_is_empty_except(df.iloc[i+1])
            if (not this_empty) and next_empty:
                last_idx = i

        newest_code = None
        if last_idx is not None and col_found:
            newest_code = _extract_report_code(df.iloc[last_idx].get(col_found))

        if not newest_code:
            # fallback: l·∫•y m√£ m·ªõi nh·∫•t t·ª´ to√†n b·ªô DS ho·∫∑c sinh m√£ k·∫ø ti·∫øp
            newest_code = _pick_latest_from_codes(all_codes) or _next_code_from_codes(all_codes)

        if newest_code:
            return f"‚ö† The latest report: {newest_code}"
        return None

    ds_report_info = _build_ds_report_info()

    return render_template(
        "tfr_request_status.html",
        requests=show_requests,
        is_admin=is_admin,
        real_indices=real_indices,
        viewer_name=viewer_name,
        viewer_emp_id=viewer_emp_id,
        ds_report_info=ds_report_info,  # <== truy·ªÅn cho template
    )

@app.route("/tfr_request_archive")
def tfr_request_archive():
    # 1) ƒê·ªçc archive
    archive = safe_read_json(ARCHIVE_LOG) or []

    # 2) Gom ·∫£nh t·ª´ to√†n b·ªô requests (gi·ªëng Status)
    try:
        tfr_all = safe_read_json(TFR_LOG_FILE) or []
    except Exception:
        tfr_all = []

    by_trq = {}
    for r in tfr_all:
        trq = (r.get("trq_id") or "").strip()
        if not trq:
            continue

        # Gom c√°c key ·∫£nh gi·ªëng Status
        merged = {
            "initial_img": r.get("initial_img") or r.get("initial_image") or r.get("initial_image_url") or "",
            "initial_images": list(r.get("initial_images") or []),
            "form_image": r.get("form_image") or "",
            "form_images": list(r.get("form_images") or []),
            "uploaded_images": list(r.get("uploaded_images") or []),
            "product_images": list(r.get("product_images") or []),
            "images": list(r.get("images") or []),
        }

        # N·∫øu 'images' r·ªóng th√¨ g·ªôp t·∫•t c·∫£ m·∫£ng c√≤n l·∫°i v√†o 'images'
        if not merged["images"]:
            tmp = []
            for k in ("initial_images", "form_images", "uploaded_images", "product_images"):
                vs = merged.get(k) or []
                if isinstance(vs, list):
                    tmp.extend(vs)
            # single-value
            for k in ("initial_img", "form_image"):
                v = merged.get(k)
                if isinstance(v, str) and v:
                    tmp.append(v)
            # kh·ª≠ tr√πng
            seen, out = set(), []
            for x in tmp:
                x = (x or "").strip()
                if x and x not in seen:
                    seen.add(x); out.append(x)
            merged["images"] = out

        by_trq[trq.upper()] = merged

    # Nh√©t ·∫£nh v√†o t·ª´ng record archive (kh√¥ng ph√° d·ªØ li·ªáu s·∫µn c√≥)
    for rec in archive:
        trq = (rec.get("trq_id") or "").strip().upper()
        if not trq:
            continue
        imgpack = by_trq.get(trq)
        if imgpack:
            rec.setdefault("initial_img", imgpack.get("initial_img", ""))
            rec.setdefault("initial_images", imgpack.get("initial_images", []))
            rec.setdefault("form_image", imgpack.get("form_image", ""))
            rec.setdefault("form_images", imgpack.get("form_images", []))
            rec.setdefault("uploaded_images", imgpack.get("uploaded_images", []))
            rec.setdefault("product_images", imgpack.get("product_images", []))
            if not rec.get("images"):
                rec["images"] = imgpack.get("images", [])

    # 3) Fallback: n·∫øu v·∫´n ch∆∞a c√≥ ·∫£nh, qu√©t static/TFR_INIT theo TRQ
    TFR_INIT_DIR = os.path.join('static', 'TFR_INIT')
    exts = ('.jpg', '.jpeg', '.png', '.webp', '.gif')

    def _find_init_by_trq(trq_id: str):
        if not trq_id:
            return []
        pattern = os.path.join(TFR_INIT_DIR, f"{trq_id}_*")
        out = []
        for p in glob.glob(pattern):
            if os.path.isfile(p) and os.path.splitext(p)[1].lower() in exts:
                # tr·∫£ v·ªÅ ƒë∆∞·ªùng d·∫´n t∆∞∆°ng ƒë·ªëi d∆∞·ªõi /static
                rel = os.path.relpath(p, start='static').replace('\\', '/')
                out.append(rel)
        out.sort()
        return out

    for rec in archive:
        if rec.get("initial_images") or rec.get("images") or rec.get("initial_img"):
            continue  # ƒë√£ c√≥ ·∫£nh t·ª´ JSON
        trq = (rec.get("trq_id") or "").strip()
        if not trq:
            continue
        picks = _find_init_by_trq(trq)
        if picks:
            rec["initial_images"] = picks
            rec["images"] = list(picks)

    # 4) ƒê·ªçc Excel -> t·∫°o rating_map, status_map, etd_map
    rating_map, status_map, etd_map = {}, {}, {}
    try:
        wb = safe_load_excel(local_main)  # d√πng helper s·∫µn c√≥
        ws = wb.active

        def find_col(*aliases):
            # th·ª≠ alias tr·ª±c ti·∫øp
            for name in aliases:
                c = get_col_idx(ws, name)
                if c:
                    return c
            # fallback: qu√©t header g·∫ßn-ƒë√∫ng
            def norm(s): return re.sub(r"[^a-z0-9#]+", "", str(s).strip().lower())
            alias_norm = {norm(a) for a in aliases}
            # ƒëo√°n m·ª•c ti√™u
            want = "report"
            if any("status" in a for a in alias_norm): want = "status"
            elif any("rating" in a for a in alias_norm): want = "rating"
            elif any("etd" in a for a in alias_norm) or any("expect" in a for a in alias_norm):
                want = "etd"
            targets = {
                "status": {"status"},
                "rating": {"rating", "result"},
                "etd": {"etd", "expecteddate", "deliverydate", "expecteddelivery", "expectedfinish", "completeddate"},
                "report": {"report#", "reportno", "report", "reportnumber"},
            }[want]
            for col in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=col).value
                if h is None:
                    continue
                h_norm = norm(h)
                if h_norm in targets or any(t in h_norm for t in targets):
                    return col
            return None

        col_report = find_col("Report #", "Report#", "Report No", "Report", "report #", "report no")
        col_rating = find_col("Rating", "RATING", "rating", "Result", "RESULT", "result")
        col_status = find_col("Status", "STATUS", "status", "Current Status", "current status")
        col_etd    = find_col("ETD", "etd", "Delivery Date", "delivery date",
                              "Expected Date", "expected date",
                              "Expected Delivery", "expected delivery",
                              "Expected Finish", "expected finish",
                              "Completed Date", "completed date")

        if col_report:

            for r in range(2, ws.max_row + 1):
                key_raw = ws.cell(row=r, column=col_report).value
                if key_raw is None:
                    continue
                key = str(key_raw).strip()
                if not key:
                    continue

                # Rating / Result
                if col_rating:
                    vr = ws.cell(row=r, column=col_rating).value
                    vr_str = "" if vr is None else str(vr).strip()
                    if vr_str:
                        rating_map[key] = vr_str
                        rating_map[key.lstrip("0")] = vr_str  # fallback kh√¥ng 0 ƒë·∫ßu

                # Status -> status_display
                if col_status:
                    vs = ws.cell(row=r, column=col_status).value
                    vs_str_orig = "" if vs is None else str(vs).strip()
                    vs_upper = vs_str_orig.upper()
                    if vs_upper in {"ACTIVE", "MUST", "DUE", "LATE"}:
                        disp = "ON PROGRESS"
                    elif vs_upper in {"COMPLETE", "DONE"}:
                        disp = "DONE"
                    else:
                        disp = vs_str_orig
                    status_map[key] = disp
                    status_map[key.lstrip("0")] = disp

                # ETD -> chu·∫©n ho√° text
                if col_etd:
                    ev = ws.cell(row=r, column=col_etd).value
                    if isinstance(ev, (datetime, date)):
                        etd_text = ev.strftime("%Y-%m-%d")
                    else:
                        etd_text = ("" if ev is None else str(ev)).strip()
                        # th·ª≠ parse nhanh v√†i format ph·ªï bi·∫øn -> 'YYYY-MM-DD'
                        if etd_text:
                            parsed = None
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                                try:
                                    parsed = datetime.strptime(etd_text, fmt).date()
                                    break
                                except Exception:
                                    pass
                            if parsed:
                                etd_text = parsed.strftime("%Y-%m-%d")
                    if etd_text:
                        etd_map[key] = etd_text
                        etd_map[key.lstrip("0")] = etd_text
    except Exception:
        rating_map, status_map, etd_map = {}, {}, {}

    # 5) G·∫Øn rating/status/etd/employee_id v√†o t·ª´ng record archive
    for rec in archive:
        rep = str(rec.get("report_no", "") or "").strip()
        if not rep:
            continue
        rec["rating"] = rating_map.get(rep, rec.get("rating", "") or "")
        rec["status_display"] = status_map.get(rep, rec.get("status_display", "") or "")
        if etd_map.get(rep):  # ETD ∆∞u ti√™n l·∫•y t·ª´ Excel theo y√™u c·∫ßu
            rec["etd"] = etd_map[rep]
        rec.setdefault("employee_id", rec.get("employee_id", "") or "")

    # 6) S·∫Øp x·∫øp: Report No m·ªõi n·∫±m tr√™n
    def report_sort_key(rec):
        s = str(rec.get("report_no", "") or "")
        nums = re.findall(r"\d+", s)
        num = int(nums[-1]) if nums else -1
        return (num, s)

    archive.sort(key=report_sort_key, reverse=True)

    # ƒë·∫£m b·∫£o m·ªói record khi ƒë∆∞a sang template ƒë·ªÅu c√≥ item_code
    for rec in archive:
        # n·∫øu tr∆∞·ªõc ƒë√¢y l∆∞u d∆∞·ªõi t√™n 'item' ho·∫∑c field kh√°c, r∆°i v·ªÅ lu√¥n
        rec["item_code"] = rec.get("item_code") or rec.get("item") or ""

    # 7) Render
    return render_template("tfr_request_archive.html", requests=archive)

@app.post("/save_etd")
def save_etd():
    if not request.is_json:
        return jsonify(success=False, message="Invalid request"), 400
    data = request.get_json()
    trq_id = (data.get("trq_id") or "").strip()
    etd = (data.get("etd") or "").strip()
    if not trq_id or not etd:
        return jsonify(success=False, message="Thi·∫øu trq_id ho·∫∑c etd"), 400

    try:
        current = _load_pending_locked()
        found = False
        for row in current:
            if (row.get("trq_id") or "").strip() == trq_id:
                row["etd"] = etd
                found = True
                break
        if not found:
            return jsonify(success=False, message="Kh√¥ng t√¨m th·∫•y TRQ-ID trong pending!"), 404
        _write_pending_locked(current)
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, message="L·ªói: " + str(e)), 500

@app.route('/run_export_excel', methods=['POST'])
def run_export_excel():
    if session.get('role') not in ['stl', 'superadmin']:
        return jsonify({'success': False, 'message': 'B·∫°n kh√¥ng c√≥ quy·ªÅn s·ª≠ d·ª•ng ch·ª©c nƒÉng n√†y!'}), 403
    try:
        # === G·ªåI TR·ª∞C TI·∫æP PYTHON CH·∫†Y SCRIPT ===
        python_path = r"C:\VFR\lab_update_app\.venv\Scripts\python.exe"  # d√πng python c·ªßa venv
        script_path = r"C:\VFR\lab_update_app\upload TRF.py"
        result = subprocess.run([python_path, script_path],
                                shell=False, capture_output=True, text=True, timeout=900)
        if result.returncode == 0:
            return jsonify({'success': True, 'message': 'ƒê√£ ch·∫°y xong TRF!', 'reload': True})
        else:
            # Log th√™m stderr n·∫øu l·ªói
            return jsonify({'success': False, 'message': f'L·ªói: {result.stderr}', 'reload': False})
    except Exception as e:
        return jsonify({'success': False, 'message': f'L·ªói: {e}', 'reload': False})

    
@app.route("/go_report")
def go_report():
    report = request.args.get("report", "").strip()
    if report:
        return redirect(url_for("update", report=report))
    return redirect(url_for("home"))

# Tr·∫£ ·∫£nh t·ªïng quan/c√¢n n·∫∑ng
@app.route('/images/<report>/<filename>')
def serve_general_img(report, filename):
    folder = os.path.join(UPLOAD_FOLDER, report)
    return send_from_directory(folder, filename)

@app.route("/delete_image/<report>/<imgfile>", methods=["POST"])
def delete_image_main(report, imgfile):
    img_path = os.path.join(UPLOAD_FOLDER, report, imgfile)
    # Th√™m try-except ƒë·ªÉ tr√°nh l·ªói race condition khi x√≥a c√πng l√∫c
    try:
        if os.path.exists(img_path):
            os.remove(img_path)
    except Exception as e:
        print(f"L·ªói khi x√≥a ·∫£nh: {img_path} - {e}")
    return redirect(url_for('update', report=report))

@app.route("/delete_test_group_image/<report>/<group>/<key>/<imgfile>", methods=["POST"])
def delete_test_group_image(report, group, key, imgfile):
    img_path = os.path.join(UPLOAD_FOLDER, report, imgfile)
    try:
        if os.path.exists(img_path):
            os.remove(img_path)
    except Exception as e:
        print(f"L·ªói khi x√≥a ·∫£nh: {img_path} - {e}")
    return redirect(url_for("test_group_item_dynamic", report=report, group=group, test_key=key))

@app.route("/logout")
def logout():
    for k in ("auth_ok", "staff_id", "user_type", "role", "last_test_type", "last_test_code"):
        session.pop(k, None)
    return "<h3 style='text-align:center;margin-top:80px;'>ƒê√£ ƒëƒÉng xu·∫•t!<br><a href='/' style='color:#4d665c;'>V·ªÅ trang ch·ªçn s·∫£n ph·∫©m</a></h3>"

@app.route("/update", methods=["GET", "POST"])
def update():
    report = request.args.get("report")
    if not report:
        return redirect("/")

    item_id, row_idx = None, None
    lines = []
    message = ""
    is_logged_in = session.get("auth_ok", False)
    valid = False

    # ==== helpers cho comment file ====
    def _ensure_dir(p):
        os.makedirs(p, exist_ok=True)

    def _parse_kv_file(path):
        """ƒê·ªçc file comment_{group}.txt -> dict {key: value} (key ƒë∆∞·ª£c chu·∫©n h√≥a lowercase)."""
        data = {}
        if not os.path.exists(path):
            return data
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                m = re.split(r"\s*[:=]\s*", line, maxsplit=1)
                if len(m) == 2:
                    k, v = m[0].strip(), m[1].strip()
                    if k:
                        data[k.lower()] = v
        return data

    def _upsert_kv_file(path, kv: dict):
        """Ghi/ghi ƒë√® c√°c kh√≥a trong file comment_{group}.txt, gi·ªØ nguy√™n c√°c d√≤ng kh√°c."""
        lines = []
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()

        # map key -> ƒë√£ thay ch∆∞a
        replaced = {k: False for k in kv.keys()}
        new_lines = []
        for line in lines:
            line_stripped = line.strip()
            done = False
            for k, v in kv.items():
                if re.match(rf"^\s*{re.escape(k)}\s*[:=]", line_stripped):
                    new_lines.append(f"{k}: {v}\n")
                    replaced[k] = True
                    done = True
                    break
            if not done:
                new_lines.append(line)

        # append c√°c key ch∆∞a c√≥
        for k, v in kv.items():
            if not replaced[k]:
                new_lines.append(f"{k}: {v}\n")

        with open(path, "w", encoding="utf-8") as f:
            f.writelines(new_lines)

    # ==== ƒë·ªçc Excel, t√¨m d√≤ng ====
    try:
        wb = safe_load_excel(local_main)
        ws = wb.active
        report_col = get_col_idx(ws, "report#") or get_col_idx(ws, "report")
        if report_col is None:
            return "‚ùå Kh√¥ng t√¨m th·∫•y c·ªôt REPORT# ho·∫∑c REPORT trong file Excel!", 500

        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=report_col).value
            if v and str(v).strip() == str(report):
                row_idx = row
                break
        if row_idx is None:
            return f"‚ùå Kh√¥ng t√¨m th·∫•y m√£ report {report} trong file Excel!", 404

        valid = True

        # --- l·∫•y rating_value ƒë·ªÉ ki·ªÉm so√°t UI ---
        rating_value = ""
        rating_col_idx = get_col_idx(ws, "rating")
        if rating_col_idx:
            rv = ws.cell(row=row_idx, column=rating_col_idx).value
            rating_value = (str(rv).strip() if rv not in (None, "") else "")

        # --- build lines hi·ªÉn th·ªã ---
        if not is_logged_in:
            summary_keys = [
                ('TRQ ID', 'TRQ ID'),
                ('report#', 'REPORT#'),
                ('item#', 'ITEM#'),
                ('type of', 'TYPE OF'),
                ('furniture testing', 'FURNITURE TESTING'),
                ('remark', 'REMARK'),
                ('qa comment', 'QA COMMENT'),
                ('etd', 'ETD'),
                ('rating', 'RATING')
            ]
            for key, label in summary_keys:
                idx_col = get_col_idx(ws, key)
                value = ws.cell(row=row_idx, column=idx_col).value if idx_col else ""
                show_value = str(value).strip() if value not in ("", None) else ""
                lines.append((label, show_value))
        else:
            for col in range(1, ws.max_column):
                label = ws.cell(row=1, column=col).value
                value = ws.cell(row=row_idx, column=col).value
                if label and value not in (None, ""):
                    lines.append((str(label).upper(), str(value)))

    except Exception as e:
        print("L·ªói khi ƒë·ªçc file excel:", e)
        print(traceback.format_exc())
        return f"L·ªói khi x·ª≠ l√Ω file: {e}", 500

    # ==== l·∫•y s·∫µn Sample Info (∆∞u ti√™n nh√≥m test g·∫ßn nh·∫•t, fallback main) ====
    report_folder = os.path.join(UPLOAD_FOLDER, str(report))
    _ensure_dir(report_folder)

    group_pref = get_last_test_code(report, default="main")

    def _read_notes(gname):
        return _parse_kv_file(os.path.join(report_folder, f"comment_{gname}.txt"))

    sample_notes = _read_notes(group_pref)
    if not sample_notes or (not sample_notes.get("sample_weight") and not sample_notes.get("sample_size")):
        base_notes = _read_notes("main")
        if base_notes:
            for k, v in base_notes.items():
                sample_notes.setdefault(k, v)

    # --- Chu·∫©n h√≥a sample_weight: l·∫•y s·ªë thu·∫ßn ƒë·ªÉ render v√†o <input type="number">
    def _extract_number_str(s):
        if not s:
            return ""
        s = str(s)
        m = re.search(r'[-+]?\d+(?:[.,]\d+)?', s)  # l·∫•y s·ªë ƒë·∫ßu ti√™n
        if not m:
            return ""
        return m.group(0).replace(",", ".")  # 12,3 -> 12.3

    sample_weight_val_raw = sample_notes.get("sample_weight", "")
    sample_weight_val = _extract_number_str(sample_weight_val_raw)

    # --- Parse sample_size "L x W x H" -> ba tr∆∞·ªùng ƒë·ªÉ bind v√†o input
    sample_size_val = sample_notes.get("sample_size", "")

    def _parse_size_triplet(s):
        if not s:
            return "", "", ""
        s = str(s)
        m = re.search(
            r'([0-9]+(?:[.,][0-9]+)?)\s*[x√ó*]\s*([0-9]+(?:[.,][0-9]+)?)\s*[x√ó*]\s*([0-9]+(?:[.,][0-9]+)?)',
            s, re.I
        )
        if not m:
            return "", "", ""
        a, b, c = m.group(1), m.group(2), m.group(3)
        return a.replace(",", "."), b.replace(",", "."), c.replace(",", ".")

    size_length_val, size_width_val, size_height_val = _parse_size_triplet(sample_size_val)

    # --- n·∫øu ch∆∞a login: x·ª≠ l√Ω login form ---
    if not is_logged_in:
        if request.method == "POST" and request.form.get("action") == "login":
            password_input = request.form.get("password")
            if login(password_input):
                return redirect(url_for("update", report=report))
            else:
                message = "Sai m·∫≠t kh·∫©u!"

        return render_template(
            "info_line.html",
            lines=lines,
            message=message,
            logout=False,
            img_overview=get_img_urls(report, "overview"),
            img_weight=[],
            show_hint=True,
            show_func=False,
            report_id=report,
            test_groups=TEST_GROUPS,
            rating_value=rating_value,
            sample_weight=sample_weight_val,
            sample_size=sample_size_val,
            show_line_test_done_notice=False,  # ·∫©n ph·∫ßn notice khi ch∆∞a login
            so_gio_test=SO_GIO_TEST,
        )

    # ==== ƒê√É ƒêƒÇNG NH·∫¨P: X·ª¨ L√ù POST ====
    if request.method == "POST":
        action = request.form.get("action")

        # --- Upload overview images ---
        if action == "upload_overview":
            files = request.files.getlist('overview_imgs')
            folder = os.path.join(UPLOAD_FOLDER, report)
            os.makedirs(folder, exist_ok=True)
            for i, file in enumerate(files):
                if file and allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"overview_{int(datetime.now().timestamp())}_{i}.{ext}"
                    file.save(os.path.join(folder, filename))
            return redirect(url_for("update", report=report))

        # --- Upload weight images ---
        elif action == "upload_weight":
            files = request.files.getlist('weight_imgs')
            folder = os.path.join(UPLOAD_FOLDER, report)
            os.makedirs(folder, exist_ok=True)
            for i, file in enumerate(files):
                if file and allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"weight_{int(datetime.now().timestamp())}_{i}.{ext}"
                    file.save(os.path.join(folder, filename))
            return redirect(url_for("update", report=report))

        # --- L∆ØU SAMPLE INFO v√†o comment_{group}.txt (KH√îNG ghi Excel) ---
        elif action == "save_sample_info":
            weight = request.form.get("sample_weight", "").strip()
            size_length = request.form.get("size_length", "").strip()
            size_width  = request.form.get("size_width", "").strip()
            size_height = request.form.get("size_height", "").strip()

            size_str = ""
            if size_length and size_width and size_height:
                size_str = f"{size_length} x {size_width} x {size_height}"

            group_code = get_last_test_code(report, default="main")
            comment_file = os.path.join(UPLOAD_FOLDER, str(report), f"comment_{group_code}.txt")
            os.makedirs(os.path.dirname(comment_file), exist_ok=True)

            updates = {}
            if weight:
                # L∆∞u nguy√™n s·ªë th·∫≠p ph√¢n, th√™m " kg" ƒë·ªÉ ng∆∞·ªùi ƒë·ªçc d·ªÖ hi·ªÉu
                # (n·∫øu kh√¥ng mu·ªën 'kg' th√¨ b·ªè ƒëi)
                updates["sample_weight"] = f"{weight} kg"
            if size_str:
                updates["sample_size"] = size_str

            if updates:
                _upsert_kv_file(comment_file, updates)

            # PRG: ƒë·∫£m b·∫£o reload trang s·∫Ω hi·ªÉn th·ªã gi√° tr·ªã m·ªõi trong input
            return redirect(url_for("update", report=report))
        
        # --- ƒê√°nh d·∫•u "testing" ---
        elif valid and action == "testing":
            wb = safe_load_excel(local_main)
            ws = wb.active
            test_date_col = get_col_idx(ws, "test date")
            rating_col = get_col_idx(ws, "rating")
            vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')
            now = datetime.now(vn_tz).strftime("%d/%m/%Y %H:%M").upper()
            if test_date_col:
                ws.cell(row=row_idx, column=test_date_col).value = now
            if rating_col:
                ws.cell(row=row_idx, column=rating_col).value = "PENDING"
            safe_save_excel(wb, local_main)
            message = f"ƒê√£ ghi th·ªùi gian ki·ªÉm tra v√† c·∫≠p nh·∫≠t tr·∫°ng th√°i PENDING cho {report}!"

        # --- ƒê√°nh d·∫•u "test_done" ---
        elif valid and action == "test_done":
            wb = safe_load_excel(local_main)
            ws = wb.active
            complete_col = get_col_idx(ws, "complete date")
            vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')
            now = datetime.now(vn_tz).strftime("%d/%m/%Y %H:%M").upper()
            if complete_col:
                ws.cell(row=row_idx, column=complete_col).value = now
            safe_save_excel(wb, local_main)
            message = f"ƒê√£ ghi ho√†n th√†nh test cho {report}!"

        # --- Rating PASS/FAIL/DATA (ghi Excel nh∆∞ c≈©) ---
        elif valid and action and action.startswith("rating_"):
            print("==> ƒêANG X·ª¨ L√ù RATING:", action, "CHO REPORT", report)
            value = action.replace("rating_", "").upper()

            wb = safe_load_excel(local_main)
            ws = wb.active

            rating_col = get_col_idx(ws, "rating")
            status_col = get_col_idx(ws, "status")
            if rating_col:
                ws.cell(row=row_idx, column=rating_col).value = value

            # (‚Ä¶ gi·ªØ nguy√™n ph·∫ßn c√≤n l·∫°i x·ª≠ l√Ω Teams / copy completed / log nh∆∞ code g·ªëc ‚Ä¶)
            # --- L·∫§Y LO·∫†I TEST G·∫¶N NH·∫§T ---
            group_code = get_last_test_code(report)
            group_title = get_group_title(group_code) if group_code else None

            if not group_code:
                last_test_type = get_last_test_type(report)
                if last_test_type:
                    for g_id, g_name in TEST_GROUPS:
                        if g_name == last_test_type:
                            group_code = g_id
                            group_title = g_name
                            break

            if not group_code:
                type_of_col = get_col_idx(ws, "type of")
                type_of = ws.cell(row=row_idx, column=type_of_col).value if type_of_col else ""
                # n·∫øu c√≥ b·∫£ng map chu·∫©n h√≥a ri√™ng th√¨ √°p d·ª•ng ·ªü ƒë√¢y thay v√¨ replace space.
                if type_of:
                    t_key = str(type_of).strip().lower()
                    mapped_code = None
                    try:
                        from config import TYPE_OF_MAP  # optional mapping: {raw_lower: group_code}
                    except Exception:
                        TYPE_OF_MAP = {}
                    if isinstance(TYPE_OF_MAP, dict):
                        mapped_code = TYPE_OF_MAP.get(t_key) or TYPE_OF_MAP.get(t_key.replace(" ", "_"))
                    if not mapped_code:
                        for g_id, g_name in TEST_GROUPS:
                            if t_key == str(g_name).strip().lower():
                                mapped_code = g_id
                                break
                    group_code = mapped_code or t_key.replace(" ", "_")
                else:
                    group_code = None
                group_title = get_group_title(group_code) or (type_of or "")

            country_col = get_col_idx(ws, "country of destination")
            furniture_testing_col = get_col_idx(ws, "furniture testing")
            country = ws.cell(row=row_idx, column=country_col).value if country_col else ""
            furniture_testing = ws.cell(row=row_idx, column=furniture_testing_col).value if furniture_testing_col else ""

            trq_col = get_col_idx(ws, "trq id")
            item_col = get_col_idx(ws, "item#")
            desc_col = get_col_idx(ws, "item name/ description")
            requestor_col = get_col_idx(ws, "submiter in") or get_col_idx(ws, "submitter in charge") or get_col_idx(ws, "requestor")

            trq = ws.cell(row=row_idx, column=trq_col).value if trq_col else ""
            item = ws.cell(row=row_idx, column=item_col).value if item_col else ""
            desc = ws.cell(row=row_idx, column=desc_col).value if desc_col else ""
            requestor = ws.cell(row=row_idx, column=requestor_col).value if requestor_col else ""

            report_url = f"{request.url_root.rstrip('/')}/update?report={report}"
            staff_id = session.get("staff_id", "Kh√¥ng r√µ")

            teams_msg = None
            if value == "PASS":
                teams_msg = (
                    f"‚úÖ **PASS** {group_title}\n"
                    f"- TRQ: {trq}\n"
                    f"- Report#: {report}\n"
                    f"- Item#: {item}\n"
                    f"- Description: {desc}\n"
                    f"- Group: {group_title}\n"
                    f"- Country of Destination: {country}\n"
                    f"- Furniture Testing: {furniture_testing}\n"
                    f"- Requestor: {requestor}\n"
                    f"- Nh√¢n vi√™n thao t√°c: {staff_id}\n"
                    f"Chi ti·∫øt: {report_url}"
                )
            elif value in ["FAIL", "DATA"]:
                report_folder = os.path.join(UPLOAD_FOLDER, str(report))
                status_file = os.path.join(report_folder, f"status_{group_code}.txt")
                comment_file = os.path.join(report_folder, f"comment_{group_code}.txt")
                group_titles = TEST_GROUP_TITLES.get(group_code, {})
                status_notes = load_group_notes(status_file)
                comment_notes = load_group_notes(comment_file)
                group_fails = []
                for key, title in group_titles.items():
                    status_val = status_notes.get(key)
                    if status_val == "FAIL":
                        comment_val = comment_notes.get(key, "")
                        group_fails.append(f"- {title['short']}: {comment_val if comment_val else '(Kh√¥ng c√≥ ghi ch√∫)'}")
                status_text = "‚ùå **FAIL**" if value == "FAIL" else "‚ö†Ô∏è **DATA**"
                if group_fails:
                    teams_msg = (
                        f"{status_text} {group_title}\n"
                        f"- TRQ: {trq}\n"
                        f"- Report#: {report}\n"
                        f"- Item#: {item}\n"
                        f"- Description: {desc}\n"
                        f"- Group: {group_title}\n"
                        f"- Country of Destination: {country}\n"
                        f"- Furniture Testing: {furniture_testing}\n"
                        f"- Requestor: {requestor}\n"
                        f"- Nh√¢n vi√™n thao t√°c: {staff_id}\n"
                        f"- C√°c m·ª•c FAIL:\n"
                        + "\n".join(group_fails)
                        + f"\nChi ti·∫øt: {report_url}"
                    )
                else:
                    teams_msg = (
                        f"{status_text} {group_title}\n"
                        f"- TRQ: {trq}\n"
                        f"- Report#: {report}\n"
                        f"- Item#: {item}\n"
                        f"- Description: {desc}\n"
                        f"- Group: {group_title}\n"
                        f"- Country of Destination: {country}\n"
                        f"- Furniture Testing: {furniture_testing}\n"
                        f"- Requestor: {requestor}\n"
                        f"- Nh√¢n vi√™n thao t√°c: {staff_id}\n"
                        f"- Kh√¥ng c√≥ m·ª•c n√†o FAIL trong nh√≥m n√†y."
                        + f"\nChi ti·∫øt: {report_url}"
                    )

            if teams_msg:
                send_teams_message(TEAMS_WEBHOOK_URL_RATE, teams_msg)

            if status_col:
                ws.cell(row=row_idx, column=status_col).value = "COMPLETE"
                fill_complete = PatternFill("solid", fgColor="BFBFBF")
                for col in range(2, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).fill = fill_complete

            # Copy sang completed file (gi·ªØ nguy√™n logic c≈©)
            if os.path.exists(local_complete):
                wb_c = safe_load_excel(local_complete)
                ws_c = wb_c.active
            else:
                wb_c = Workbook()
                ws_c = wb_c.active
                for col in range(1, ws.max_column + 1):
                    from_cell = ws.cell(row=1, column=col)
                    to_cell = ws_c.cell(row=1, column=col)
                    to_cell.value = from_cell.value
                    if from_cell.font:        to_cell.font = from_cell.font.copy()
                    if from_cell.border:      to_cell.border = from_cell.border.copy()
                    if from_cell.fill:        to_cell.fill = from_cell.fill.copy()
                    if from_cell.protection:  to_cell.protection = from_cell.protection.copy()
                    if from_cell.alignment:   to_cell.alignment = from_cell.alignment.copy()
                    to_cell.number_format = from_cell.number_format
                    col_letter = from_cell.column_letter
                    ws_c.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
                ws_c.row_dimensions[1].height = ws.row_dimensions[1].height
                safe_save_excel(wb_c, local_complete)

            report_idx_in_c = get_col_idx(ws_c, "report#") or get_col_idx(ws_c, "report") or 2
            found_row = None
            for r in range(2, ws_c.max_row + 1):
                v = ws_c.cell(row=r, column=report_idx_in_c).value
                if v and str(v).strip().upper() == str(report).upper():
                    found_row = r
                    break

            if found_row:
                copy_row_with_style(ws, ws_c, row_idx, to_row=found_row)
            else:
                copy_row_with_style(ws, ws_c, row_idx)

            safe_save_excel(wb_c, local_complete)
            safe_save_excel(wb, local_main)

            # Log ho√†n th√†nh
            type_of_col = get_col_idx(ws, "type of")
            type_of = ws.cell(row=row_idx, column=type_of_col).value if type_of_col else ""
            vn_tz = pytz.timezone("Asia/Ho_Chi_Minh")
            now = datetime.now(vn_tz)
            tval = now.hour * 60 + now.minute
            office_start = 8 * 60
            office_end = 16 * 60 + 45
            ot_end = 23 * 60 + 59
            if office_start <= tval < office_end:
                ca = "office"
            elif office_end <= tval <= ot_end:
                ca = "ot"
            else:
                ca = ""
            employee_id = session.get("staff_id", "")
            log_report_complete(report, type_of, ca, employee_id)

            message = f"ƒê√£ c·∫≠p nh·∫≠t ƒë√°nh gi√°: <b>{value}</b> cho {report}!"
            check_and_reset_counter()
            update_counter()

    # === L·∫•y lo·∫°i test g·∫ßn nh·∫•t (last_test_type) ===
    last_test_type = get_last_test_type(report)

    # === Ki·ªÉm tra ƒë√£ ƒë·ªß s·ªë gi·ªù line test ch∆∞a ===
    elapsed = get_line_test_elapsed(report)
    show_line_test_done = elapsed is not None and elapsed >= SO_GIO_TEST

    # === Ki·ªÉm tra ƒë√£ c√≥ ·∫£nh after ch∆∞a ===
    folder = os.path.join(UPLOAD_FOLDER, str(report))
    imgs_after = []
    after_tag = "line_after"
    if os.path.exists(folder):
        for f in sorted(os.listdir(folder)):
            if allowed_file(f) and f.startswith(after_tag):
                imgs_after.append(f"/images/{report}/{f}")
    has_after_img = len(imgs_after) > 0

    show_line_test_done_notice = show_line_test_done and not has_after_img

    return render_template(
        "info_line.html",
        lines=lines,
        message=message,
        logout=True,
        img_overview=get_img_urls(report, "overview"),
        img_weight=get_img_urls(report, "weight"),
        show_hint=False,
        show_func=True,
        report_id=report,
        test_groups=TEST_GROUPS,
        last_test_type=last_test_type,
        so_gio_test=SO_GIO_TEST,
        show_line_test_done_notice=show_line_test_done_notice,
        rating_value=rating_value,
        # Sample info
        sample_weight=sample_weight_val,
        sample_size=sample_size_val,
        size_length=size_length_val,
        size_width=size_width_val,
        size_height=size_height_val,
    )

def _has_images(report_folder: str, group: str, key: str, is_hotcold_like: bool) -> bool:
    if not os.path.exists(report_folder):
        return False
    try:
        files = os.listdir(report_folder)
    except Exception:
        return False

    if is_hotcold_like:
        # ch·∫•p nh·∫≠n t√™n c√≥/kh√¥ng k√®m group sau before/after
        prefixes = (
            f"{key}_before_{group}",
            f"{key}_after_{group}",
            f"{key}_before_",
            f"{key}_after_",
        )
        return any(allowed_file(fn) and fn.startswith(prefixes) for fn in files)
    else:
        pref = f"test_{group}_{key}_"
        return any(allowed_file(fn) and fn.startswith(pref) for fn in files)

@app.route("/download_images/<report>")
def download_all_images(report):
    base_dir = os.path.join("images", report)

    if not os.path.exists(base_dir):
        return f"Kh√¥ng t√¨m th·∫•y th∆∞ m·ª•c h√¨nh cho report {report}", 404

    # T·∫°o ZIP trong RAM
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(base_dir):
            for file in files:
                if file.lower().endswith((".jpg", ".jpeg", ".png", ".webp")):
                    full_path = os.path.join(root, file)

                    # ƒê∆∞·ªùng d·∫´n b√™n trong ZIP gi·ªØ nguy√™n c·∫•u tr√∫c
                    arcname = os.path.relpath(full_path, base_dir)

                    zipf.write(full_path, arcname)

    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=f"{report}_images.zip",
        mimetype="application/zip"
    )
# >>> ADD
@app.get('/api/report/detect')
def api_report_detect():
    report = (request.args.get('report') or '').strip()
    if not report:
        return jsonify({'ok': False, 'error': 'missing report'}), 400

    # TODO: vi·∫øt logic t·ª± ƒëo√°n n·∫øu mu·ªën (t·ª´ last_test_type, file/·∫£nh, v.v.)
    # T·∫°m th·ªùi: lu√¥n y√™u c·∫ßu ng∆∞·ªùi d√πng ch·ªçn
    return jsonify({'ok': False})

# ==== ROUTE ====
@app.route("/api/report/create")
def api_report_create():
    """
    Create .docx report for any type using auto-mapping to TEST_GROUP_TITLES.
    Templates are placed next to app.py.
    """
    report_id = (request.args.get("report") or "").strip()
    rtype     = (request.args.get("type") or "").strip().lower()

    tpl_name = TEMPLATE_MAP.get(rtype)
    if not tpl_name:
        return jsonify({"ok": False, "error": f"Unknown type: {rtype}"}), 400

    tpl_path = os.path.join(os.path.dirname(__file__), tpl_name)
    if not os.path.exists(tpl_path):
        return jsonify({"ok": False, "error": f"Template not found: {tpl_name}"}), 404

    # Use the same Excel as for bed (adjust if you later separate per-type files)
    excel_path = os.path.join(local_main, "ds san pham test voi qr.xlsx")

    try:
        # Generic fill for ALL types with auto mapping
        generated_by = session.get('staff_id') or session.get('user_type') or "VFR User"

        bio = fill_cover_from_excel_generic(
            template_docx_path=tpl_path,
            excel_path_or_name=excel_path,
            report_id=report_id,
            template_key=rtype,
            generated_by=generated_by,   # NEW
        )
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"{report_id or 'report'}_{rtype}.docx",
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
    except Exception as ex:
        # Fallback: return raw template to not block users
        print("Autofill error:", ex)
        return send_file(
            tpl_path,
            as_attachment=True,
            download_name=f"{report_id or 'report'}_{rtype}_TEMPLATE.docx"
        )

# --- THAY TH·∫æ H·∫≤N h√†m test_group_page ---
@app.route("/test_group/<report>/<group>", methods=["GET", "POST"])
def test_group_page(report, group):
    # L∆∞u context g·∫ßn nh·∫•t
    set_last_test(report, group)

    group_titles = TEST_GROUP_TITLES.get(group)
    if not group_titles:
        return "Nh√≥m ki·ªÉm tra kh√¥ng t·ªìn t·∫°i!", 404

    report_folder = os.path.join(UPLOAD_FOLDER, str(report))
    os.makedirs(report_folder, exist_ok=True)

    # N∆°i c√°c trang hot/cold ghi v√†o:
    status_file  = os.path.join(report_folder, f"status_{group}.txt")
    comment_file = os.path.join(report_folder, f"comment_{group}.txt")
    all_status   = load_group_notes(status_file)    # {key -> PASS/FAIL/N/A/DATA...}
    all_comment  = load_group_notes(comment_file)   # {key -> comment string}

    test_status = {}

    for key in group_titles:
        # 1) L·∫•y t·ª´ file t·ªïng (status_{group}.txt / comment_{group}.txt)
        st = all_status.get(key)
        cm = all_comment.get(key)

        # 2) N·∫øu ch∆∞a c√≥, ƒë·ªçc fallback theo c·∫£ 2 pattern file ri√™ng:
        #    - M·ªõi:  status_{group}_{key}.txt / comment_{group}_{key}.txt
        #    - C≈©:   {key}_{group}_status.txt / {key}_{group}_comment.txt
        if not st:
            for st_path in [
                os.path.join(report_folder, f"status_{group}_{key}.txt"),
                os.path.join(report_folder, f"{key}_{group}_status.txt"),
            ]:
                if os.path.exists(st_path):
                    try:
                        v = (safe_read_text(st_path) or "").strip()
                        if v:
                            st = v
                            break
                    except Exception:
                        pass

        if not cm:
            for cm_path in [
                os.path.join(report_folder, f"comment_{group}_{key}.txt"),
                os.path.join(report_folder, f"{key}_{group}_comment.txt"),
            ]:
                if os.path.exists(cm_path):
                    try:
                        v = (safe_read_text(cm_path) or "").strip()
                        if v:
                            cm = v
                            break
                    except Exception:
                        pass

        # 3) Ki·ªÉm tra ·∫£nh (ƒë√£ ok cho c·∫£ hot_cold & th∆∞·ªùng)
        has_img = _has_images(report_folder, group, key, key in HOTCOLD_LIKE)

        test_status[key] = {"status": st, "comment": cm, "has_img": has_img}
    # Ri√™ng t·ªß US (n·∫øu c√≥)
    f2057_status = {}
    if group == "tu_us":
        for fkey in F2057_TEST_TITLES:
            f2057_status[fkey] = get_group_test_status(report, group, fkey)

    return render_template(
        "test_group_menu.html",
        report=report,
        group=group,
        test_titles=group_titles,
        test_status=test_status,
        F2057_TEST_TITLES=F2057_TEST_TITLES,
        f2057_status=f2057_status
    )

@app.route('/test_group/<report>/<group>/<test_key>', methods=['GET', 'POST'])
def test_group_item_dynamic(report, group, test_key):
    # L∆∞u l·∫°i lo·∫°i test g·∫ßn nh·∫•t
    set_last_test(report, group)

    # Hot/Cold chuy·ªÉn sang route ri√™ng
    if test_key in HOTCOLD_LIKE and group in INDOOR_GROUPS:
        return redirect(url_for("hot_cold_test", report=report, group=group, test_key=test_key))

    # Ki·ªÉm tra t·ªìn t·∫°i test key
    group_titles = TEST_GROUP_TITLES.get(group)
    if not group_titles or test_key not in group_titles:
        return "M·ª•c ki·ªÉm tra kh√¥ng t·ªìn t·∫°i!", 404
    title = group_titles[test_key]

    # Th∆∞ m·ª•c theo report
    report_folder = os.path.join(UPLOAD_FOLDER, str(report))
    os.makedirs(report_folder, exist_ok=True)
    status_file = os.path.join(report_folder, f"status_{group}.txt")
    comment_file = os.path.join(report_folder, f"comment_{group}.txt")

    # ƒê·∫∑c th√π nh√≥m TRANSIT
    is_rh_np = (group == "transit_RH_np")
    is_drop = (is_drop_test(title) if group.startswith("transit") else False) or (group == "transit_181_lt68" and test_key == "step4")
    is_impact = is_impact_test(title) if group.startswith("transit") else False
    is_rot = is_rotational_test(title) if group.startswith("transit") else False
    rh_impact_zones = RH_IMPACT_ZONES if is_rh_np and test_key == "step3" else []
    rh_vib_zones = RH_VIB_ZONES if is_rh_np and test_key == "step4" else []
    rh_second_impact_zones = RH_SECOND_IMPACT_ZONES if is_rh_np and test_key == "step5" else []
    rh_step12_zones = RH_STEP12_ZONES if is_rh_np and test_key == "step12" else []

    # ------------- AJAX IMAGE UPLOAD/DELETE (JSON RESPONSE) -------------
    if request.method == "POST" and request.headers.get("X-Requested-With") == "XMLHttpRequest":
        # Ki·ªÉm tra c√°c v√πng ·∫£nh ƒë·∫∑c bi·ªát RH
        imgs = {}

        # ========== GT68 FACE ZONES (ch·ªâ x·ª≠ l√Ω GT68 ·ªü ƒë√¢y) ==========
        if group == "transit_181_gt68" and test_key == "step4":
            for idx, zone in enumerate(GT68_FACE_ZONES):
                files = request.files.getlist(f'gt68_face_img_{zone}')
                if files:
                    imgs[str(idx)] = []  # FIX: ƒë·ªìng b·ªô key "0".."5" ƒë·ªÉ FE ƒë·ªçc data.imgs[zone]
                    for file in files:
                        if file and allowed_file(file.filename):
                            ext = file.filename.rsplit('.', 1)[-1].lower()
                            prefix = f"test_{group}_{test_key}_gt68_face_{zone}_"
                            nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                            next_num = max(nums, default=0) + 1
                            fname = f"{prefix}{next_num}.{ext}"
                            file.save(os.path.join(report_folder, fname))
                            imgs[str(idx)].append(f"/images/{report}/{fname}")

        # ========== RH Impact zones (t√°ch ra ngo√†i nh√°nh GT68) ==========
        # FIX: c√°c kh·ªëi RH/Drop/Impact/Rot KH√îNG c√≤n l·ªìng trong nh√°nh GT68
        for zone, _ in rh_impact_zones:
            files = request.files.getlist(f'rh_impact_img_{zone}')
            if files:
                imgs.setdefault(zone, [])
                for file in files:
                    if file and allowed_file(file.filename):
                        ext = file.filename.rsplit('.', 1)[-1].lower()
                        prefix = f"test_{group}_{test_key}_{zone}_"
                        nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                        next_num = max(nums, default=0) + 1
                        fname = f"{prefix}{next_num}.{ext}"
                        file.save(os.path.join(report_folder, fname))
                        imgs[zone].append(f"/images/{report}/{fname}")

        # ========== RH Vib zones ==========
        for zone, _ in rh_vib_zones:
            files = request.files.getlist(f'rh_vib_img_{zone}')
            if files:
                imgs.setdefault(zone, [])
                for file in files:
                    if file and allowed_file(file.filename):
                        ext = file.filename.rsplit('.', 1)[-1].lower()
                        prefix = f"test_{group}_{test_key}_{zone}_"
                        nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                        next_num = max(nums, default=0) + 1
                        fname = f"{prefix}{next_num}.{ext}"
                        file.save(os.path.join(report_folder, fname))
                        imgs[zone].append(f"/images/{report}/{fname}")

        # ========== RH Second impact zones ==========
        for zone, _ in rh_second_impact_zones:
            files = request.files.getlist(f'rh_second_impact_img_{zone}')
            if files:
                imgs.setdefault(zone, [])
                for file in files:
                    if file and allowed_file(file.filename):
                        ext = file.filename.rsplit('.', 1)[-1].lower()
                        prefix = f"test_{group}_{test_key}_{zone}_"
                        nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                        next_num = max(nums, default=0) + 1
                        fname = f"{prefix}{next_num}.{ext}"
                        file.save(os.path.join(report_folder, fname))
                        imgs[zone].append(f"/images/{report}/{fname}")

        # ========== RH step12 zones ==========
        for zone, _ in rh_step12_zones:
            files = request.files.getlist(f'rh_step12_img_{zone}')
            if files:
                imgs.setdefault(zone, [])
                for file in files:
                    if file and allowed_file(file.filename):
                        ext = file.filename.rsplit('.', 1)[-1].lower()
                        prefix = f"test_{group}_{test_key}_{zone}_"
                        nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                        next_num = max(nums, default=0) + 1
                        fname = f"{prefix}{next_num}.{ext}"
                        file.save(os.path.join(report_folder, fname))
                        imgs[zone].append(f"/images/{report}/{fname}")

        # ========== DROP, IMPACT, ROTATION (t√°ch ra ngo√†i nh√°nh GT68) ==========
        # Drop
        if is_drop:
            for idx, zone in enumerate(DROP_ZONES):
                files = request.files.getlist(f'drop_img_{zone}')
                if files:
                    imgs.setdefault(idx, [])
                    for file in files:
                        if file and allowed_file(file.filename):
                            ext = file.filename.rsplit('.', 1)[-1].lower()
                            prefix = f"test_{group}_{test_key}_drop_{zone}_"
                            nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                            next_num = max(nums, default=0) + 1
                            fname = f"{prefix}{next_num}.{ext}"
                            file.save(os.path.join(report_folder, fname))
                            imgs[idx].append(f"/images/{report}/{fname}")

        # Impact
        if is_impact:
            for idx, zone in enumerate(IMPACT_ZONES):
                files = request.files.getlist(f'impact_img_{zone}')
                if files:
                    imgs.setdefault(idx, [])
                    for file in files:
                        if file and allowed_file(file.filename):
                            ext = file.filename.rsplit('.', 1)[-1].lower()
                            prefix = f"test_{group}_{test_key}_impact_{zone}_"
                            nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                            next_num = max(nums, default=0) + 1
                            fname = f"{prefix}{next_num}.{ext}"
                            file.save(os.path.join(report_folder, fname))
                            imgs[idx].append(f"/images/{report}/{fname}")

        # Rotation
        if is_rot:
            for idx, zone in enumerate(ROT_ZONES):
                files = request.files.getlist(f'rot_img_{zone}')
                if files:
                    imgs.setdefault(idx, [])
                    for file in files:
                        if file and allowed_file(file.filename):
                            ext = file.filename.rsplit('.', 1)[-1].lower()
                            prefix = f"test_{group}_{test_key}_rotation_{zone}_"
                            nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                            next_num = max(nums, default=0) + 1
                            fname = f"{prefix}{next_num}.{ext}"
                            file.save(os.path.join(report_folder, fname))
                            imgs[idx].append(f"/images/{report}/{fname}")

        # TH∆Ø·ªúNG
        if request.files.getlist('test_imgs'):
            imgs['normal'] = []
            for file in request.files.getlist('test_imgs'):
                if file and allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[-1].lower()
                    prefix = f"test_{group}_{test_key}_"
                    nums = [int(f[len(prefix):].split('.')[0]) for f in os.listdir(report_folder) if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()]
                    next_num = max(nums, default=0) + 1
                    fname = f"{prefix}{next_num}.{ext}"
                    file.save(os.path.join(report_folder, fname))
                    imgs['normal'].append(f"/images/{report}/{fname}")

        # X√≥a ·∫£nh AJAX
        if 'delete_img' in request.form:
            fname = request.form['delete_img']
            img_path = os.path.join(report_folder, fname)
            if os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except Exception:
                    pass  # ƒê√£ b·ªã x√≥a b·ªüi thread kh√°c
            # Tr·∫£ l·∫°i danh s√°ch ·∫£nh c√≤n l·∫°i
            if 'kind' in request.form and 'zone_idx' in request.form:
                kind = request.form['kind']
                idx = request.form['zone_idx']
                if kind in ['drop', 'impact', 'rot']:
                    # L·∫•y l·∫°i danh s√°ch ·∫£nh c√≤n l·∫°i cho zone idx
                    if kind == 'drop':
                        zone = DROP_ZONES[int(idx)]
                        prefix = f"test_{group}_{test_key}_drop_{zone}_"
                    elif kind == 'impact':
                        zone = IMPACT_ZONES[int(idx)]
                        prefix = f"test_{group}_{test_key}_impact_{zone}_"
                    elif kind == 'rot':
                        zone = ROT_ZONES[int(idx)]
                        prefix = f"test_{group}_{test_key}_rotation_{zone}_"
                    imgs[int(idx)] = []
                    for f in os.listdir(report_folder):
                        if allowed_file(f) and f.startswith(prefix):
                            imgs[int(idx)].append(f"/images/{report}/{f}")
                elif kind == 'gt68_face' and group == "transit_181_gt68" and test_key == "step4":
                    idx = int(idx)
                    zone = GT68_FACE_ZONES[idx]
                    prefix = f"test_{group}_{test_key}_gt68_face_{zone}_"
                    imgs[str(idx)] = []  # FIX: tr·∫£ v·ªÅ key "0".."5" ƒë·ªÉ kh·ªõp FE
                    for f in os.listdir(report_folder):
                        if allowed_file(f) and f.startswith(prefix):
                            imgs[str(idx)].append(f"/images/{report}/{f}")
                else:
                    # RH zones
                    zone = idx
                    prefix = f"test_{group}_{test_key}_{zone}_"
                    imgs[zone] = []
                    for f in os.listdir(report_folder):
                        if allowed_file(f) and f.startswith(prefix):
                            imgs[zone].append(f"/images/{report}/{f}")
            elif 'delete_img' in request.form:
                # ·∫¢nh th∆∞·ªùng
                imgs['normal'] = []
                for f in os.listdir(report_folder):
                    if allowed_file(f) and f.startswith(f"test_{group}_{test_key}_"):
                        imgs['normal'].append(f"/images/{report}/{f}")

        return jsonify(imgs=imgs)

    # --- Tr·∫°ng th√°i PASS/FAIL/N/A ---
    all_status = load_group_notes(status_file)
    status_value = all_status.get(test_key, "")

    # --- Comment ---
    comment = get_group_note_value(comment_file, test_key) 
    
    def get_comment(file_path, key):
        return get_group_note_value(file_path, key)

    # L·∫•y comment c·ªßa m·ª•c n√†y
    comment = get_comment(comment_file, test_key)

    # --- X√°c ƒë·ªãnh lo·∫°i test ƒë·∫∑c bi·ªát ---
    is_rh_np = (group == "transit_RH_np")
    is_drop = (is_drop_test(title) if group.startswith("transit") else False) or (group == "transit_181_lt68" and test_key == "step4")
    is_impact = is_impact_test(title) if group.startswith("transit") else False
    is_rot = is_rotational_test(title) if group.startswith("transit") else False

    # --- RH Non Pallet zones ---
    rh_impact_zones = RH_IMPACT_ZONES if is_rh_np and test_key == "step3" else []
    rh_vib_zones = RH_VIB_ZONES if is_rh_np and test_key == "step4" else []
    rh_second_impact_zones = RH_SECOND_IMPACT_ZONES if is_rh_np and test_key == "step5" else []
    rh_step12_zones = RH_STEP12_ZONES if is_rh_np and test_key == "step12" else []

    # --- X·ª≠ l√Ω upload ·∫£nh, x√≥a ·∫£nh, comment, status ---
    if request.method == 'POST':
        # Ch·ªâ upload ·∫£nh lo·∫°i th∆∞·ªùng (test_imgs)
        if 'test_imgs' in request.files:
            files = request.files.getlist('test_imgs')
            for file in files:
                if file and allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[-1].lower()
                    prefix = f"test_{group}_{test_key}_"
                    current_nums = [
                        int(f[len(prefix):].split('.')[0])
                        for f in os.listdir(report_folder)
                        if f.startswith(prefix) and f[len(prefix):].split('.')[0].isdigit()
                    ]
                    next_num = max(current_nums) + 1 if current_nums else 1
                    new_fname = f"{prefix}{next_num}.{ext}"
                    file.save(os.path.join(report_folder, new_fname))
        # X√≥a ·∫£nh th∆∞·ªùng
        if 'delete_img' in request.form:
            del_img = request.form['delete_img']
            img_path = os.path.join(report_folder, del_img)
            if os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except Exception:
                    pass
        # Ghi status PASS/FAIL/N/A
        if 'status' in request.form:
            update_group_note_file(status_file, test_key, request.form['status'])  # D√ôNG SAFE
        # Ghi comment
        if 'save_comment' in request.form:
            comment_val = request.form.get('comment_input', '').strip()
            update_group_note_file(comment_file, test_key, comment_val)  # D√ôNG SAFE
        return redirect(request.url)

    # --- Chu·∫©n b·ªã d·ªØ li·ªáu ·∫£nh v√πng RH (step3/4/5/12) ---
    zone_imgs = {}
    for zone, label in rh_impact_zones + rh_vib_zones + rh_second_impact_zones + rh_step12_zones:
        imgs_zone = []
        for f in os.listdir(report_folder):
            if allowed_file(f) and f.startswith(f"test_{group}_{test_key}_{zone}_"):
                imgs_zone.append(f"/images/{report}/{f}")
        zone_imgs[zone] = imgs_zone

    # --- Chu·∫©n b·ªã d·ªØ li·ªáu ·∫£nh th∆∞·ªùng ---
    imgs = []
    for f in sorted(os.listdir(report_folder)):
        # Ch·ªâ l·∫•y ·∫£nh lo·∫°i th∆∞·ªùng, kh√¥ng l·∫•y ·∫£nh v√πng
        if allowed_file(f) and f.startswith(f"test_{group}_{test_key}_") and all(not f.startswith(f"test_{group}_{test_key}_{zone}_") for zone, _ in rh_impact_zones + rh_vib_zones + rh_second_impact_zones + rh_step12_zones):
            imgs.append(f"/images/{report}/{f}")

    # --- Chu·∫©n b·ªã ·∫£nh drop, impact, rot n·∫øu c√≥ ---
    drop_imgs, impact_imgs, rot_imgs = [], [], []
    if is_drop:
        for zone in DROP_ZONES:
            di = []
            for f in os.listdir(report_folder):
                if allowed_file(f) and f.startswith(f"test_{group}_{test_key}_drop_{zone}_"):
                    di.append(f"/images/{report}/{f}")
            drop_imgs.append(di)
    if is_impact:
        for zone in IMPACT_ZONES:
            ii = []
            for f in os.listdir(report_folder):
                if allowed_file(f) and f.startswith(f"test_{group}_{test_key}_impact_{zone}_"):
                    ii.append(f"/images/{report}/{f}")
            impact_imgs.append(ii)
    if is_rot:
        for zone in ROT_ZONES:
            ri = []
            for f in os.listdir(report_folder):
                if allowed_file(f) and f.startswith(f"test_{group}_{test_key}_rotation_{zone}_"):
                    ri.append(f"/images/{report}/{f}")
            rot_imgs.append(ri)

    # --- Tr·∫£ v·ªÅ template ---
    return render_test_group_item(report, group, test_key, group_titles, comment=comment)

def render_test_group_item(report, group, key, group_titles, comment):
    title = group_titles[key]
    report_folder = os.path.join(UPLOAD_FOLDER, str(report))
    os.makedirs(report_folder, exist_ok=True)
    status_file = os.path.join(report_folder, f"status_{group}.txt")
    comment_file = os.path.join(report_folder, f"comment_{group}.txt")

    # RH Non Pallet zone logic
    is_rh_np = (group == "transit_RH_np")
    is_rh_np_step3 = is_rh_np and key == "step3"
    is_rh_np_step4 = is_rh_np and key == "step4"
    is_rh_np_step5 = is_rh_np and key == "step5"
    rh_impact_zones = RH_IMPACT_ZONES if is_rh_np_step3 else []
    rh_vib_zones = RH_VIB_ZONES if is_rh_np_step4 else []
    rh_second_impact_zones = RH_SECOND_IMPACT_ZONES if is_rh_np_step5 else []
    allow_na = is_rh_np and (key in ["step6", "step7", "step8", "step11", "step12"])

    # X·ª≠ l√Ω ·∫£nh v√πng RH (zone_imgs)
    zone_imgs = {}
    for zone, label in rh_impact_zones + rh_vib_zones + rh_second_impact_zones:
        imgs = []
        if os.path.exists(report_folder):
            for f in os.listdir(report_folder):
                if allowed_file(f) and f.startswith(f"test_{group}_{key}_{zone}_"):
                    imgs.append(f"/images/{report}/{f}")
        zone_imgs[zone] = imgs

    # V√πng Face cho transit_181_gt68 step4
    gt68_face_zones, gt68_face_labels, gt68_face_imgs = [], [], []
    if group == "transit_181_gt68" and key == "step4":
        gt68_face_zones = GT68_FACE_ZONES
        gt68_face_labels = GT68_FACE_LABELS
        for zone in gt68_face_zones:
            imgs = []
            if os.path.exists(report_folder):
                for f in os.listdir(report_folder):
                    if allowed_file(f) and f.startswith(f"test_{group}_{key}_gt68_face_{zone}_"):
                        imgs.append(f"/images/{report}/{f}")
            gt68_face_imgs.append(imgs)

    # Nh√≥m transit 2C logic
    TRANSIT_2C_GROUPS = ("transit_2c_np", "transit_2c_pallet")
    if not (is_rh_np_step3 or is_rh_np_step4 or is_rh_np_step5):
        is_transit_2c = group in TRANSIT_2C_GROUPS
        is_drop = (is_drop_test(title) and is_transit_2c) or (group == "transit_181_lt68" and key == "step4")
        is_impact = is_impact_test(title) and is_transit_2c
        is_rot = is_rotational_test(title) and is_transit_2c
    else:
        is_drop = is_impact = is_rot = False

    # Drop, Impact, Rot imgs
    drop_imgs = []
    if is_drop:
        for zone in DROP_ZONES:
            imgs = []
            if os.path.exists(report_folder):
                for f in os.listdir(report_folder):
                    if allowed_file(f) and f.startswith(f"test_{group}_{key}_drop_{zone}_"):
                        imgs.append(f"/images/{report}/{f}")
            drop_imgs.append(imgs)

    impact_imgs = []
    rot_imgs = []
    if is_impact:
        for zone in IMPACT_ZONES:
            imgs = []
            if os.path.exists(report_folder):
                for f in os.listdir(report_folder):
                    if allowed_file(f) and f.startswith(f"test_{group}_{key}_impact_{zone}_"):
                        imgs.append(f"/images/{report}/{f}")
            impact_imgs.append(imgs)
    if is_rot:
        for zone in ROT_ZONES:
            imgs = []
            if os.path.exists(report_folder):
                for f in os.listdir(report_folder):
                    if allowed_file(f) and f.startswith(f"test_{group}_{key}_rotation_{zone}_"):
                        imgs.append(f"/images/{report}/{f}")
            rot_imgs.append(imgs)
    gt68_face_imgs = []
    if group == "transit_181_gt68" and key == "step4":
        for zone in GT68_FACE_ZONES:
            imgs = []
            if os.path.exists(report_folder):
                for f in os.listdir(report_folder):
                    if allowed_file(f) and f.startswith(f"test_{group}_{key}_gt68_face_{zone}_"):
                        imgs.append(f"/images/{report}/{f}")
            gt68_face_imgs.append(imgs)

    # === Status/comment helper ===
    def update_group_note_file(file_path, key, value):
        # ƒê·ªçc file d√πng lock
        lines = []
        found = False
        content = safe_read_text(file_path)
        if content:
            lines = content.splitlines(keepends=True)
        new_lines = []
        for line in lines:
            if line.strip().startswith(f"M·ª•c {key}:"):
                new_lines.append(f"M·ª•c {key}: {value}\n")
                found = True
            else:
                new_lines.append(line)
        if not found:
            new_lines.append(f"M·ª•c {key}: {value}\n")
        # Ghi l·∫°i d√πng lock
        safe_write_text(file_path, "".join(new_lines))

    def get_group_note_value(file_path, key):
        content = safe_read_text(file_path)
        if content:
            for line in content.splitlines():
                s = line.strip()
                if s.startswith(f"M·ª•c {key}:") or s.startswith(f"{key}:"):
                    return s.split(":", 1)[1].strip()
        return None

    status_value = get_group_note_value(status_file, key)

    # === X·ª≠ l√Ω POST: ch·ªâ x·ª≠ l√Ω x√≥a ·∫£nh, status, comment (KH√îNG UPLOAD ·∫¢NH V√ôNG ZONE ·ªû ƒê√ÇY) ===
    if request.method == 'POST':
        # X√≥a ·∫£nh th∆∞·ªùng ho·∫∑c v√πng
        if 'delete_img' in request.form:
            del_img = request.form['delete_img']
            img_path = os.path.join(report_folder, del_img)
            if os.path.exists(img_path):
                os.remove(img_path)
        # Ghi status PASS/FAIL/N/A
        if 'status' in request.form:
            status = request.form['status']
            update_group_note_file(status_file, key, status)
        # Ghi comment
        if 'save_comment' in request.form:
            comment = request.form.get('comment_input', '').strip()
            update_group_note_file(comment_file, key, comment)
        # C·∫≠p nh·∫≠t lo·∫°i ki·ªÉm tra g·∫ßn nh·∫•t
        vi_name = TEST_TYPE_VI.get(group, group.upper())
        set_last_test_type(report, vi_name)
        return redirect(request.url)

    # === T√≠nh tr·∫°ng th√°i t·ªïng th·ªÉ t·ª´ng m·ª•c cho menu group ===
    test_status = {}
    for k in group_titles:
        st = get_group_note_value(status_file, k) if not group.startswith("transit") else None
        cm = get_group_note_value(comment_file, k)
        has_img = False
        prefix = f"test_{group}_{k}_"
        if os.path.exists(report_folder):
            for fn in os.listdir(report_folder):
                if allowed_file(fn) and fn.startswith(prefix):
                    has_img = True
                    break
        test_status[k] = {
            'status': st,
            'comment': cm,
            'has_img': has_img
        }

    # === L·∫•y ·∫£nh th∆∞·ªùng cho m·ª•c kh√¥ng ph·∫£i drop/impact/rot/RH np ===
    imgs = []
    if os.path.exists(report_folder) and not is_drop:
        prefix = f"test_{group}_{key}_"
        for f in sorted(os.listdir(report_folder)):
            if allowed_file(f) and f.startswith(prefix):
                # Ch·ªâ nh·∫≠n file c√≥ s·ªë th·ª© t·ª± ngay sau prefix (vd: ..._1.jpg, ..._2.png)
                tail = f[len(prefix):].split('.')[0]
                if tail.isdigit():
                    imgs.append(f"/images/{report}/{f}")

    # === Ch·ªçn template (transit d√πng test_transit_item.html) ===
    TRANSIT_GROUPS = (
        "transit_2c_np", "transit_2c_pallet",
        "transit_RH_np", "transit_RH_pallet",
        "transit_181_lt68", "transit_181_gt68",
        "transit_3b_np", "transit_3b_pallet", "transit_3a_np"
    )
    if group in TRANSIT_GROUPS:
        template_name = "test_transit_item.html"
    else:
        template_name = "test_group_item.html"

    return render_template(
        template_name,
        report=report,
        group=group,
        test_titles=group_titles,
        test_status=test_status,
        key=key,
        imgs=imgs,
        status=status_value,
        comment=get_group_note_value(comment_file, key),
        title=title,
        is_drop=is_drop,
        drop_labels=DROP_LABELS,
        drop_zones=DROP_ZONES,
        drop_imgs=drop_imgs,
        is_impact=is_impact,
        impact_labels=IMPACT_LABELS,
        impact_zones=IMPACT_ZONES,
        impact_imgs=impact_imgs,
        is_rot=is_rot,
        rot_labels=ROT_LABELS,
        rot_zones=ROT_ZONES,
        rot_imgs=rot_imgs,
        is_rh_np=is_rh_np,
        rh_impact_zones=rh_impact_zones,
        rh_vib_zones=rh_vib_zones,
        rh_second_impact_zones=rh_second_impact_zones,
        allow_na=allow_na,
        zone_imgs=zone_imgs,
        gt68_face_labels=GT68_FACE_LABELS,
        gt68_face_zones=GT68_FACE_ZONES,
        gt68_face_imgs=gt68_face_imgs,
    )

def read_kv_file(path):
    """
    ƒê·ªçc file d·∫°ng:
        key: value
        key2: value2
    -> tr·∫£ v·ªÅ dict {key: value}
    """
    data = {}
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                for raw in f.readlines():
                    line = raw.strip()
                    if not line:
                        continue
                    # ch·∫•p nh·∫≠n c·∫£ ':' l·∫´n '=' cho b·ªÅn
                    if ":" in line:
                        k, v = line.split(":", 1)
                    elif "=" in line:
                        k, v = line.split("=", 1)
                    else:
                        # n·∫øu l√† file ki·ªÉu c≈© ch·ªâ c√≥ 1 gi√° tr·ªã (PASS/FAIL/...)
                        # g√°n v√†o key 'default'
                        data["default"] = line
                        continue
                    data[k.strip()] = v.strip()
    except Exception:
        pass
    return data


def upsert_kv_line(path, key, value):
    """
    Ghi/ c·∫≠p nh·∫≠t m·ªôt d√≤ng 'key: value' v√†o file.
    - N·∫øu ch∆∞a c√≥ file => t·∫°o m·ªõi.
    - N·∫øu c√≥ => c·∫≠p nh·∫≠t ƒë√∫ng key, gi·ªØ c√°c key kh√°c.
    """
    d = read_kv_file(path)
    d[key] = value
    try:
        with open(path, "w", encoding="utf-8") as f:
            for k, v in d.items():
                f.write(f"{k}: {v}\n")
    except Exception:
        pass


# ====== Route hot_cold (ghi status/comment theo t·ª´ng test_key) ===============
# Cho ph√©p URL c√≥/kh√¥ng c√≥ test_key (m·∫∑c ƒë·ªãnh l√† 'hot_cold' ƒë·ªÉ kh√¥ng ph√° link c≈©)
@app.route("/hot_cold_test/<report>/<group>", defaults={'test_key': 'hot_cold'}, methods=["GET", "POST"])
@app.route("/hot_cold_test/<report>/<group>/<test_key>", methods=["GET", "POST"])
def hot_cold_test(report, group, test_key):
    from_line = request.args.get("from_line")

    # ====== L·∫•y t√™n hi·ªÉn th·ªã ƒë√∫ng theo test_key ======
    try:
        raw_title = TEST_GROUP_TITLES.get(group, {}).get(test_key)
    except Exception:
        raw_title = None

    if isinstance(raw_title, dict):
        display_title = raw_title.get('short') or raw_title.get('full') or test_key.replace('_', ' ').title()
    elif raw_title:
        display_title = raw_title
    else:
        display_title = test_key.replace('_', ' ').title()

    set_last_test_type(report, f"{display_title} ({group.upper()})")

    # ====== Chu·∫©n b·ªã ƒë∆∞·ªùng d·∫´n/l∆∞u tr·ªØ ======
    vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')
    folder = os.path.join(UPLOAD_FOLDER, str(report))
    os.makedirs(folder, exist_ok=True)

    # File CHUNG theo group (m·ªói n√∫t 1 d√≤ng)
    common_prefix = f"{group}"
    status_file   = os.path.join(folder, f"status_{common_prefix}.txt")
    comment_file  = os.path.join(folder, f"comment_{common_prefix}.txt")

    # ·∫¢nh & m·ªëc th·ªùi gian v·∫´n theo test_key (kh√¥ng ƒë·ªïi)
    test_prefix       = f"{test_key}_{group}"   # v√≠ d·ª•: hot_cold_indoor_thuong
    before_tag        = f"{test_key}_before_{group}"
    after_tag         = f"{test_key}_after_{group}"
    before_time_file  = os.path.join(folder, f"{test_prefix}_before_time.txt")
    after_time_file   = os.path.join(folder, f"{test_prefix}_after_time.txt")
    duration_file     = os.path.join(folder, f"{test_prefix}_duration.txt")

    # ====== X·ª≠ l√Ω POST ======
    if request.method == "POST":
        # 1) C·∫≠p nh·∫≠t tr·∫°ng th√°i -> ghi/upsert theo test_key (KH√îNG ghi ƒë√® c·∫£ file)
        if "status" in request.form:
            status_value = (request.form.get("status") or "").strip()
            if status_value:
                upsert_kv_line(status_file, test_key, status_value)

        # 2) L∆∞u ghi ch√∫ -> c≈©ng upsert theo test_key
        if "save_comment" in request.form:
            cmt = (request.form.get("comment_input") or "").strip()
            # ƒë·ªÉ tr√°nh ph√° format m·ªôt d√≤ng, thay newline b·∫±ng ' / '
            cmt_one_line = " / ".join([s.strip() for s in cmt.splitlines() if s.strip()]) if cmt else ""
            upsert_kv_line(comment_file, test_key, cmt_one_line)

        # 3) Upload ·∫£nh (before/after) + ghi m·ªëc th·ªùi gian t∆∞∆°ng ·ª©ng
        for tag, time_file in [(before_tag, before_time_file), (after_tag, after_time_file)]:
            field_name = f"{tag}_imgs"
            if field_name in request.files:
                files = request.files.getlist(field_name)
                count = 0
                for file in files:
                    if file and allowed_file(file.filename):
                        ext = file.filename.rsplit('.', 1)[-1].lower()
                        prefix_img = f"{tag}_"
                        nums = [
                            int(fname[len(prefix_img):].split('.')[0])
                            for fname in os.listdir(folder)
                            if fname.startswith(prefix_img) and fname[len(prefix_img):].split('.')[0].isdigit()
                        ]
                        next_num = max(nums) + 1 if nums else 1
                        new_fname = f"{tag}_{next_num}.{ext}"
                        file.save(os.path.join(folder, new_fname))
                        count += 1
                if count > 0:
                    now = datetime.now(vn_tz).strftime("%d/%m/%Y %H:%M")
                    safe_write_text(time_file, now)

        # 4) Xo√° ·∫£nh (v√† d·ªçn m·ªëc th·ªùi gian n·∫øu h·∫øt ·∫£nh)
        if "delete_img" in request.form:
            img = request.form["delete_img"]
            img_path = os.path.join(folder, img)
            if os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except Exception:
                    pass
            # N·∫øu kh√¥ng c√≤n ·∫£nh before/after th√¨ xo√° file time t∆∞∆°ng ·ª©ng
            if img.startswith(before_tag):
                still = [f for f in os.listdir(folder) if allowed_file(f) and f.startswith(before_tag)]
                if not still and os.path.exists(before_time_file):
                    try: os.remove(before_time_file)
                    except Exception: pass
            if img.startswith(after_tag):
                still = [f for f in os.listdir(folder) if allowed_file(f) and f.startswith(after_tag)]
                if not still and os.path.exists(after_time_file):
                    try: os.remove(after_time_file)
                    except Exception: pass

        # 5) C·∫≠p nh·∫≠t th·ªùi gian test (gi·ªù)
        if "set_duration" in request.form:
            raw = (request.form.get("duration") or "").strip()
            try:
                dur = float(raw)
                if dur <= 0: raise ValueError
                safe_write_text(duration_file, str(dur))
                flash("ƒê√£ c·∫≠p nh·∫≠t th·ªùi gian test.", "success")
            except Exception:
                flash("Gi√° tr·ªã th·ªùi gian kh√¥ng h·ª£p l·ªá.", "danger")

        # tr√°nh resubmit
        session[f"last_test_type_{report}"] = f"{display_title} ({group.upper()})"
        return redirect(request.url)

    # ====== ƒê·ªçc d·ªØ li·ªáu ƒë·ªÉ render (l·∫•y ƒë√∫ng m·ª•c theo test_key) ======
    status_map  = read_kv_file(status_file)
    comment_map = read_kv_file(comment_file)

    # ∆Øu ti√™n key c·ª• th·ªÉ; n·∫øu tr∆∞·ªõc ƒë√¢y file c≈© l∆∞u 1 d√≤ng kh√¥ng key th√¨ d√πng 'default'
    status  = (status_map.get(test_key) or status_map.get("default") or "").strip()
    comment = (comment_map.get(test_key) or comment_map.get("default") or "").strip()

    # H√¨nh m√¥ t·∫£ (n·∫øu c√≥ trong TEST_GROUP_TITLES)
    try:
        imgs_mo_ta = (TEST_GROUP_TITLES.get(group, {}).get(test_key) or {}).get("img", [])
    except Exception:
        imgs_mo_ta = []

    # Danh s√°ch ·∫£nh before/after
    imgs_before, imgs_after = [], []
    for fname in sorted(os.listdir(folder)):
        if allowed_file(fname):
            if fname.startswith(before_tag):
                imgs_before.append(f"/images/{report}/{fname}")
            elif fname.startswith(after_tag):
                imgs_after.append(f"/images/{report}/{fname}")

    # Th·ªùi gian upload
    before_upload_time = (safe_read_text(before_time_file) or "").strip() if os.path.exists(before_time_file) else None
    after_upload_time  = (safe_read_text(after_time_file) or "").strip()  if os.path.exists(after_time_file)  else None

    # Th·ªùi l∆∞·ª£ng test (gi·ªù)
    raw_duration = safe_read_text(duration_file)
    try:
        so_gio_test = float(raw_duration) if raw_duration not in (None, "") else float(SO_GIO_TEST)
    except Exception:
        so_gio_test = 4.0

    # ====== Render ======
    return render_template(
        "hot_cold_test.html",
        report=report,
        group=group,
        test_key=test_key,
        title={'short': display_title, 'full': display_title},
        status=status,
        comment=comment,
        imgs_mo_ta=imgs_mo_ta,
        imgs_before=imgs_before,
        imgs_after=imgs_after,
        before_upload_time=before_upload_time,
        after_upload_time=after_upload_time,
        so_gio_test=so_gio_test,
        from_line=from_line,
        before_tag=before_tag,
        after_tag=after_tag,
    )

def get_hotcold_elapsed(report, group):
    folder = os.path.join(UPLOAD_FOLDER, str(report))
    time_file = os.path.join(folder, f"hot_cold_upload_time_{group}.txt")
    tstr = safe_read_text(time_file).strip() if os.path.exists(time_file) else ""
    if tstr:
        try:
            vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')
            dt = datetime.strptime(tstr, "%d/%m/%Y %H:%M")
            dt = vn_tz.localize(dt)
            now = datetime.now(vn_tz)
            return (now - dt).total_seconds() / 3600
        except Exception:
            return None
    return None

@app.route("/line_test/<report>", methods=["GET", "POST"])
def line_test(report):
    set_last_test_type(report, "LINE TEST")
    folder = os.path.join(UPLOAD_FOLDER, str(report))
    os.makedirs(folder, exist_ok=True)
    before_tag, after_tag = "line_before", "line_after"
    files_map = {
        "status": os.path.join(folder, "line_status.txt"),
        "comment": os.path.join(folder, "line_comment.txt"),
        "before_time": os.path.join(folder, "before_upload_time.txt"),
        "after_time": os.path.join(folder, "after_upload_time.txt"),
    }
    fail_reasons_list = [
        "V·∫≠t li·ªáu b·ªã ·∫©m.",
        "V·ªã tr√≠ b·ªã t√°ch l·ªõp, m·∫∑t d∆∞·ªõi veneer c√≥ ph·ªß keo.",
        "V·ªã tr√≠ b·ªã t√°ch l·ªõp, m·∫∑t d∆∞·ªõi veneer kh√¥ng ph·ªß ƒë·ªÅu keo."
    ]

    # --- POST ---
    if request.method == "POST":
        # L∆∞u tr·∫°ng th√°i PASS/FAIL/DATA
        if "status" in request.form:
            safe_write_text(files_map["status"], request.form["status"])
            if request.form["status"] != "FAIL":
                if os.path.exists(files_map["comment"]):
                    os.remove(files_map["comment"])
        # L∆∞u fail reason
        if "save_fail_reason" in request.form:
            reasons = request.form.getlist("fail_reason")
            other = request.form.get("fail_reason_other", "").strip()
            if other: reasons.append(other)
            safe_write_text(files_map["comment"], "; ".join(reasons))
        # Upload ·∫£nh before/after
        for tag, time_file in [(before_tag, files_map["before_time"]), (after_tag, files_map["after_time"])]:
            if f"{tag}_imgs" in request.files:
                files = request.files.getlist(f"{tag}_imgs")
                nums = [int(f[len(tag)+1:].split('.')[0]) for f in os.listdir(folder)
                        if f.startswith(f"{tag}_") and f[len(tag)+1:].split('.')[0].isdigit()]
                next_num = max(nums, default=0) + 1
                count = 0
                for file in files:
                    if file and allowed_file(file.filename):
                        ext = file.filename.rsplit('.', 1)[-1].lower()
                        file.save(os.path.join(folder, f"{tag}_{next_num}.{ext}"))
                        next_num += 1
                        count += 1
                if count:
                    vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')
                    safe_write_text(time_file, datetime.now(vn_tz).strftime("%d/%m/%Y %H:%M"))
        # X√≥a ·∫£nh
        if "delete_img" in request.form:
            img = request.form["delete_img"]
            img_path = os.path.join(folder, img)
            if os.path.exists(img_path): os.remove(img_path)
            for tag, time_file in [(before_tag, files_map["before_time"]), (after_tag, files_map["after_time"])]:
                if img.startswith(tag):
                    if not any(allowed_file(f) and f.startswith(tag) for f in os.listdir(folder)):
                        if os.path.exists(time_file): os.remove(time_file)
        set_last_test_type(report, "LINE TEST")
        return redirect(request.url)

    # --- GET: ƒê·ªçc d·ªØ li·ªáu ƒë√£ l∆∞u ---
    status = safe_read_text(files_map["status"])
    fail_reason_raw = safe_read_text(files_map["comment"])
    fail_reasons, fail_reason_other = [], ""
    if fail_reason_raw:
        all_reasons = [r.strip() for r in fail_reason_raw.split(";") if r.strip()]
        for r in all_reasons[:]:
            if r not in fail_reasons_list:
                fail_reason_other = r
                all_reasons.remove(r)
        fail_reasons = all_reasons
    imgs_before = [f"/images/{report}/{f}" for f in sorted(os.listdir(folder)) if allowed_file(f) and f.startswith(before_tag)]
    imgs_after  = [f"/images/{report}/{f}" for f in sorted(os.listdir(folder)) if allowed_file(f) and f.startswith(after_tag)]
    before_upload_time = safe_read_text(files_map["before_time"])
    after_upload_time  = safe_read_text(files_map["after_time"])

    return render_template(
        "line_test.html",
        report=report,
        status=status,
        fail_reasons=fail_reasons,
        fail_reason_other=fail_reason_other,
        fail_reasons_list=fail_reasons_list,
        imgs_before=imgs_before,
        imgs_after=imgs_after,
        before_upload_time=before_upload_time,
        after_upload_time=after_upload_time,
        so_gio_test=SO_GIO_TEST,
    )

def get_line_test_elapsed(report):
    folder = os.path.join(UPLOAD_FOLDER, str(report))
    before_time_file = os.path.join(folder, "before_upload_time.txt")
    tstr = safe_read_text(before_time_file)  # D√πng h√†m an to√†n, ƒë√£ c√≥ filelock
    if tstr:
        try:
            vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')
            dt = datetime.strptime(tstr, "%d/%m/%Y %H:%M")
            dt = vn_tz.localize(dt)
            now = datetime.now(vn_tz)
            elapsed = (now - dt).total_seconds() / 3600
            return elapsed
        except Exception as e:
            print("Parse time error:", e)
            return None
    return None

SAMPLE_STORAGE_FILE = os.path.join(BASE_DIR, "sample_storage.json")

@app.route("/store_sample", methods=["GET", "POST"])
def store_sample():
    """
    Route l∆∞u m·∫´u theo format m·ªõi (box_code, save_date, discard_date...),
    cho ph√©p nhi·ªÅu m·∫´u trong c√πng 1 box (list d·∫°ng).
    """
    report = request.args.get("report")
    item_code = get_item_code(report)
    error_msg = ""

    # --- ƒê·ªçc d·ªØ li·ªáu an to√†n ---
    SAMPLE_STORAGE = safe_read_json(SAMPLE_STORAGE_FILE)
    if not isinstance(SAMPLE_STORAGE, dict):
        SAMPLE_STORAGE = {}

    # --- Ki·ªÉm tra xem m·∫´u ƒë√£ t·ªìn t·∫°i ch∆∞a (theo report + item_code) ---
    existing_box = None
    for box, info in SAMPLE_STORAGE.items():
        if isinstance(info, list):
            # box ch·ª©a nhi·ªÅu m·∫´u
            for entry in info:
                if entry.get("report") == report and entry.get("item_code") == item_code:
                    existing_box = box
                    break
        elif isinstance(info, dict):
            # d·ªØ li·ªáu c≈© ch·ªâ c√≥ 1 m·∫´u
            if info.get("report") == report and info.get("item_code") == item_code:
                existing_box = box
        if existing_box:
            break

    # N·∫øu ƒë√£ t·ªìn t·∫°i m·∫´u => chuy·ªÉn sang trang info
    if existing_box:
        return redirect(url_for("sample_infor", report=report, box_code=existing_box))

    # --- N·∫øu ch∆∞a c√≥ m·∫´u => x·ª≠ l√Ω form POST ---
    if request.method == "POST":
        sample_type = request.form.get("sample_type", "").strip()
        box_code = request.form.get("box_code", "").strip().upper()
        note = request.form.get("note", "").strip()

        save_date_str = request.form.get("save_date")
        discard_date_str = request.form.get("discard_date")

        from datetime import datetime, date
        from dateutil.relativedelta import relativedelta

        try:
            save_date = datetime.strptime(save_date_str, "%Y-%m-%d").date() if save_date_str else date.today()
        except:
            save_date = date.today()

        try:
            discard_date = datetime.strptime(discard_date_str, "%Y-%m-%d").date() if discard_date_str else (save_date + relativedelta(months=3))
        except:
            discard_date = save_date + relativedelta(months=3)

        # --- Ghi d·ªØ li·ªáu ---
        if box_code not in SAMPLE_STORAGE:
            SAMPLE_STORAGE[box_code] = []
        elif isinstance(SAMPLE_STORAGE[box_code], dict):
            # n·∫øu d·ªØ li·ªáu c≈© ch·ªâ c√≥ 1 m·∫´u th√¨ chuy·ªÉn th√†nh list
            SAMPLE_STORAGE[box_code] = [SAMPLE_STORAGE[box_code]]

        # th√™m m·∫´u m·ªõi v√†o danh s√°ch trong box
        SAMPLE_STORAGE[box_code].append({
            "report": report,
            "item_code": item_code,
            "sample_type": sample_type,
            "box_code": box_code,
            "save_date": save_date.isoformat(),
            "discard_date": discard_date.isoformat(),
            "note": note
        })

        safe_write_json(SAMPLE_STORAGE_FILE, SAMPLE_STORAGE)
        return redirect(url_for("sample_infor", report=report, box_code=box_code))

    # --- GET request: render form ---
    return render_template(
        "sample_form.html",
        report=report,
        item_code=item_code
    )

@app.route("/report/<report>/sample/edit/<box_code>", methods=["GET", "POST"])
def edit_sample(report, box_code):
    """Ch·ªânh s·ª≠a m·∫´u c√≥ s·∫µn trong box (t∆∞∆°ng th√≠ch nhi·ªÅu m·∫´u)."""
    SAMPLE_STORAGE = safe_read_json(SAMPLE_STORAGE_FILE)
    if not isinstance(SAMPLE_STORAGE, dict):
        SAMPLE_STORAGE = {}

    data = None
    if box_code in SAMPLE_STORAGE:
        box_data = SAMPLE_STORAGE[box_code]
        if isinstance(box_data, list):
            for entry in box_data:
                if entry.get("report") == report:
                    data = entry
                    break
        elif isinstance(box_data, dict):
            data = box_data

    if not data:
        return f"Kh√¥ng t√¨m th·∫•y m·∫´u trong box {box_code}", 404

    from datetime import datetime, date
    from dateutil.relativedelta import relativedelta

    if request.method == "POST":
        sample_type = request.form.get("sample_type")
        note = request.form.get("note", "")
        save_date_str = request.form.get("save_date")
        discard_date_str = request.form.get("discard_date")

        try:
            save_date = datetime.strptime(save_date_str, "%Y-%m-%d").date()
        except:
            save_date = date.today()

        try:
            discard_date = datetime.strptime(discard_date_str, "%Y-%m-%d").date()
        except:
            discard_date = save_date + relativedelta(months=3)

        # c·∫≠p nh·∫≠t d·ªØ li·ªáu trong list ƒë√∫ng m·∫´u ƒë√≥
        if isinstance(SAMPLE_STORAGE[box_code], list):
            for i, entry in enumerate(SAMPLE_STORAGE[box_code]):
                if entry.get("report") == report:
                    SAMPLE_STORAGE[box_code][i].update({
                        "sample_type": sample_type,
                        "note": note,
                        "save_date": save_date.isoformat(),
                        "discard_date": discard_date.isoformat()
                    })
                    break
        elif isinstance(SAMPLE_STORAGE[box_code], dict):
            SAMPLE_STORAGE[box_code].update({
                "sample_type": sample_type,
                "note": note,
                "save_date": save_date.isoformat(),
                "discard_date": discard_date.isoformat()
            })

        safe_write_json(SAMPLE_STORAGE_FILE, SAMPLE_STORAGE)
        return redirect(url_for("list_samples", report=report))

    # --- GET request: render form ---
    return render_template(
        "sample_form.html",
        report=report,
        item_code=data.get("item_code", ""),
        box_code=box_code,
        sample_type=data.get("sample_type", ""),
        note=data.get("note", ""),
        save_date=data.get("save_date", ""),
        discard_date=data.get("discard_date", "")
    )

@app.route("/boxes", methods=["GET","POST","DELETE"])
def boxes_api():
    boxes = read_boxes()

    if request.method == "GET":
        return jsonify({"boxes": boxes})

    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        code = (payload.get("code") or "").strip().upper()
        if code and code not in boxes:
            boxes.append(code)
            boxes = sorted(set(boxes))
            write_boxes(boxes)
        return jsonify({"boxes": boxes})

    if request.method == "DELETE":
        code = (request.args.get("code") or "").strip().upper()
        if code in boxes:
            boxes = [b for b in boxes if b != code]
            write_boxes(boxes)
        return jsonify({"boxes": boxes})
    
@app.route("/report/<report>/sample/info/<box_code>")
def sample_infor(report, box_code):
    SAMPLE_STORAGE = safe_read_json(SAMPLE_STORAGE_FILE)
    info = None
    if box_code in SAMPLE_STORAGE:
        box_data = SAMPLE_STORAGE[box_code]
        if isinstance(box_data, list):
            # t√¨m m·∫´u ƒë√∫ng report trong box
            for entry in box_data:
                if entry.get("report") == report:
                    info = entry
                    break
        elif isinstance(box_data, dict):
            info = box_data
    if not info:
        return f"Kh√¥ng t√¨m th·∫•y m·∫´u trong box {box_code}", 404

    return render_template(
        "sample_infor.html",
        info=info,
        report_id=report,
        item_code=info.get("item_code", ""),
        box_code=box_code
    )

@app.route("/list_samples", methods=["GET"])
def list_samples():
    SAMPLE_STORAGE = safe_read_json(SAMPLE_STORAGE_FILE)
    if not isinstance(SAMPLE_STORAGE, dict):
        SAMPLE_STORAGE = {}

    samples = []

    for box_code, info in SAMPLE_STORAGE.items():
        if isinstance(info, list):
            # box c√≥ nhi·ªÅu m·∫´u
            for entry in info:
                samples.append({
                    "box_code": box_code,
                    "report": entry.get("report", ""),
                    "item_code": entry.get("item_code", ""),
                    "sample_type": entry.get("sample_type", ""),
                    "save_date": entry.get("save_date", ""),
                    "discard_date": entry.get("discard_date", ""),
                    "note": entry.get("note", "")
                })
        elif isinstance(info, dict):
            # d·ªØ li·ªáu c≈© ch·ªâ c√≥ 1 m·∫´u
            samples.append({
                "box_code": box_code,
                "report": info.get("report", ""),
                "item_code": info.get("item_code", ""),
                "sample_type": info.get("sample_type", ""),
                "save_date": info.get("save_date", ""),
                "discard_date": info.get("discard_date", ""),
                "note": info.get("note", "")
            })

    # Debug log (t√πy ch·ªçn)
    print(f"‚úÖ Loaded {len(samples)} samples from {len(SAMPLE_STORAGE)} boxes")

    # Truy·ªÅn JSON v√†o template ƒë·ªÉ render
    from markupsafe import Markup
    import json
    samples_json = Markup(json.dumps(samples, ensure_ascii=False))

    return render_template(
        "list_samples.html",
        samples_json=samples_json,
        report_id=request.args.get("report", "")
    )

# ====== API: X√ìA M·∫™U THEO BOX_CODE ======
@app.route("/samples", methods=["DELETE"])
def delete_sample():
    box_code = (request.args.get("box_code") or "").strip().upper()
    report = (request.args.get("report") or "").strip()

    SAMPLE_STORAGE = safe_read_json(SAMPLE_STORAGE_FILE)
    if not isinstance(SAMPLE_STORAGE, dict):
        SAMPLE_STORAGE = {}

    if box_code not in SAMPLE_STORAGE:
        return {"ok": False, "error": "not found"}, 404

    box_data = SAMPLE_STORAGE[box_code]

    if isinstance(box_data, list) and report:
        box_data = [r for r in box_data if r.get("report") != report]
        if box_data:
            SAMPLE_STORAGE[box_code] = box_data
        else:
            del SAMPLE_STORAGE[box_code]
    else:
        del SAMPLE_STORAGE[box_code]

    safe_write_json(SAMPLE_STORAGE_FILE, SAMPLE_STORAGE)
    return {"ok": True}

@app.post('/export_expired_samples')
def export_expired_samples():
    """Nh·∫≠n list m·∫´u t·ª´ client, l·ªçc c√°c m·∫´u h·∫øt h·∫°n (ƒëang c√≥ status-danger) ƒë·ªÉ xu·∫•t Excel."""
    data = request.get_json(force=True) or []
    # l·ªçc nh·ªØng m·∫´u c√≥ ng√†y h·ªßy < h√¥m nay
    today = datetime.now().date()
    expired = [r for r in data if r.get('discard_date') and r.get('discard_date') < str(today)]

    wb_bytes = export_expired_samples_to_excel(expired)
    return send_file(
        BytesIO(wb_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='expired_samples.xlsx'
    )

@app.route('/images/<report>/imgs_<group>_<test_key>/<filename>')
def serve_test_img(report, group, test_key, filename):
    folder = os.path.join(UPLOAD_FOLDER, report, f"imgs_{group}_{test_key}")
    if not os.path.exists(folder):
        # B√°o l·ªói r√µ r√†ng ho·∫∑c tr·∫£ v·ªÅ 404
        return "Kh√¥ng t√¨m th·∫•y th∆∞ m·ª•c ·∫£nh!", 404
    return send_from_directory(folder, filename)

@app.route("/api/report_comment")
def api_report_comment():
    report = (request.args.get("report") or "").strip()
    if not report:
        return jsonify({"comment": ""})

    base_dir = os.path.dirname(os.path.abspath(__file__))
    rep_dir = os.path.join(base_dir, "images", report)
    if not os.path.isdir(rep_dir):
        return jsonify({"comment": ""})

    # ∆Øu ti√™n file c√≥ ch·ªØ "comment" v√† ƒëu√¥i .txt/.md/.json/.log; fallback c√°c .txt kh√°c
    patterns = ["comment*.txt", "comment*.md", "comment*.json", "comment*.log",
                "*.txt", "*.md", "*.json"]
    candidates = []
    for pat in patterns:
        candidates.extend(glob.glob(os.path.join(rep_dir, pat)))

    # S·∫Øp theo ∆∞u ti√™n 'comment' r·ªìi m·ªõi ƒë·∫øn th·ªùi gian m·ªõi nh·∫•t
    candidates.sort(key=lambda p: (0 if "comment" in os.path.basename(p).lower() else 1, -os.path.getmtime(p)))

    text = ""
    for p in candidates:
        try:
            if p.lower().endswith(".json"):
                import json
                with open(p, "r", encoding="utf-8", errors="ignore") as f:
                    obj = json.load(f)
                # ch·∫•p nh·∫≠n kh√≥a 'comment' ho·∫∑c to√†n b·ªô n·ªôi dung l√† chu·ªói
                if isinstance(obj, dict) and "comment" in obj:
                    text = str(obj["comment"])
                elif isinstance(obj, str):
                    text = obj
                else:
                    # g·ªôp c√°c value th√†nh text
                    text = "\n".join(f"{k}: {v}" for k, v in obj.items())
            else:
                with open(p, "r", encoding="utf-8", errors="ignore") as f:
                    text = f.read()
            if text.strip():
                break
        except Exception:
            continue

    return jsonify({"comment": text})

@app.route("/view_counter_log")
def view_counter_log():

    excel_path = "counter_detail_log.xlsx"
    rows = []
    type_of_set = set()
    ca_map = {"office": "HC", "hc": "HC", "ot": "OT"}
    header = ["Ng√†y", "Ca", "T·ªïng"]  # Default

    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
            # Build column name -> index map
            col_idx = {str(cell.value).strip().lower(): i for i, cell in enumerate(ws[1], 0)}
            date_idx = col_idx.get("ng√†y", 0)
            ca_idx = col_idx.get("ca", 2)
            type_idx = col_idx.get("type of", 4)

            # summary[day][ca][type_of_short] = count
            summary = OrderedDict()
            for row in ws.iter_rows(min_row=2, values_only=True):
                day = row[date_idx]
                ca_raw = str(row[ca_idx]).strip().lower() if row[ca_idx] else ""
                ca = "HC" if "office" in ca_raw or ca_raw == "hc" else "OT"
                type_of_raw = (row[type_idx] or "UNKNOWN").strip().upper()
                if type_of_raw.startswith("OUTSOURCE"):
                    parts = type_of_raw.split("-")
                    type_of_short = (parts[-1][-3:] if len(parts) >= 2 else type_of_raw[-3:])
                else:
                    type_of_short = type_of_raw[:3]
                type_of_short = type_of_raw[:3]
                type_of_set.add(type_of_short)
                if day not in summary:
                    summary[day] = {"HC": defaultdict(int), "OT": defaultdict(int)}
                summary[day][ca][type_of_short] += 1

            # Gi·ªØ 10 ng√†y m·ªõi nh·∫•t
            day_keys = list(summary.keys())[-10:]
            summary = OrderedDict((k, summary[k]) for k in day_keys)
            type_of_list = sorted([t for t in type_of_set if t != "UNK"])
            if "UNK" in type_of_set:
                type_of_list.append("UNK")
            header = ["Ng√†y", "Ca"] + type_of_list + ["T·ªïng"]

            # T·∫°o rows cho template (2 d√≤ng/ng√†y: HC, OT)
            rows = []
            for day in summary:
                for ca in ("HC", "OT"):
                    type_counts = [summary[day][ca].get(t, 0) for t in type_of_list]
                    rows.append({
                        "date": day if ca == "HC" else "",
                        "ca": ca,
                        "types": type_counts,
                        "total": sum(type_counts)
                    })
        except Exception as e:
            # Log l·ªói n·∫øu c·∫ßn, nh∆∞ng tr·∫£ template b√¨nh th∆∞·ªùng
            print("[view_counter_log] Error:", e)
            rows = []
            type_of_list = []
    else:
        type_of_list = []

    return render_template(
        "counter_log.html",
        header=header,
        rows=rows,
        type_of_list=type_of_list,
    )

DISPLAY = {
    "hot_cold": "Hot & Cold cycle test",
    "standing_water": "Standing water test",
    "stain": "Stain test",
    "corrosion": "Corrosion test",
}

@app.route("/set_pref", methods=["POST"])
def set_pref():
    key = (request.json.get("key") or "").strip()
    value = str(request.json.get("value") or "").strip()

    if key == "darkmode":
        # Chu·∫©n h√≥a v·ªÅ "0"/"1"
        norm = "1" if value.lower() in ("1", "true", "on", "yes") else "0"
        session["darkmode"] = norm
        return jsonify({"success": True})

    if key == "lang":
        # Gi·ªõi h·∫°n m√£ ng√¥n ng·ªØ v√† ƒë·ªô d√†i
        norm = value.lower()[:5]
        if norm not in ("vi", "en"):
            norm = "vi"
        session["lang"] = norm
        return jsonify({"success": True})

    return jsonify({"success": False}), 400

KIOSK_TOKEN = os.environ.get("KIOSK_TOKEN") or ("kiosk-" + secrets.token_urlsafe(18))

def _kiosk_ok(req):
    """Ch·ªâ cho ph√©p truy c·∫≠p khi c√≥ token ƒë√∫ng (ƒë·∫∑t ?t=<token> tr√™n URL)."""
    return req.args.get("t") == KIOSK_TOKEN

# ------------------------------------------------------------------------------
# 2) B·ªô nh·ªõ ƒë·ªám d·ªØ li·ªáu cu·ªëi c√πng c·ªßa trang home.html (ƒë·ªÉ c·∫•p cho kiosk)
#    -> Kh√¥ng ƒë·ª•ng v√†o h√†m load d·ªØ li·ªáu c·ªßa b·∫°n, ch·ªâ nghe l√∫c template render.
# ------------------------------------------------------------------------------
_last_home_ctx = {
    "summary_by_type": [],           # d·∫°ng: [{"short":"TR","late":1,"due":2,"must":0,"active":5,"total":8}, ...]
    "report_list": [],               # d·∫°ng: [{"report":"25-xxxx","item":"...","type_of":"...","status":"DUE","log_date":"YYYY-MM-DD","etd":"YYYY-MM-DD"}, ...]
    "counter": {"office": 0, "ot": 0},  # {"office": <HC done>, "ot": <OT done>}
    "generated_at": None
}

def _extract_for_kiosk(context: dict):
    """
    T·ª´ context render c·ªßa home.html, r√∫t g·ªçn d·ªØ li·ªáu c·∫ßn cho kiosk.
    H√†m n√†y an to√†n n·∫øu thi·∫øu bi·∫øn (s·∫Ω d√πng default).
    """
    summary_by_type = context.get("summary_by_type") or []
    report_list     = context.get("report_list") or []
    counter         = context.get("counter") or {"office": 0, "ot": 0}

    # Chu·∫©n ho√° t·ª´ng ph·∫ßn t·ª≠ ƒë·ªÉ ƒë·∫£m b·∫£o key ƒë·∫ßy ƒë·ªß
    def _norm_summary(x):
        return {
            "short":  x.get("short", ""),
            "late":   int(x.get("late", 0) or 0),
            "due":    int(x.get("due", 0) or 0),
            "must":   int(x.get("must", 0) or 0),
            "active": int(x.get("active", 0) or 0),
            "total":  int(x.get("total", 0) or 0),
        }

    def _norm_report(r):
        return {
            "report":   r.get("report", "") or "",
            "item":     r.get("item", "") or "",
            "type_of":  r.get("type_of", "") or "",
            "status":   r.get("status", "") or "",
            "log_date": r.get("log_date", "") or "",
            "etd":      r.get("etd", "") or "",
        }

    norm_summary = [_norm_summary(x) for x in summary_by_type if isinstance(x, dict)]
    norm_reports = [_norm_report(r)  for r in report_list      if isinstance(r, dict)]
    norm_counter = {
        "office": int((counter or {}).get("office", 0) or 0),
        "ot":     int((counter or {}).get("ot", 0) or 0),
    }

    return {
        "summary_by_type": norm_summary,
        "report_list": norm_reports,
        "counter": norm_counter,
        "generated_at": datetime.now().isoformat(timespec="seconds")
    }

@template_rendered.connect_via(app)
def _capture_home_context(sender, template, context, **extra):
    """
    Nghe signal m·ªói khi Flask render template n√†o ƒë√≥.
    Khi template l√† 'home.html', l·∫•y nh·ªØng bi·∫øn c·∫ßn v√† cache v√†o _last_home_ctx.
    """
    try:
        # N·∫øu t√™n file template c·ªßa trang ch√≠nh kh√¥ng ph·∫£i 'home.html', ƒë·ªïi l·∫°i t·∫°i ƒë√¢y:
        if getattr(template, "name", None) == "home.html":
            data = _extract_for_kiosk(context or {})
            _last_home_ctx.update(copy.deepcopy(data))
    except Exception:
        # kh√¥ng ƒë·ªÉ l·ªói t·∫°i ƒë√¢y ph√° render c·ªßa app
        pass

# ------------------------------------------------------------------------------
# 3) API d·ªØ li·ªáu cho kiosk: /api/display_data?t=<KIOSK_TOKEN>
# ------------------------------------------------------------------------------
@app.route("/api/display_data")
def api_display_data():
    if not _kiosk_ok(request):
        abort(403)

    # N·∫øu mu·ªën fallback (khi server m·ªõi kh·ªüi ƒë·ªông, ch∆∞a render home l·∫ßn n√†o),
    # b·∫°n c√≥ th·ªÉ t·ª± g·ªçi h√†m load d·ªØ li·ªáu c·ªßa b·∫°n ·ªü ƒë√¢y, v√≠ d·ª•:
    #
    # try:
    #     from yourmodule import load_home_data    # n·∫øu b·∫°n c√≥ s·∫µn h√†m n√†y
    #     summary_by_type, report_list, counter = load_home_data()
    #     data = _extract_for_kiosk({
    #         "summary_by_type": summary_by_type,
    #         "report_list": report_list,
    #         "counter": counter
    #     })
    #     _last_home_ctx.update(copy.deepcopy(data))
    # except Exception:
    #     pass
    #
    # M·∫∑c ƒë·ªãnh s·∫Ω tr·∫£ v·ªÅ cache g·∫ßn nh·∫•t ƒë√£ b·∫Øt ƒë∆∞·ª£c khi render home.html

    return jsonify({
        "generated_at": _last_home_ctx.get("generated_at") or datetime.now().isoformat(timespec="seconds"),
        "summary": _last_home_ctx.get("summary_by_type", []),
        "reports": _last_home_ctx.get("report_list", []),
        "counter": _last_home_ctx.get("counter", {"office": 0, "ot": 0})
    })

# ------------------------------------------------------------------------------
# 4) Trang kiosk: /display?t=<KIOSK_TOKEN>&page_len=15&rotate_sec=60&refresh_sec=30&dark=1
# ------------------------------------------------------------------------------
@app.route("/display")
def display_board():
    if not _kiosk_ok(request):
        abort(403)

    # Tham s·ªë c·∫•u h√¨nh nhanh qua URL
    page_len    = int(request.args.get("page_len", 15))     # s·ªë d√≤ng m·ªói trang chi ti·∫øt
    rotate_sec  = int(request.args.get("rotate_sec", 60))   # l·∫≠t trang m·ªói X gi√¢y
    refresh_sec = int(request.args.get("refresh_sec", 30))  # n·∫°p l·∫°i d·ªØ li·ªáu m·ªói X gi√¢y
    dark        = request.args.get("dark", "1").lower() in ("1", "true", "yes")

    return render_template(
        "display.html",
        token=KIOSK_TOKEN,
        page_len=page_len,
        rotate_sec=rotate_sec,
        refresh_sec=refresh_sec,
        dark=dark
    )

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8246,debug=True)
