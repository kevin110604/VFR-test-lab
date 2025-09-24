import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import os
import io
import json
import re
from datetime import datetime, timedelta, date
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# SharePoint CSOM
from office365.sharepoint.client_context import ClientContext

# MSAL cho OAuth (Delegated + cache)
import msal

import sys
sys.stdout.reconfigure(encoding='utf-8')

# ================== CẤU HÌNH OAUTH (DELEGATED + CACHE) ==================
TENANT_ID = "064944f6-1e04-4050-b3e1-e361758625ec"       # Directory (tenant) ID
CLIENT_ID = "9abf6ee2-50c8-47c8-a9f2-8cf18587c6ea"       # Application (client) ID
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"

SP_HOST = "https://jonathancharles.sharepoint.com"

# Scope hợp lệ của SharePoint (Delegated)
SPO_SCOPES = [
    f"{SP_HOST}/AllSites.Read",
    f"{SP_HOST}/AllSites.Write",
]

TOKEN_CACHE_FILE = "token_cache.bin"


def _load_token_cache():
    cache = msal.SerializableTokenCache()
    if os.path.exists(TOKEN_CACHE_FILE):
        with open(TOKEN_CACHE_FILE, "r") as f:
            cache.deserialize(f.read())
    return cache


def _save_token_cache(cache: msal.SerializableTokenCache):
    if cache.has_state_changed:
        with open(TOKEN_CACHE_FILE, "w") as f:
            f.write(cache.serialize())


def acquire_spo_access_token() -> str:
    cache = _load_token_cache()
    app = msal.PublicClientApplication(
        client_id=CLIENT_ID,
        authority=AUTHORITY,
        token_cache=cache,
    )

    result = None
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SPO_SCOPES, account=accounts[0])

    if not result:
        flow = app.initiate_device_flow(scopes=SPO_SCOPES)
        if "user_code" not in flow:
            raise RuntimeError(f"Không khởi tạo được device flow: {json.dumps(flow, indent=2)}")
        print(flow["message"])
        result = app.acquire_token_by_device_flow(flow)

    _save_token_cache(cache)

    if "access_token" not in result:
        raise RuntimeError(f"Không lấy được access token: {result.get('error_description', str(result))}")

    return result["access_token"]


def get_ctx(site_url: str) -> ClientContext:
    ctx = ClientContext(site_url)

    def _auth(request):
        token = acquire_spo_access_token()
        request.ensure_header("Authorization", "Bearer " + token)

    ctx.authentication_context.authenticate_request = _auth
    print("AUTH MODE:", "Delegated (Device Code + Token Cache) [custom auth]")
    return ctx


# ================== HỖ TRỢ ==================
CURRENT_YEAR = datetime.now().year

def normalize_col(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def find_login_date_col(df: pd.DataFrame):
    targets = {
        "log in date", "login date", "log-in date", "logindate", "log in-date",
        "logged in date", "log_date", "log date"
    }
    cmap = {c: normalize_col(c) for c in df.columns}
    for orig, low in cmap.items():
        if low in targets:
            return orig
    for orig, low in cmap.items():
        if "log" in low and "date" in low:
            return orig
    return None

def hide_rows_by_login_date(ws, login_col_name_or_idx, today=None):
    if today is None:
        today = datetime.now().date()
    hide_threshold = today - timedelta(days=2)

    if isinstance(login_col_name_or_idx, int):
        login_col_idx = login_col_name_or_idx
    else:
        login_col_idx = None
        header = [cell.value for cell in ws[1]]
        for i, name in enumerate(header, start=1):
            if name and normalize_col(name) == normalize_col(login_col_name_or_idx):
                login_col_idx = i
                break
        if login_col_idx is None:
            for i, name in enumerate(header, start=1):
                if name and ("log" in normalize_col(name) and "date" in normalize_col(name)):
                    login_col_idx = i
                    break
    if not login_col_idx:
        return

    for r in range(2, ws.max_row + 1):
        raw = ws.cell(row=r, column=login_col_idx).value
        if not raw:
            continue
        try:
            dt = pd.to_datetime(str(raw), errors="coerce", dayfirst=True)
            if pd.isna(dt):
                dt = pd.to_datetime(str(raw), errors="coerce", dayfirst=False)
        except Exception:
            dt = None
        if dt is not None and not pd.isna(dt):
            if dt.date() <= hide_threshold:
                ws.row_dimensions[r].hidden = True

def ensure_folder(ctx, folder_url):
    folder_url = folder_url.rstrip("/")
    root_url = "/".join(folder_url.strip("/").split("/")[:4])
    parts = folder_url.strip("/").split("/")[4:]
    current_url = root_url
    for part in parts:
        current_url = current_url + "/" + part
        try:
            ctx.web.folders.add(current_url).execute_query()
        except Exception as e:
            if "already exists" not in str(e).lower() and "conflict" not in str(e).lower():
                print(f"Lỗi tạo folder {current_url}: {e}")
                raise
    return ctx.web.get_folder_by_server_relative_url(folder_url)

def _parse_to_datetime_or_none(text: str):
    s = (text or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        dt = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if pd.isna(dt):
        return None
    return pd.Timestamp(dt).to_pydatetime()

def format_date_columns(ws, target_headers=("log in date", "etd"), number_format="DD-MMM"):
    header_cells = [cell.value for cell in ws[1]]
    target_norm = {normalize_col(h) for h in target_headers}
    target_col_indices = []
    for idx, name in enumerate(header_cells, start=1):
        if name and normalize_col(name) in target_norm:
            target_col_indices.append(idx)

    for col_idx in target_col_indices:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if isinstance(val, datetime):
                cell.number_format = number_format
            elif isinstance(val, date):
                cell.value = datetime(val.year, val.month, val.day)
                cell.number_format = number_format
            elif isinstance(val, (int, float)):
                cell.number_format = number_format
            elif isinstance(val, str):
                dt = _parse_to_datetime_or_none(val)
                if dt is not None:
                    cell.value = dt
                    cell.number_format = number_format

# ==== Cấu hình SharePoint ====
site_url = f"{SP_HOST}/sites/TESTLAB-VFR9"
ctx = get_ctx(site_url)

# ================= PHẦN DUY NHẤT: TRF + COMPLETED =================
output_file = "TRF_complete.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # Completed
    if os.path.exists("completed_items.xlsx"):
        df_cpl = pd.read_excel("completed_items.xlsx", dtype=str)
        df_cpl.to_excel(writer, sheet_name="Completed", index=False)
    else:
        print("⚠️ Không tìm thấy completed_items.xlsx")

    # TRF
    if os.path.exists("TRF.xlsx"):
        df_trf = pd.read_excel("TRF.xlsx", dtype=str)
        df_trf.to_excel(writer, sheet_name="TRF", index=False)
    else:
        print("⚠️ Không tìm thấy TRF.xlsx")

# ===== Định dạng sau khi ghi =====
wb = load_workbook(output_file)
thin = Side(border_style="thin", color="888888")
header_fill = PatternFill("solid", fgColor="B7E1CD")

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

    # Format cột ngày
    format_date_columns(ws, target_headers=("log in date", "etd"), number_format="DD-MMM")

    # Hide rows theo login date
    df_tmp = pd.DataFrame(ws.values)
    login_col = find_login_date_col(df_tmp)
    if login_col:
        hide_rows_by_login_date(ws, login_col, today=datetime.now().date())

wb.save(output_file)
print("✅ Đã xuất file:", output_file)

# ===== Upload SharePoint =====
upload_relative_url = "/sites/TESTLAB-VFR9/Shared Documents/DATA DAILY/TRF_complete.xlsx"
folder_excel = os.path.dirname(upload_relative_url)
ensure_folder(ctx, folder_excel)
with open(output_file, "rb") as f:
    ctx.web.get_folder_by_server_relative_url(folder_excel)\
       .upload_file(os.path.basename(upload_relative_url), f.read()).execute_query()
print("🚀 Đã upload file", output_file, "lên SharePoint:", upload_relative_url)