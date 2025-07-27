import os
import json
import pytz
import openpyxl
from datetime import datetime, time

COUNT_JSON = "counter_stats.json"
DETAIL_XLSX = "counter_detail_log.xlsx"
SHIFT_END = time(16, 45)  # 16:45, hết ca

def read_counter():
    today_str = datetime.now(pytz.timezone("Asia/Ho_Chi_Minh")).strftime("%Y-%m-%d")
    if os.path.exists(COUNT_JSON):
        with open(COUNT_JSON, "r", encoding="utf-8") as f:
            counter = json.load(f)
        if counter.get("date") != today_str:
            counter = {"date": today_str, "office": 0, "ot": 0}
    else:
        counter = {"date": today_str, "office": 0, "ot": 0}
    return counter

def write_counter_json(counter):
    with open(COUNT_JSON, "w", encoding="utf-8") as f:
        json.dump(counter, f, ensure_ascii=False)

def update_counter(ca=None):
    now = datetime.now(pytz.timezone("Asia/Ho_Chi_Minh"))
    today_str = now.strftime("%Y-%m-%d")
    tval = now.hour * 60 + now.minute

    office_start = 8 * 60
    office_end = 16 * 60 + 45
    ot_end = 23 * 60 + 59

    counter = read_counter()
    # Reset nếu qua ngày mới
    if counter.get("date") != today_str:
        counter = {"date": today_str, "office": 0, "ot": 0}

    # Xác định ca nếu không truyền vào
    if not ca:
        if office_start <= tval < office_end:
            ca = "office"
        elif office_end <= tval <= ot_end:
            ca = "ot"
        else:
            ca = None

    if ca in ("office", "ot"):
        counter[ca] += 1

    write_counter_json(counter)
    return counter

def log_report_complete(report_number, type_of, ca, employee_id=""):
    from openpyxl import load_workbook, Workbook
    import os

    excel_path = DETAIL_XLSX
    vn_tz = pytz.timezone("Asia/Ho_Chi_Minh")
    now = datetime.now(vn_tz)
    date_str = now.strftime("%d/%m/%Y")

    # Nếu file chưa có thì tạo mới và ghi header
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Ngày", "Ca", "Type of", "Report#", "ID"])
        wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active

    # Tìm dòng đã có report_number VÀ cùng ngày, cùng ca
    found_row = None
    for row in range(2, ws.max_row + 1):
        day = ws.cell(row=row, column=1).value
        ca_val = ws.cell(row=row, column=2).value
        report_val = ws.cell(row=row, column=4).value
        if (str(day) == date_str) and (str(ca_val) == ca) and (str(report_val).strip().upper() == str(report_number).strip().upper()):
            found_row = row
            break

    # Nếu đã tồn tại report# (cùng ngày và ca) thì update lại dòng đó
    if found_row:
        ws.cell(row=found_row, column=1).value = date_str
        ws.cell(row=found_row, column=2).value = ca
        ws.cell(row=found_row, column=3).value = type_of
        ws.cell(row=found_row, column=4).value = report_number
        ws.cell(row=found_row, column=5).value = employee_id
    else:
        ws.append([date_str, ca, type_of, report_number, employee_id])

    wb.save(excel_path)

def check_and_reset_counter():
    now = datetime.now(pytz.timezone("Asia/Ho_Chi_Minh"))
    today_str = now.strftime("%Y-%m-%d")
    counter = read_counter()
    need_reset = False

    if counter.get("date") != today_str:
        counter = {"date": today_str, "office": 0, "ot": 0}
        need_reset = True

    if need_reset:
        write_counter_json(counter)
