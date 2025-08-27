# qr_print_v10.py — Center header to QR + Back button in template
# • REPORT: small, not bold, single line (truncate)
# • Type of: regular, wraps as needed
# • Entire header block vertically centers relative to QR using font metrics
# • Body always starts below both header & QR; extra left padding
# • Remove Submitter* and QR* columns; aligned "Label: Value"
# • Unicode font if available, else auto strip accents
#
# Usage:
#   from qr_print_v10 import qr_bp
#   app.register_blueprint(qr_bp)

import io, os, re, unicodedata
import datetime as dt
from flask import Blueprint, request, send_file, render_template_string, session, current_app
import openpyxl
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

qr_bp = Blueprint("qr_bp", __name__)

# ---- Config ----
PAGE_W_MM    = 50
PAGE_H_MM    = 65
M_MM         = 2
QR_SIZE_MM   = 14

HDR_FONT_REPORT = 8.6   # REPORT font (smaller, not bold)
HDR_FONT_TYPEOF = 9.2   # Type of font (regular, can wrap)
HDR_LINE_GAP    = 11.0  # baseline-to-baseline gap for header lines (pt)
HDR_BOTTOM_GAP_MM = 2.0 # gap below header block before body

BODY_FONT_SZ    = 7.0
BODY_LINE_GAP   = 7.4
LABEL_RATIO     = 0.40
BODY_LEFT_PAD_MM= 1.4

ALWAYS_ASCII          = False
FORCE_ASCII_IF_NOFONT = True

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# ---- Helpers ----
def _strip_accents(s: str) -> str:
    try:
        s = unicodedata.normalize("NFD", s)
        s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
        s = unicodedata.normalize("NFC", s)
        return s
    except Exception:
        return s

def _to_ascii_if_needed(s: str, ascii_only: bool) -> str:
    s = re.sub(r"\s+", " ", str(s)).strip()
    return _strip_accents(s) if ascii_only else s

def _norm_header(h: str) -> str:
    h = re.sub(r"\s+", " ", str(h)).strip().lower()
    h = _strip_accents(h)
    return re.sub(r"[\s:_\-\./]+", "", h)

def _excel_col_label(idx_zero_based: int) -> str:
    idx = idx_zero_based
    label = ""
    while True:
        idx, rem = divmod(idx, 26)
        label = chr(65 + rem) + label
        if idx == 0:
            break
        idx -= 1
    return label

def _get_trf_path():
    try:
        cfg = current_app.config.get("TRF_XLSX_PATH")
        if cfg:
            return cfg
    except Exception:
        pass
    env = os.environ.get("TRF_XLSX_PATH")
    if env:
        return env
    try:
        app_root = current_app.root_path
        cand = os.path.join(app_root, "TRF.xlsx")
        if os.path.exists(cand):
            return cand
    except Exception:
        pass
    return os.path.join(BASE_DIR, "TRF.xlsx")

def _try_register_fonts():
    candidates = [
        (os.path.join(BASE_DIR, "static", "fonts", "DejaVuSans.ttf"),
         os.path.join(BASE_DIR, "static", "fonts", "DejaVuSans-Bold.ttf"),
         "DejaVuSans", "DejaVuSans-Bold"),
        (r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\arialbd.ttf", "Arial", "Arial-Bold"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "DejaVuSans", "DejaVuSans-Bold"),
        ("/Library/Fonts/Arial.ttf", "/Library/Fonts/Arial Bold.ttf", "Arial", "Arial-Bold"),
    ]
    for reg, bold, rname, bname in candidates:
        if os.path.exists(reg) and os.path.exists(bold):
            try:
                pdfmetrics.registerFont(TTFont(rname, reg))
                pdfmetrics.registerFont(TTFont(bname, bold))
                return (rname, bname, True)
            except Exception:
                pass
    return ("Helvetica", "Helvetica-Bold", False)

def _parse_report_code(code):
    m = re.match(r"^\s*(\d{2})\s*-\s*(\d+)\s*$", str(code or ""))
    if not m:
        return None
    return (int(m.group(1)), int(m.group(2)))

def _safe_str(v, ascii_only=False):
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%d-%b-%Y")
    return _to_ascii_if_needed(v, ascii_only)

def _load_trf_table(xlsx_path):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Không thấy file TRF.xlsx tại: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = None
    for name in ("TRF", "Sheet1"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    headers = [cell.value for cell in ws[1]]
    headers = [h if h is not None else "" for h in headers]
    norm_map = {_norm_header(str(h)): i for i, h in enumerate(headers)}

    idx_report = None
    for k in ("report", "reportno", "reportnumber", "report#"):
        if k in norm_map:
            idx_report = norm_map[k]
            break
    idx_typeof = None
    for k in ("typeof", "type", "type_of", "typeof:"):
        if k in norm_map:
            idx_typeof = norm_map[k]
            break

    rows_values = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows_values.append([r[i] if i < len(r) else None for i in range(len(headers))])

    return headers, rows_values, idx_report, idx_typeof

def _filter_rows_by_range_table(headers, rows_values, idx_report, start_code, end_code):
    a = _parse_report_code(start_code)
    b = _parse_report_code(end_code)
    if not a or not b or idx_report is None:
        return []
    if a > b:
        a, b = b, a

    kept = []
    for row in rows_values:
        rep_str = _safe_str(row[idx_report], ascii_only=False)
        t = _parse_report_code(rep_str)
        if t and (a <= t <= b):
            kept.append(row)
    kept.sort(key=lambda r: _parse_report_code(_safe_str(r[idx_report])) or (0, 0))
    return kept

def _wrap_no_split(cvs, text, max_w, font, size):
    words = re.sub(r"\s+", " ", str(text)).strip().split()
    if not words:
        return []
    lines, current = [], ""
    for w in words:
        candidate = (current + " " + w).strip()
        if cvs.stringWidth(candidate, font, size) <= max_w:
            current = candidate
        else:
            if current:
                lines.append(current)
                current = ""
            if cvs.stringWidth(w, font, size) > max_w:
                base, ell = "", "…"
                ell_w = cvs.stringWidth(ell, font, size)
                for ch in w:
                    if cvs.stringWidth(base + ch, font, size) + ell_w <= max_w:
                        base += ch
                    else:
                        break
                lines.append(base + ell if base else w[:1] + ell)
            else:
                current = w
    if current:
        lines.append(current)
    return lines

def _is_qr_header(header: str) -> bool:
    h = _norm_header(header)
    return h in {
        "qr", "qrcode", "qrcodeimage", "qrimage", "qrlink", "qrurl",
        "qrcodeurl", "qrcodeink", "qrcodepath"
    } or header.strip().lower() in {"qr code", "qr-code"}

def _is_submitter_header(header: str) -> bool:
    h = _norm_header(header)
    return h.startswith("submitter")

def _rank_header_for_beauty(header_ascii: str) -> int:
    h = re.sub(r"[^a-z0-9]+", "", header_ascii.lower())
    if h in {"trqid", "trq", "trqnumber", "trqno", "trqidcode"}:
        return 0
    if h in {"item", "itemno", "itemnumber", "itemcode", "item#", "code"}:
        return 1
    if h in {"itemname", "description", "itemnamedescription"}:
        return 2
    if h in {"furnituretesting", "testing", "testtype"}:
        return 3
    if h in {"submitteddept", "department", "dept"}:
        return 4
    if h in {"remark", "remarks", "note", "notes", "purpose"}:
        return 5
    if h in {"etd", "logindate", "requestdate", "duedate"}:
        return 6
    if h in {"qacomment", "qa"}:
        return 7
    return 10

def _truncate_to_width(cvs, text, max_w, font, size):
    s = re.sub(r"\s+", " ", str(text)).strip()
    if cvs.stringWidth(s, font, size) <= max_w:
        return s
    ell = "…"
    ell_w = cvs.stringWidth(ell, font, size)
    out = ""
    for ch in s:
        if cvs.stringWidth(out + ch, font, size) + ell_w <= max_w:
            out += ch
        else:
            break
    return out + ell if out else s[:1] + ell

# ---- PDF builder ----
def _build_qr_label_pdf(headers, table_rows, idx_report, idx_typeof):
    buf = io.BytesIO()

    PAGE_W = PAGE_W_MM * mm
    PAGE_H = PAGE_H_MM * mm
    cvs = canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    font_reg, font_bold, unicode_ok = _try_register_fonts()
    ascii_only = ALWAYS_ASCII or (FORCE_ASCII_IF_NOFONT and not unicode_ok)

    M = M_MM * mm
    QR_SIZE = QR_SIZE_MM * mm
    info_x = M + QR_SIZE + (2 * mm)
    info_w = PAGE_W - info_x - M

    # Display header names and printable indices
    display_headers = []
    seen = {}
    for i, h in enumerate(headers):
        disp = _to_ascii_if_needed(h if h else f"Col {_excel_col_label(i)}", ascii_only)
        cnt = seen.get(disp, 0) + 1
        seen[disp] = cnt
        if cnt > 1:
            disp = f"{disp} ({cnt})"
        display_headers.append(disp)

    printable_indices = []
    for i, h in enumerate(headers):
        if i == idx_report or i == idx_typeof:
            continue
        if _is_qr_header(str(h)) or _is_submitter_header(str(h)):
            continue
        printable_indices.append(i)
    printable_indices.sort(key=lambda i: (_rank_header_for_beauty(display_headers[i]), i))

    # Pre-calc font metrics for accurate centering
    ascent_rep = pdfmetrics.getAscent(font_reg) / 1000.0 * HDR_FONT_REPORT
    descent_type = abs(pdfmetrics.getDescent(font_reg)) / 1000.0 * HDR_FONT_TYPEOF

    for row in table_rows:
        # Border
        cvs.setLineWidth(0.5)
        cvs.setDash(1, 2)
        cvs.rect(0.8 * mm, 0.8 * mm, PAGE_W - 1.6 * mm, PAGE_H - 1.6 * mm)

        # QR
        report_val = _safe_str(row[idx_report], ascii_only=ascii_only) if idx_report is not None else ""
        url = f"http://103.77.166.187:8246/update?report={report_val}"
        qr_obj = qrcode.make(url, box_size=3, border=1)
        try:
            pil_img = qr_obj.get_image()
        except AttributeError:
            pil_img = qr_obj
        try:
            if getattr(pil_img, "mode", "RGB") not in ("RGB", "RGBA", "L"):
                pil_img = pil_img.convert("RGB")
        except Exception:
            pass
        img_buf = io.BytesIO()
        pil_img.save(img_buf, format="PNG")
        img_buf.seek(0)
        qr_reader = ImageReader(img_buf)
        qr_x = M
        qr_y = PAGE_H - M - QR_SIZE
        cvs.drawImage(qr_reader, qr_x, qr_y, width=QR_SIZE, height=QR_SIZE,
                      preserveAspectRatio=True, mask="auto")

        # Header block: vertically center to QR using metrics
        cvs.setFont(font_reg, HDR_FONT_REPORT)
        rep_line = _to_ascii_if_needed(f"REPORT: {report_val}", ascii_only)
        rep_fit = _truncate_to_width(cvs, rep_line, info_w, font_reg, HDR_FONT_REPORT)

        cvs.setFont(font_reg, HDR_FONT_TYPEOF)
        typeof_val = _safe_str(row[idx_typeof], ascii_only=ascii_only) if idx_typeof is not None else ""
        type_line = _to_ascii_if_needed(f"Type of: {typeof_val}" if typeof_val else "Type of: -", ascii_only)
        type_lines = _wrap_no_split(cvs, type_line, info_w, font_reg, HDR_FONT_TYPEOF)
        N = max(1, len(type_lines))

        # Height from top of REPORT (ascent above baseline) to bottom of last Type (descent below baseline)
        block_h = ascent_rep + N * HDR_LINE_GAP + descent_type

        # Desired block center equals QR center
        qr_center_y = qr_y + QR_SIZE / 2.0

        # Solve for report baseline Y0 so that center(top_edge..bottom_edge) == qr_center_y
        # top_edge = Y0 + ascent_rep; bottom_edge = (Y0 - N*HDR_LINE_GAP) - descent_type
        # => center = (top_edge + bottom_edge)/2 = Y0 - (N*HDR_LINE_GAP)/2 + (ascent_rep - descent_type)/2
        Y0 = qr_center_y + (N * HDR_LINE_GAP) / 2.0 - (ascent_rep - descent_type) / 2.0

        # Clamp top edge to top margin
        top_edge = Y0 + ascent_rep
        max_top = PAGE_H - M - 6
        if top_edge > max_top:
            shift = top_edge - max_top
            Y0 -= shift
        # Clamp bottom edge to be above body minimum area if extreme (rare)
        bottom_edge = (Y0 - N * HDR_LINE_GAP) - descent_type
        min_bottom = M + 12  # keep a little breathing space
        if bottom_edge < min_bottom:
            shift = min_bottom - bottom_edge
            Y0 += shift

        # Draw REPORT and Type of
        y_right = Y0
        cvs.setFont(font_reg, HDR_FONT_REPORT)
        cvs.drawString(info_x, y_right, rep_fit)

        cvs.setFont(font_reg, HDR_FONT_TYPEOF)
        for ln in type_lines:
            y_right -= HDR_LINE_GAP
            cvs.drawString(info_x, y_right, ln)

        # Body start strictly below both QR and header
        body_start = min(qr_y, y_right) - (HDR_BOTTOM_GAP_MM * mm)
        y = body_start

        # Body columns
        body_w = PAGE_W - 2*M - (BODY_LEFT_PAD_MM * mm)
        label_w = body_w * LABEL_RATIO
        value_w = body_w - label_w
        body_x = M + (BODY_LEFT_PAD_MM * mm)

        cvs.setFont(font_reg, BODY_FONT_SZ)
        for i in printable_indices:
            sval = _safe_str(row[i], ascii_only=ascii_only)
            if not sval:
                continue
            label = display_headers[i] + ":"
            label = _truncate_to_width(cvs, label, label_w, font_reg, BODY_FONT_SZ)

            val_lines = _wrap_no_split(cvs, sval, value_w, font_reg, BODY_FONT_SZ)
            needed = max(1, len(val_lines)) * BODY_LINE_GAP
            if y - needed < (M + 9 * mm):
                break

            cvs.drawString(body_x, y, label)
            vy = y
            for ln in val_lines:
                cvs.drawString(body_x + label_w + 1.6*mm, vy, ln)
                vy -= BODY_LINE_GAP
            y = vy

        # Notes
        cvs.setDash(1, 0)
        cvs.setLineWidth(0.7)
        note_y1 = M + 8 * mm
        note_y2 = M + 4 * mm
        cvs.line(body_x, note_y2, PAGE_W - M, note_y2)
        cvs.line(body_x, note_y1, PAGE_W - M, note_y1)
        cvs.setFont(font_reg, 7.0)
        cvs.drawString(body_x, note_y1 + 2.0 * mm, _to_ascii_if_needed("Notes / Ghi chú:", ascii_only))

        cvs.showPage()

    cvs.save()
    buf.seek(0)
    return buf

# ---- Route ----
@qr_bp.route("/print_qr", methods=["GET", "POST"])
def print_qr():
    if request.method == "GET":
        return render_template_string("""
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <title>Print QR</title>
  <style>
    :root{--bg:#fffbe8;--border:#ffe082;--accent:#b78e24;--btn:#c49c48;--btnh:#35523A}
    *{box-sizing:border-box}
    body{font-family:Segoe UI,Arial,sans-serif;background:var(--bg);color:#333;margin:0}
    a.back{position:fixed;top:10px;left:10px;background:#fff;border:1px solid var(--border);border-radius:8px;padding:6px 10px;text-decoration:none;color:#4b3e0f;font-weight:700;box-shadow:0 2px 8px #c49c4840}
    a.back:hover{background:#35523A;color:#ffe082}
    .wrap{max-width:720px;margin:36px auto;background:#fff;border:1.5px solid var(--border);border-radius:14px;box-shadow:0 8px 24px #c49c4840;padding:22px}
    h2{margin:0 0 10px;color:var(--accent)}
    p.sub{margin:0 0 16px;color:#6f5b18}
    .row{display:flex;gap:12px;flex-wrap:wrap}
    .col{flex:1 1 280px}
    label{display:block;font-weight:700;margin:10px 0 6px}
    input[type=text]{width:100%;padding:10px 12px;border-radius:9px;border:1.5px solid var(--border);font-size:16px}
    .actions{display:flex;gap:10px;margin-top:14px;flex-wrap:wrap}
    button{background:var(--btn);color:#fff;border:none;border-radius:10px;padding:10px 18px;font-weight:800;cursor:pointer}
    button:hover{background:var(--btnh);color:#ffe082}
    .hint{font-size:13px;color:#7a6520;margin-top:6px}
  </style>
</head>
<body>
  <a class="back" href="/">← Back</a>
  <div class="wrap">
    <h2>Print QR Code (50×65mm)</h2>
    <form method="POST">
      <div class="row">
        <div class="col">
          <label>Start (Ex: 25-5000)</label>
          <input type="text" name="start_report" required placeholder="25-5000">
        </div>
        <div class="col">
          <label>Finish (Ex: 25-5050)</label>
          <input type="text" name="end_report" placeholder="25-5050">
        </div>
      </div>
      <div class="actions">
        <button type="submit">Create PDF</button>
      </div>
    </form>
  </div>
</body>
</html>
        """)

    start_code = (request.form.get("start_report") or "").strip()
    end_code   = (request.form.get("end_report") or "").strip() or start_code
    if not start_code:
        return "Thiếu mã bắt đầu", 400

    xlsx_path = _get_trf_path()
    try:
        headers, rows_values, idx_report, idx_typeof = _load_trf_table(xlsx_path)
    except Exception as e:
        return f"Lỗi đọc TRF.xlsx: {e}", 500

    if idx_report is None:
        return "Không tìm thấy cột REPORT trong TRF.xlsx", 500

    chosen_rows = _filter_rows_by_range_table(headers, rows_values, idx_report, start_code, end_code)
    if not chosen_rows:
        return f"Không tìm thấy REPORT trong khoảng [{start_code}..{end_code}]", 404

    pdf = _build_qr_label_pdf(headers, chosen_rows, idx_report, idx_typeof)
    fname = f"QR_{start_code}_to_{end_code}.pdf"
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name=fname)
