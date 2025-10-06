import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import os
import io
import json
import re
from datetime import datetime, timedelta, date
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# SharePoint CSOM
from office365.sharepoint.client_context import ClientContext

# MSAL cho OAuth (Delegated + cache)
import msal

# ================== C·∫§U H√åNH OAUTH (DELEGATED + CACHE) ==================
TENANT_ID = "064944f6-1e04-4050-b3e1-e361758625ec"       # Directory (tenant) ID
CLIENT_ID = "9abf6ee2-50c8-47c8-a9f2-8cf18587c6ea"       # Application (client) ID
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"

SP_HOST = "https://jonathancharles.sharepoint.com"

# Scope h·ª£p l·ªá c·ªßa SharePoint (Delegated) - KH√îNG d√πng openid/profile/offline_access ·ªü ƒë√¢y
SPO_SCOPES = [
    f"{SP_HOST}/AllSites.Read",
    f"{SP_HOST}/AllSites.Write",
]

TOKEN_CACHE_FILE = "token_cache.bin"


def _load_token_cache():
    cache = msal.SerializableTokenCache()
    if os.path.exists(TOKEN_CACHE_FILE):
        with open(TOKEN_CACHE_FILE, "r") as f:
            cache.deserialize(f.read())
    return cache


def _save_token_cache(cache: msal.SerializableTokenCache):
    if cache.has_state_changed:
        with open(TOKEN_CACHE_FILE, "w") as f:
            f.write(cache.serialize())


def acquire_spo_access_token() -> str:
    """
    L·∫•y access token cho SharePoint Online (Delegated).
    - ∆Øu ti√™n: acquire_token_silent() t·ª´ cache
    - N·∫øu ch∆∞a c√≥/expired: Device Code Flow (in m√£ ra console), login 1 l·∫ßn
    - Cache t·ª± l∆∞u ƒë·ªÉ l·∫ßn sau silent
    """
    cache = _load_token_cache()
    app = msal.PublicClientApplication(
        client_id=CLIENT_ID,
        authority=AUTHORITY,
        token_cache=cache,
    )

    result = None
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SPO_SCOPES, account=accounts[0])

    if not result:
        flow = app.initiate_device_flow(scopes=SPO_SCOPES)
        if "user_code" not in flow:
            raise RuntimeError(f"Kh√¥ng kh·ªüi t·∫°o ƒë∆∞·ª£c device flow: {json.dumps(flow, indent=2)}")
        # H∆∞·ªõng d·∫´n login 1 l·∫ßn (copy link + code ra tr√¨nh duy·ªát)
        print(flow["message"])
        result = app.acquire_token_by_device_flow(flow)

    _save_token_cache(cache)

    if "access_token" not in result:
        raise RuntimeError(f"Kh√¥ng l·∫•y ƒë∆∞·ª£c access token: {result.get('error_description', str(result))}")

    return result["access_token"]


def get_ctx(site_url: str) -> ClientContext:
    """
    ClientContext s·ª≠ d·ª•ng custom authenticate_request:
    - M·ªói request t·ª± g·∫Øn 'Authorization: Bearer <token>' l·∫•y t·ª´ MSAL cache
    - Tr√°nh ho√†n to√†n bug timezone (naive/aware) trong lib
    """
    ctx = ClientContext(site_url)

    def _auth(request):
        token = acquire_spo_access_token()  # l·∫•y t·ª´ cache, t·ª± refresh khi c·∫ßn
        request.ensure_header("Authorization", "Bearer " + token)

    # Ghi ƒë√® c∆° ch·∫ø auth m·∫∑c ƒë·ªãnh
    ctx.authentication_context.authenticate_request = _auth
    print("AUTH MODE:", "Delegated (Device Code + Token Cache) [custom auth]")
    return ctx


# ================== PH·∫¶N SHAREPOINT / X·ª¨ L√ù EXCEL ==================
CURRENT_YEAR = datetime.now().year

def normalize_col(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def find_login_date_col(df: pd.DataFrame):
    targets = {
        "log in date", "login date", "log-in date", "logindate", "log in-date",
        "logged in date", "log_date", "log date"
    }
    cmap = {c: normalize_col(c) for c in df.columns}
    for orig, low in cmap.items():
        if low in targets:
            return orig
    for orig, low in cmap.items():
        if "log" in low and "date" in low:
            return orig
    return None

_has_year_pat = re.compile(r"\b(\d{4}|\d{2})\b")
_missing_year_pat = re.compile(r"^\s*(\d{1,2})[./\- ]([A-Za-z]{3}|\d{1,2})\s*$")
MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct":10, "nov":11, "dec":12
}

def _month_to_int(token: str):
    t = token.strip().lower()
    if t.isdigit():
        m = int(t)
        return m if 1 <= m <= 12 else None
    return MONTH_MAP.get(t[:3])

def _attach_current_year_if_missing(text: str) -> str:
    s = (text or "").strip()
    if not s:
        return s
    if _has_year_pat.search(s):
        return s
    m = _missing_year_pat.match(s)
    if m:
        mon_token = m.group(2)
        mon = _month_to_int(mon_token)
        year = CURRENT_YEAR if mon != 12 else CURRENT_YEAR - 1
        if mon_token.isdigit():
            s2 = re.sub(r"[.\- ]", "/", s)
            return f"{s2}/{year}"
        else:
            return f"{s}-{year}"
    return f"{s} {CURRENT_YEAR}"

def parse_login_dates(series: pd.Series) -> pd.Series:
    """
    ∆Øu ti√™n parse ISO '%Y-%m-%d %H:%M:%S' (kh√¥ng c·∫£nh b√°o),
    n·∫øu kh√¥ng ƒë∆∞·ª£c th√¨ fallback dayfirst True/False.
    """
    tmp = series.astype(str).map(_attach_current_year_if_missing)

    # ∆Øu ti√™n parse ISO chu·∫©n
    dt = pd.to_datetime(tmp, errors="coerce", format="%Y-%m-%d %H:%M:%S", exact=False)
    if dt.isna().any():
        # Th·ª≠ parse generic v·ªõi dayfirst
        dt2 = pd.to_datetime(tmp, errors="coerce", dayfirst=True, infer_datetime_format=True)
        if dt2.notna().sum() > dt.notna().sum():
            dt = dt2
        else:
            # fallback cu·ªëi c√πng
            dt3 = pd.to_datetime(tmp, errors="coerce", dayfirst=False, infer_datetime_format=True)
            if dt3.notna().sum() > dt.notna().sum():
                dt = dt3
    return dt.fillna(pd.Timestamp(1900, 1, 1))

def hide_rows_by_login_date(ws, login_col_name_or_idx, today=None):
    """
    ·∫®n c√°c d√≤ng c√≥ ng√†y ƒëƒÉng nh·∫≠p (login date) <= today - 2 ng√†y.
    Parse ng√†y theo th·ª© t·ª±: ISO -> dayfirst=True -> dayfirst=False.
    """
    if today is None:
        today = datetime.now().date()
    hide_threshold = today - timedelta(days=2)
    if isinstance(login_col_name_or_idx, int):
        login_col_idx = login_col_name_or_idx
    else:
        login_col_idx = None
        header = [cell.value for cell in ws[1]]
        for i, name in enumerate(header, start=1):
            if name and normalize_col(name) == normalize_col(login_col_name_or_idx):
                login_col_idx = i
                break
        if login_col_idx is None:
            for i, name in enumerate(header, start=1):
                if name and ("log" in normalize_col(name) and "date" in normalize_col(name)):
                    login_col_idx = i
                    break
    if not login_col_idx:
        return
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(row=r, column=login_col_idx).value
        raw_str = "" if raw is None else str(raw).strip()
        if not raw_str:
            continue
        s_with_year = _attach_current_year_if_missing(raw_str)

        # Try ISO first
        try:
            dt = pd.to_datetime(s_with_year, errors="coerce", format="%Y-%m-%d %H:%M:%S")
        except Exception:
            dt = None
        if dt is None or pd.isna(dt):
            dt = pd.to_datetime(s_with_year, errors="coerce", dayfirst=True)
        if pd.isna(dt):
            dt = pd.to_datetime(s_with_year, errors="coerce", dayfirst=False)
        if pd.isna(dt):
            continue
        if dt.date() <= hide_threshold:
            ws.row_dimensions[r].hidden = True

def ensure_folder(ctx, folder_url):
    folder_url = folder_url.rstrip("/")
    root_url = "/".join(folder_url.strip("/").split("/")[:4])
    parts = folder_url.strip("/").split("/")[4:]
    current_url = root_url
    for part in parts:
        current_url = current_url + "/" + part
        try:
            ctx.web.folders.add(current_url).execute_query()
        except Exception as e:
            if "already exists" not in str(e).lower() and "conflict" not in str(e).lower():
                print(f"Loi tao folder {current_url}: {e}")
                raise
    return ctx.web.get_folder_by_server_relative_url(folder_url)

# ====== H·ªñ TR·ª¢: ƒê·ªãnh d·∫°ng c·ªôt ng√†y th√†nh dd-mmm cho openpyxl ======
def _parse_to_datetime_or_none(text: str):
    """
    C·ªë g·∫Øng parse chu·ªói sang datetime (kh√¥ng timezone).
    ∆Øu ti√™n ISO '%Y-%m-%d %H:%M:%S', sau ƒë√≥ th·ª≠ dayfirst True/False.
    """
    s = (text or "").strip()
    if not s:
        return None
    # ISO tr∆∞·ªõc
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    # Pandas fallback
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        dt = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if pd.isna(dt):
        return None
    return pd.Timestamp(dt).to_pydatetime()

def format_date_columns(ws, target_headers=("log in date", "etd"), number_format="DD-MMM"):
    """
    - T√¨m c√°c c·ªôt header n·∫±m trong target_headers (so s√°nh normalize).
    - V·ªõi t·ª´ng √¥ d·ªØ li·ªáu:
        + N·∫øu l√† datetime/date: set number_format = 'DD-MMM'
        + N·∫øu l√† chu·ªói: th·ª≠ parse; n·∫øu parse ƒë∆∞·ª£c -> g√°n datetime + number_format
        + N·∫øu ƒë√£ ƒë√∫ng ƒë·ªãnh d·∫°ng th√¨ set l·∫°i number_format (idempotent, an to√†n)
    """
    header_cells = [cell.value for cell in ws[1]]
    target_norm = {normalize_col(h) for h in target_headers}
    target_col_indices = []
    for idx, name in enumerate(header_cells, start=1):
        if name and normalize_col(name) in target_norm:
            target_col_indices.append(idx)

    for col_idx in target_col_indices:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is None or (isinstance(val, str) and not val.strip()):
                continue
            if isinstance(val, datetime):
                # ƒë√£ l√† datetime -> ch·ªâ c·∫ßn set number_format
                cell.number_format = number_format
                continue
            if isinstance(val, date):
                # date -> cast sang datetime ƒë·ªÉ Excel hi·ªÉn th·ªã ƒë√∫ng
                cell.value = datetime(val.year, val.month, val.day)
                cell.number_format = number_format
                continue
            if isinstance(val, (int, float)):
                # c√≥ th·ªÉ l√† serial date; ƒë·ªÉ nguy√™n, ch·ªâ set number_format
                cell.number_format = number_format
                continue
            # N·∫øu l√† chu·ªói -> c·ªë parse
            if isinstance(val, str):
                dt = _parse_to_datetime_or_none(val)
                if dt is not None:
                    cell.value = dt
                    cell.number_format = number_format
                # n·∫øu kh√¥ng parse ƒë∆∞·ª£c th√¨ b·ªè qua (gi·ªØ nguy√™n)

# ==== C·∫•u h√¨nh SharePoint ====
site_url = f"{SP_HOST}/sites/TESTLAB-VFR9"

# ================= PH·∫¶N 1: XU·∫§T FILE DS S·∫¢N PH·∫®M TEST V·ªöI QR =================
relative_url = "/sites/TESTLAB-VFR9/Shared Documents/DATA DAILY/QAD-Outstanding list-2025.xlsx"
excel_file_out = "ds san pham test voi qr.xlsx"

# --- K·∫æT N·ªêI SharePoint b·∫±ng Delegated (Device Code + Cache) ---
ctx = get_ctx(site_url)

download = io.BytesIO()
ctx.web.get_file_by_server_relative_url(relative_url).download(download).execute_query()
download.seek(0)

excel_file = pd.ExcelFile(download)
sheet_name = next((n for n in excel_file.sheet_names if n.strip().lower() == "ol list"), None)
if not sheet_name:
    raise Exception("Khong tim thay sheet 'OL list'!")
df = excel_file.parse(sheet_name)

def norm(s):
    return (str(s).strip().lower() if pd.notnull(s) else "")

def find_col_exact(df, name):
    """T√¨m c·ªôt kh·ªõp ch√≠nh x√°c theo normalize ('test date', 'complete date', ...)"""
    want = normalize_col(name)
    for c in df.columns:
        if normalize_col(c) == want:
            return c
    return None

def find_col_any(df, keywords):
    """T√¨m c·ªôt ch·ª©a t·∫•t c·∫£ keywords (chu·ªói con, kh√¥ng ph√¢n bi·ªát hoa/th∆∞·ªùng)."""
    for col in df.columns:
        low = str(col).lower()
        if all(k in low for k in keywords):
            return col
    for col in df.columns:
        low = str(col).lower()
        if any(k in low for k in keywords):
            return col
    return None

# C·ªôt ch·ªçn t·ª´ ngu·ªìn (gi·ªØ logic c≈©, nh∆∞ng s·∫Ω ƒë·∫£m b·∫£o th√™m Test Date & Complete Date n·∫øu c√≥)
cols_selected = list(df.columns[1:23])

# ƒê·∫£m b·∫£o c√≥ c·ªôt Rating (n·∫øu c√≥ tr√™n ngu·ªìn)
rating_col = next((c for c in df.columns if "rating" in str(c).lower()), None)
if rating_col and rating_col not in cols_selected:
    cols_selected.append(rating_col)

# ƒê·∫£m b·∫£o c√≥ c·ªôt 'Test Date' v√† 'Complete Date' t·ª´ ngu·ªìn (n·∫øu t·ªìn t·∫°i)
test_date_col = find_col_exact(df, "Test Date") or find_col_any(df, ["test", "date"])
complete_date_col = find_col_exact(df, "Complete Date") or find_col_any(df, ["complete", "date"])

if test_date_col and test_date_col not in cols_selected:
    cols_selected.append(test_date_col)
if complete_date_col and complete_date_col not in cols_selected:
    cols_selected.append(complete_date_col)

# X√°c ƒë·ªãnh c√°c c·ªôt ƒë·ªÉ l·ªçc OUTSOURCE
type_of_cols = [c for c in df.columns if "type of" in str(c).lower()]

# T√¨m c√°c c·ªôt 'status' v√† 'report' (gi·ªØ nguy√™n c√°ch t√¨m nh∆∞ c≈©)
def find_col(keywords):
    for col in df.columns:
        if all(k in str(col).lower() for k in keywords):
            return col
    for col in df.columns:
        if any(k in str(col).lower() for k in keywords):
            return col
    return None

status_col = find_col(["status"])
report_col = find_col(["report", "#"])

rows = list(df.index)

# df_out s·∫Ω gi·ªØ nguy√™n d·ªØ li·ªáu ngu·ªìn cho Test Date & Complete Date n·∫øu c√≥
df_out = df.iloc[rows].copy()[cols_selected]

# C·ªôt report ch√≠nh l√† c·ªôt th·ª© 2 c·ªßa df_out (gi·ªØ nguy√™n logic c≈©)
report_col_main = df_out.columns[1]
df_out[report_col_main] = df_out[report_col_main].astype(str).str.strip()
df_out = df_out.drop_duplicates(subset=report_col_main, keep='last')

# Ch·ªâ gi·ªØ c√°c report >= 4500 (nh∆∞ c≈©)
df_out["__report_num__"] = pd.to_numeric(
    df_out[report_col_main].astype(str).str.extract(r'(\d+)$')[0],
    errors="coerce"
)
df_out = df_out[df_out["__report_num__"] >= 0000].drop(columns="__report_num__")

# KH√îNG g√°n r·ªóng 2 c·ªôt ng√†y n·ªØa (b·ªè 2 d√≤ng c≈©):
# for col in ["Test Date", "Complete Date"]:
#     df_out[col] = ""

# T·∫°o QR code URL d·ª±a tr√™n report
qr_url_dict = {}
for _, row in df_out.iterrows():
    report_raw = str(row[report_col_main]).strip()
    if report_raw and report_raw.lower() != "nan":
        qr_url_dict[report_raw] = f"http://103.77.166.187:8246/update?report={report_raw}"
df_out["QR Code"] = df_out[report_col_main].astype(str).map(qr_url_dict).fillna("")

# Chu·∫©n h√≥a hi·ªÉn th·ªã c√°c c·ªôt ng√†y (bao g·ªìm c·∫£ Test Date & Complete Date n·∫øu c√≥)
def only_date(val):
    if pd.isnull(val) or not str(val).strip():
        return ""
    try:
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        s = str(val)
        if "-" in s and len(s.split(" ")[0]) == 10:
            return datetime.strptime(s.split(" ")[0], "%Y-%m-%d").strftime("%d/%m/%Y")
        if "/" in s and len(s.split(" ")[0]) == 10:
            return datetime.strptime(s.split(" ")[0], "%d/%m/%Y")
        if len(s) >= 10:
            return s[:10]
        return s
    except:
        return val

# Gom c√°c c·ªôt ng√†y c·∫ßn format: log in date, etd, test date, complete date (n·∫øu t·ªìn t·∫°i)
date_cols = [c for c in df_out.columns if "log in date" in str(c).lower() or "etd" in str(c).lower()]
if test_date_col:
    date_cols.append(test_date_col)
if complete_date_col:
    date_cols.append(complete_date_col)

# Lo·∫°i tr√πng (n·∫øu c√≥) + √°p d·ª•ng format
seen = set()
unique_date_cols = []
for c in date_cols:
    if c not in seen:
        seen.add(c)
        unique_date_cols.append(c)

for c in unique_date_cols:
    df_out[c] = df_out[c].apply(only_date)

# Merge v·ªõi file local n·∫øu ƒë√£ t·ªìn t·∫°i (gi·ªØ nguy√™n rating c·ªßa file local)
if os.path.exists(excel_file_out):
    df_exist = pd.read_excel(excel_file_out)
    df_exist.columns = [c.strip() for c in df_exist.columns]
    df_out.columns = [c.strip() for c in df_out.columns]

    report_col_exist = next((c for c in df_exist.columns if "report" in c.lower() and ("#" in c or "id" in c or c.lower().strip()=="report")), None)
    rating_col_exist = next((c for c in df_exist.columns if "rating" in c.lower()), None)

    report_col_main = df_out.columns[1]
    df_exist[report_col_exist] = df_exist[report_col_exist].astype(str).str.strip()
    df_out[report_col_main] = df_out[report_col_main].astype(str).str.strip()
    df_out = df_out.drop_duplicates(subset=report_col_main, keep='last')

    sharepoint_data = df_out.set_index(report_col_main).to_dict(orient="index")

    def is_valid_rating(val):
        if pd.isnull(val): return False
        sval = str(val).strip().lower()
        return sval != "" and sval != "nan"

    updated = []
    for _, row in df_exist.iterrows():
        report_val = row[report_col_exist]
        rating_val = row[rating_col_exist] if rating_col_exist in df_exist.columns else None
        if rating_col_exist and is_valid_rating(rating_val):
            # Gi·ªØ nguy√™n d√≤ng c√≥ rating h·ª£p l·ªá
            updated.append(row)
        else:
            key = str(report_val).strip()
            if pd.notnull(report_val) and key in sharepoint_data:
                new_data = sharepoint_data[key]
                # C·∫≠p nh·∫≠t to√†n b·ªô c·ªôt kh·ªõp t√™n (bao g·ªìm Test Date & Complete Date n·∫øu c√≥)
                for col in df_exist.columns:
                    if (not rating_col_exist) or (col != rating_col_exist):
                        if col in new_data:
                            row[col] = new_data[col]
            updated.append(row)
    df_final = pd.DataFrame(updated, columns=df_exist.columns)
else:
    df_final = df_out

df_final.to_excel(excel_file_out, index=False)

# ======= ƒê·ªäNH D·∫†NG, T√î M√ÄU =======
wb = load_workbook(excel_file_out)
ws = wb.active

thin = Side(border_style="thin", color="888888")
fill_late = PatternFill("solid", fgColor="FF6961")
fill_due = PatternFill("solid", fgColor="FFEB7A")
fill_must = PatternFill("solid", fgColor="FFA54F")
fill_complete = PatternFill("solid", fgColor="D3D3D3")
header_fill = PatternFill("solid", fgColor="B7E1CD")

header = [cell.value for cell in ws[1]]
# status_col c√≥ th·ªÉ None n·∫øu kh√¥ng t√¨m th·∫•y
status_col_idx = None
# N·∫øu b·∫°n mu·ªën c·ªë ƒë·ªãnh t√™n c·ªôt, c√≥ th·ªÉ set status_col = "Status"
# ·ªû ƒë√¢y m√¨nh gi·ªØ nguy√™n: n·∫øu c√≥ bi·∫øn status_col ·ªü find_col ph√≠a tr√™n th√¨ d√πng
try:
    if 'status_col' in locals() and status_col:
        for i, name in enumerate(header):
            if name and str(name).strip().lower() == status_col.strip().lower():
                status_col_idx = i + 1
                break
except Exception:
    status_col_idx = None

valid_statuses = ("active", "pending", "late", "due", "must", "complete")
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        if cell.row == 1:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 90
    if status_col_idx:
        status_val_norm = str(ws.cell(row=row, column=status_col_idx).value or "").strip().lower()
        fill = None
        if status_val_norm == "late":
            fill = fill_late
        elif status_val_norm == "due":
            fill = fill_due
        elif status_val_norm == "must":
            fill = fill_must
        elif status_val_norm == "complete":
            fill = fill_complete
        if status_val_norm and status_val_norm not in valid_statuses:
            ws.row_dimensions[row].hidden = True
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill

wb.save(excel_file_out)
print("Da xuat file:", excel_file_out)

# ================= PH·∫¶N 2: G·ªòP Completed + TRF =================
output_file = "TRF_complete.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # Completed
    if os.path.exists("completed_items.xlsx"):
        df_cpl = pd.read_excel("completed_items.xlsx", dtype=str)
        df_cpl.to_excel(writer, sheet_name="Completed", index=False)
    else:
        print("‚ö†Ô∏è Kh√¥ng t√¨m th·∫•y completed_items.xlsx")
    # TRF
    if os.path.exists("TRF.xlsx"):
        df_trf = pd.read_excel("TRF.xlsx", dtype=str)
        df_trf.to_excel(writer, sheet_name="TRF", index=False)
    else:
        print("‚ö†Ô∏è Kh√¥ng t√¨m th·∫•y TRF.xlsx")

# ===== ƒê·ªãnh d·∫°ng sau khi ghi =====
wb = load_workbook(output_file)
thin = Side(border_style="thin", color="888888")
header_fill = PatternFill("solid", fgColor="B7E1CD")

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length+2, 15)
    # format date
    format_date_columns(ws, target_headers=("log in date","etd"), number_format="DD-MMM")
    # hide rows
    df_tmp = pd.DataFrame(ws.values)
    login_col = find_login_date_col(df_tmp)
    if login_col:
        hide_rows_by_login_date(ws, login_col, today=datetime.now().date())

wb.save(output_file)
print("‚úÖ ƒê√£ xu·∫•t file:", output_file)

# ===== Upload SharePoint =====
upload_relative_url = "/sites/TESTLAB-VFR9/Shared Documents/DATA DAILY/TRF_complete.xlsx"
folder_excel = os.path.dirname(upload_relative_url)
ensure_folder(ctx, folder_excel)
with open(output_file, "rb") as f:
    ctx.web.get_folder_by_server_relative_url(folder_excel)\
       .upload_file(os.path.basename(upload_relative_url), f.read()).execute_query()
print("üöÄ ƒê√£ upload file", output_file, "l√™n SharePoint:", upload_relative_url)