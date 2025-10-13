import pandas as pd
import re
import os
import datetime
from config import local_main
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO

# =========================
# Helpers chuẩn hoá & utils
# =========================

def normalize_colname(s):
    return re.sub(r'[\s/\.:\-\n_]+', '', str(s).strip().lower())

def _norm_str(s):
    """Chuẩn hoá chuỗi để so khớp: bỏ NBSP/zero-width/CR/LF/TAB, strip, UPPER."""
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    return (
        s.replace("\u00A0", "")  # NBSP
         .replace("\u200B", "")  # zero-width space
         .replace("\r", "")
         .replace("\n", "")
         .replace("\t", "")
         .strip()
         .upper()
    )

def _as_int_like(s):
    """Rút toàn bộ chữ số để so khớp dạng số (ví dụ 25-5364 -> 255364)."""
    if s is None:
        return None
    digits = re.sub(r'[^0-9]', '', str(s))
    if digits == "":
        return None
    try:
        return int(digits)
    except Exception:
        return None

def clean_col(s):
    s = str(s).lower()
    s = re.sub(r'[^a-z0-9#]+', '', s)
    return s

# =========================
# Các hàm giữ nguyên tên
# =========================

def get_col_idx(ws, target):
    norm_target = normalize_colname(target)
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v and normalize_colname(v) == norm_target:
            return col
    return None

def ensure_column(ws, col_name):
    header = [cell.value for cell in ws[1]]
    if col_name not in header:
        ws.cell(row=1, column=ws.max_column+1, value=col_name)
    return [cell.value for cell in ws[1]].index(col_name) + 1

def copy_row_with_style(ws_from, ws_to, row_idx, to_row=None):
    if to_row is None:
        to_row = ws_to.max_row + 1
    for col in range(1, ws_from.max_column+1):
        c1 = ws_from.cell(row=row_idx, column=col)
        c2 = ws_to.cell(row=to_row, column=col)
        c2.value = c1.value
        try:
            c2.font = c1.font.copy() if c1.font else None
            c2.border = c1.border.copy() if c1.border else None
            c2.fill = c1.fill.copy() if c1.fill else None
            c2.number_format = c1.number_format
            c2.protection = c1.protection.copy() if c1.protection else None
            c2.alignment = c1.alignment.copy() if c1.alignment else None
        except Exception:
            pass

def is_img_at_cell(img, row, col):
    try:
        anchor = getattr(img, "anchor", None)
        if hasattr(anchor, "_from"):
            return (anchor._from.row + 1 == row) and (anchor._from.col + 1 == col)
        if hasattr(anchor, "cell"):
            img_row, img_col = anchor.cell.row + 1, anchor.cell.col_idx
            return (img_row == row) and (img_col == col)
    except Exception:
        return False
    return False

# ==============
# QAD dataframe
# ==============
try:
    QAD_DF = pd.read_excel(local_main)
    QAD_DF.columns = [clean_col(c) for c in QAD_DF.columns]
except Exception:
    QAD_DF = pd.DataFrame()

def get_item_code(report):
    row = QAD_DF.loc[QAD_DF['report#'] == str(report)] if 'report#' in QAD_DF.columns else pd.DataFrame()
    if not row.empty and 'item#' in QAD_DF.columns:
        return str(row.iloc[0]['item#'])
    return ""

# ==============================
# Tìm đúng cột/row theo REPORT NO
# ==============================

def _find_report_col(ws):
    """
    Ưu tiên nhận các biến thể tiêu đề: 'Report #', 'Report#', 'Report No', 'Report Number', 'Report'
    Fallback: 2 (cột B).
    """
    import re
    def norm(s):
        return re.sub(r'[^a-z0-9#]+', '', str(s).strip().lower())

    candidates = {"report#", "reportno", "reportnumber", "report"}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if not name:
            continue
        n = norm(name)
        # match chặt cho các biến thể phổ biến hoặc match mềm: có 'report' và ('no' hoặc '#')
        if (
            n in candidates
            or n.startswith("report")
            or ("report" in n and ("no" in n or "#" in n))
        ):
            return col
    return 2  # fallback đúng cột B

def _find_row_by_report(ws, report_no, report_col=None):
    """Tìm row có giá trị cột REPORT NO == report_no (so khớp mạnh)."""
    if report_col is None:
        report_col = _find_report_col(ws)

    tgt_txt = _norm_str(report_no)
    tgt_num = _as_int_like(report_no)

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=report_col).value
        if v is None:
            continue
        if _norm_str(v) == tgt_txt:
            return r
        v_num = _as_int_like(v)
        if tgt_num is not None and v_num is not None and v_num == tgt_num:
            return r
    return None

def _build_headers_map(ws):
    """Tạo map {header_clean: col_idx} để match mềm theo từ khoá."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if name:
            clean = (
                str(name)
                .strip()
                .replace('\n', ' ')
                .replace('/', ' ')
                .replace('.', '')
                .replace('#', '')
                .lower()
            )
            clean = " ".join(clean.split())
            headers[clean] = col
    return headers

def _set_by_keywords(ws, row_idx, headers, keywords, value):
    """
    Ghi value vào cột có header thỏa mãn tất cả từ khóa (match mềm, lowercase, bỏ dấu câu).
    keywords: list[str], ví dụ ["estimated", "completion", "date"]
    """
    if value is None or value == "":
        return False
    kws = [w.lower() for w in keywords]
    for h_clean, col_idx in headers.items():
        if all(word in h_clean for word in kws):
            ws.cell(row=row_idx, column=col_idx).value = value
            return True
    return False

# ===== Helpers ghi NGÀY với number_format dd-mmm (thêm mới) =====

def _to_excel_date(value):
    """Chuẩn hoá giá trị ngày -> datetime (openpyxl) để Excel nhận dạng ngày.
    Hỗ trợ: datetime/date, 'YYYY-MM-DD', 'DD/MM/YYYY', 'DD-MM-YYYY', ..."""
    if value is None or value == "":
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
            return datetime.datetime(value.year, value.month, value.day)
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    s = str(value).strip()
    try:
        try:
            dt = pd.to_datetime(s, format="%Y-%m-%d", errors='raise')
        except Exception:
            dt = pd.to_datetime(s, dayfirst=True, errors='raise')
        if pd.isna(dt):
            return None
        if hasattr(dt, 'to_pydatetime'):
            dt = dt.to_pydatetime()
        return dt.replace(hour=0, minute=0, second=0, microsecond=0)
    except Exception:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
            try:
                return datetime.datetime.strptime(s, fmt)
            except Exception:
                pass
    return None

def _set_date_by_keywords(ws, row_idx, headers, keywords, value, fmt='dd-mmm'):
    """Ghi ngày vào cột match keywords và set number_format theo fmt (mặc định dd-mmm)."""
    dt = _to_excel_date(value)
    if dt is None:
        return _set_by_keywords(ws, row_idx, headers, keywords, value)
    kws = [w.lower() for w in keywords]
    for h_clean, col_idx in headers.items():
        if all(word in h_clean for word in kws):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = dt
            cell.number_format = fmt
            return True
    return False

# =========================
# GHI DỮ LIỆU VÀO ĐÚNG HÀNG
# =========================

def write_tfr_to_excel(excel_path, report_no, request):
    """
    Ghi dữ liệu vào **ĐÚNG HÀNG** có REPORT NO == report_no.
    - KHÔNG tự +1.
    - Tìm cột bằng header 'report' & 'no' (fallback B).
    - Ghi theo 'match mềm' header (từ khoá).
    - ĐẶC BIỆT:
        * "Log in date" = request["log_in_date"] hoặc request["request_date"] (định dạng dd-mmm)
        * ETD điền vào cột "ETD"/"Estimated Completion Date"/"Estimated Completed Date" (tùy header)
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 1) Tìm cột REPORT NO và hàng tương ứng
    report_col = _find_report_col(ws)
    row_idx = _find_row_by_report(ws, report_no, report_col)
    if row_idx is None:
        wb.close()
        raise Exception(f"Không tìm thấy mã report {report_no} trong file excel!")

    # 2) Chuẩn bị headers map
    headers = _build_headers_map(ws)

    def to_upper(val):
        return val.upper() if isinstance(val, str) else val

    # 'type of' tách từ test_group (bỏ ' TEST' nếu có)
    test_group_val = request.get("test_group", "")
    if isinstance(test_group_val, str) and test_group_val.upper().endswith(" TEST"):
        type_of_val = test_group_val[:-5].strip()
    else:
        type_of_val = test_group_val

    # 3) Ghi TRQ-ID nếu có cột tương ứng
    trq_col = (
        get_col_idx(ws, "trq-id")
        or get_col_idx(ws, "trq_id")
        or get_col_idx(ws, "trq id")
        or get_col_idx(ws, "trqid")
    )
    if trq_col:
        ws.cell(row=row_idx, column=trq_col).value = request.get("trq_id", "")

    # 4) Ghi các trường mô tả chính theo match mềm
    fields_map = [
        (["item"], to_upper(request.get("item_code", ""))),
        (["type of"], to_upper(type_of_val)),
        (["item name", "description"], to_upper(request.get("sample_description", ""))),
        (["furniture testing"], to_upper(request.get("furniture_testing", ""))),
        (["submitter in", "submitter in charge"], to_upper(request.get("requestor", ""))),
        (["submitted dept"], to_upper(request.get("department", ""))),
        (["remark"], to_upper(request.get("test_status", ""))),
    ]
    for keys, val in fields_map:
        _set_by_keywords(ws, row_idx, headers, keys, val)

    # 5) Các cột tuỳ chọn: priority, ETD/Estimated, QR link (Y=25) nếu có
    # 5.1 Priority (nếu có trong request)
    if "priority" in request:
        _set_by_keywords(ws, row_idx, headers, ["priority"], to_upper(request.get("priority")))

    # 5.2 ETD / Estimated Completion/Completed Date
    etd_val = request.get("etd") or request.get("estimated_completion_date")
    wrote_etd = (
        _set_by_keywords(ws, row_idx, headers, ["etd"], etd_val) or
        _set_by_keywords(ws, row_idx, headers, ["estimated", "completion", "date"], etd_val) or
        _set_by_keywords(ws, row_idx, headers, ["estimated", "completed", "date"], etd_val)
    )

    # 5.3 Log in date = request["log_in_date"] hoặc request["request_date"] => ghi với number_format dd-mmm
    login_date_val = request.get("log_in_date") or request.get("request_date")
    _set_date_by_keywords(ws, row_idx, headers, ["log", "in", "date"], login_date_val, fmt='dd-mmm') or \
        _set_date_by_keywords(ws, row_idx, headers, ["login", "date"], login_date_val, fmt='dd-mmm')

    # 5.4 QR link (cột Y=25) nếu backend có set
    if "qr_link" in request:
        ws.cell(row=row_idx, column=25).value = str(request["qr_link"])

    wb.save(excel_path)
    wb.close()

def append_row_to_trf(report_no, main_excel_path, trf_excel_path, trq_id=None):
    wb_main = load_workbook(main_excel_path, data_only=True)
    ws_main = wb_main.active

    # Tìm dòng theo REPORT NO (so khớp mạnh)
    report_col = _find_report_col(ws_main)
    row_idx = _find_row_by_report(ws_main, report_no, report_col)
    if row_idx is None:
        wb_main.close()
        return

    # Nếu chưa có file TRF thì tạo mới, copy FULL header layout
    if not os.path.exists(trf_excel_path):
        wb_trf_new = Workbook()
        ws_trf_new = wb_trf_new.active
        # Copy header
        for col in range(1, ws_main.max_column + 1):
            c1 = ws_main.cell(row=1, column=col)
            c2 = ws_trf_new.cell(row=1, column=col)
            c2.value = c1.value
            if c1.has_style:
                c2.font = copy(c1.font)
                c2.border = copy(c1.border)
                c2.fill = copy(c1.fill)
                c2.number_format = c1.number_format
                c2.protection = copy(c1.protection)
                c2.alignment = copy(c1.alignment)
            col_letter = get_column_letter(col)
            ws_trf_new.column_dimensions[col_letter].width = ws_main.column_dimensions[col_letter].width
        ws_trf_new.row_dimensions[1].height = ws_main.row_dimensions[1].height
        wb_trf_new.save(trf_excel_path)

    # Mở và append row
    wb_trf = load_workbook(trf_excel_path)
    ws_trf = wb_trf.active

    to_row = ws_trf.max_row + 1
    for col in range(1, ws_main.max_column + 1):
        c1 = ws_main.cell(row=row_idx, column=col)
        c2 = ws_trf.cell(row=to_row, column=col)
        c2.value = c1.value
        if c1.has_style:
            c2.font = copy(c1.font)
            c2.border = copy(c1.border)
            c2.fill = copy(c1.fill)
            c2.number_format = c1.number_format
            c2.protection = copy(c1.protection)
            c2.alignment = copy(c1.alignment)

    # Copy chiều cao & width cho chắc
    ws_trf.row_dimensions[to_row].height = ws_main.row_dimensions[row_idx].height
    for col in range(1, ws_main.max_column + 1):
        col_letter = get_column_letter(col)
        ws_trf.column_dimensions[col_letter].width = ws_main.column_dimensions[col_letter].width

    wb_trf.save(trf_excel_path)
    wb_trf.close()
    wb_main.close()

def export_expired_samples_to_excel(rows):
    """Tạo file Excel 3 cột: Report - Item - Loại mẫu (cho các mẫu hết hạn)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mẫu hết hạn"

    headers = ["Report", "Item", "Loại mẫu"]
    ws.append(headers)

    # Ghi dữ liệu
    for r in rows:
        ws.append([r.get('report',''), r.get('item_code',''), r.get('sample_type','')])

    # Style
    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='999999')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    fill = PatternFill('solid', fgColor='FFF2CC')

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = fill

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4
        for c in col:
            c.alignment = center
            c.border = border

    # Freeze header + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()